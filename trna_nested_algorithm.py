"""
Nested Algorithm for tRNA Mixtures Using "Block-Wise" Relative Intensity
=========================================================================

Implementation of Dr. Jianfeng Jiang's method for de novo base-calling of
tRNA (or other oligonucleotide) mixtures from deconvoluted LC-MS data.

Pipeline (mirrors the steps in the source document):

    Step 1  Load (M, I, T) = (monoisotopic mass, summed intensity, apex RT)
            from a deconvoluted Excel file.
    Step 2  Sort all data points by mass M, ascending.
    Step 3  Assign each point to a "block" k (k = 1..95), where block k
            covers M in ((k-0.5)*320, (k+0.5)*320) Da. 320 Da ~= the
            average single-nucleotide residue mass, so each block roughly
            corresponds to one ladder position.
    Step 4  Within each block, sort by intensity I (descending) and convert
            absolute intensity to *relative* intensity (Rel_I), using the
            most intense point in the block as the base (Rel_I = 1).
    Step 5  Visualize Rel_I vs. M (2-D scatter).
    Step 6  "Base calling": greedily walk a mass ladder, picking the next
            fragment by (1) an allowed nucleotide mass difference, (2) the
            highest Rel_I among mass-valid candidates, and (3) consistency
            with the chain's running retention-time (RT) trend. Repeat from
            the next-highest unused intensity peak to recover multiple short
            reads, classify each as a 5'- or 3'-ladder, and (optionally)
            align reads against a reference sequence to flag modifications
            or isoforms.

Because steps 6.4-6.7 describe judgment calls a person makes by eye, this
module turns them into explicit, *tunable* heuristics (see `Config`) rather
than claiming one "correct" automatic answer. Every threshold is a keyword
argument so you can adjust it to your own LC-MS method and nucleotide
chemistry.

Quick start
-----------
    python trna_nested_algorithm.py --input my_deconvoluted_data.xlsx
    python trna_nested_algorithm.py --demo          # runs on the worked
                                                      # example from the
                                                      # source document

Requires: pandas, numpy, matplotlib, openpyxl (for reading .xlsx).
"""

from __future__ import annotations

import argparse
import bisect
import itertools
import json
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")  # headless / file-output friendly
import matplotlib.pyplot as plt


# ---------------------------------------------------------------------------
# Reference data: monoisotopic *residue* masses (Da), i.e. the mass added to
# a chain when extending it by one nucleotide (nucleoside-5'-monophosphate
# minus water). Extend/replace this table for your own modification set.
# ---------------------------------------------------------------------------

CANONICAL_RNA_RESIDUE_MASS: Dict[str, float] = {
    "A": 329.0525,
    "C": 305.0413,
    "G": 345.0474,
    "U": 306.0253,
}

# A few common tRNA modifications, given as illustrative starting points.
# Real modification sets are organism/sample specific -- edit freely.
COMMON_MODIFIED_RESIDUE_MASS: Dict[str, float] = {
    "Y":    306.0253,   # pseudouridine - isobaric with U
    "D":    308.0410,   # dihydrouridine (U + 2H)
    "m1A":  343.0682,   # 1-methyladenosine (A + CH2)
    "m5C":  319.0569,   # 5-methylcytidine (C + CH2)
    "m7G":  359.0631,   # 7-methylguanosine (G + CH2)
    "m2G":  359.0631,   # N2-methylguanosine (isobaric with m7G)
    "I":    330.0365,   # inosine (A, deaminated)
    "ac4C": 347.0518,   # N4-acetylcytidine (C + C2H2O)
}

DEFAULT_ALLOWED_MASSES: Dict[str, float] = {
    **CANONICAL_RNA_RESIDUE_MASS,
    **COMMON_MODIFIED_RESIDUE_MASS,
}

BLOCK_WIDTH_DA = 320.0   # Step 3: ~ average single-nucleotide residue mass
N_BLOCKS = 95            # Step 3: k = 1..95


# ---------------------------------------------------------------------------
# Configuration for the heuristic parts of Step 6
# ---------------------------------------------------------------------------

@dataclass
class Config:
    block_width: float = BLOCK_WIDTH_DA
    n_blocks: int = N_BLOCKS

    # Step 4 (display only): which intensity column the *plots* read.
    # `compute_relative_intensity` always computes BOTH the doc's literal
    # linear Rel_I (Rel_I = I / max(I in block)) and a log10-normalized
    # alternative (log10_intensity_qc), and primary read/selection logic
    # always uses the linear column, full stop -- this setting cannot
    # change that. It only changes which column `plot_relative_intensity` /
    # `plot_chains_overlay` draw, as a QC/visualization convenience. The
    # "_qc" suffix is deliberate: this column exists to sanity-check the
    # data by eye, never to decide which peak gets picked.
    intensity_scale: str = "linear"

    # Step 6.2, criterion 1: how close (Da) a mass difference must be to a
    # table value to count as a match.
    mass_tol: float = 0.02

    # Step 6.2, criterion 1, gap handling -- two clearly separate numbers so
    # "residues per jump" and "skipped rungs" are never conflated:
    #   max_residues_per_step   = how many residues may be summed together
    #                             to test one mass-difference jump. This is
    #                             the number you set directly.
    #   max_missing_positions   = max_residues_per_step - 1 (derived, below)
    #                             = how many ladder rungs that jump skips.
    # A jump of max_residues_per_step=3 skips 2 rungs -- never describe that
    # as "a gap of 3" without saying which of the two numbers you mean.
    # Jiang-strict default (3) is the conservative, always-on setting. Turn
    # on `exploratory_mode` to opt into the wider, noisier-data-tolerant cap
    # (exploratory_max_residues_per_step) -- off by default.
    max_residues_per_step: int = 3
    exploratory_mode: bool = False
    exploratory_max_residues_per_step: int = 4

    # Step 6.2, criterion 3: a candidate's RT must fall within
    # `rt_zscore` * std of the chain's running RT-delta trend to be
    # accepted. Only enforced once >= min_rt_history points exist.
    rt_zscore: float = 2.5
    min_rt_history: int = 2
    rt_window: int = 4          # how many trailing steps define the trend
    rt_std_floor: float = 0.02  # minutes; avoids a zero-width window

    # Step 6.3 / 6.5: minimum length (# of points) for a walk to be kept as
    # a "read" rather than discarded as an isolated, unextendable peak.
    min_chain_len: int = 2

    # Step 6.4: minimum composite pairing score (0..1, see
    # `_pairwise_evidence`) for two reads to be accepted as a 5'/3' pair.
    # Below this, each read is treated as unpaired and falls back to the
    # global-median heuristic (always labeled as a fallback in its
    # evidence). Calibrate against real, Jiang-reviewed data.
    min_pair_score: float = 0.45

    # Step 6.4, optional: intact/precursor mass (Da) of the full-length
    # molecule, if known -- used as one extra piece of pairing evidence
    # (does the sum of the two reads' terminal masses close to the
    # precursor mass?). Purely optional; pairing works without it.
    precursor_mass: Optional[float] = None
    precursor_mass_tol: float = 1.0

    # Step 6.6: an optional reference sequence (single-letter / mod codes).
    # When set it (a) is used post-hoc to flag mismatches as candidate
    # modifications/isoforms (`compare_to_reference`), AND (b) is used as
    # SECOND-STAGE evidence, strictly after all primary Step 6.1-6.5 reads
    # already exist, to extend a read's ends by reusing an already-flagged
    # point -- never to override or interleave with primary read
    # generation (see `apply_reference_guided_extension`). Every reused
    # point gets an explicit evidence record. With reference=None, no
    # reuse happens and behavior is unaffected.
    reference: Optional[Sequence[str]] = None

    allowed_masses: Dict[str, float] = field(
        default_factory=lambda: dict(DEFAULT_ALLOWED_MASSES)
    )

    @property
    def effective_max_residues_per_step(self) -> int:
        """The residues-per-jump cap actually in force: the Jiang-strict
        default unless `exploratory_mode` has been explicitly turned on."""
        return self.exploratory_max_residues_per_step if self.exploratory_mode else self.max_residues_per_step

    @property
    def max_missing_positions(self) -> int:
        """Derived: skipped ladder rungs = residues per jump - 1 (the rung
        actually landed on doesn't count as "missing")."""
        return self.effective_max_residues_per_step - 1


# ---------------------------------------------------------------------------
# Step 1 & 2: load data, sort by mass
# ---------------------------------------------------------------------------

def _resolve_column(df: pd.DataFrame, wanted: str, keyword_sets: Sequence[Sequence[str]]) -> str:
    """Find a column matching `wanted` exactly, else fuzzy-match on keywords.

    `keyword_sets` is tried in order from most to least specific (e.g. first
    look for a column containing both "mono" and "mass" before settling for
    any column that merely contains "mass") so a decoy column like "Average
    Mass" doesn't get picked ahead of "Monoisotopic Mass".
    """
    if wanted in df.columns:
        return wanted
    lowered = {c.lower(): c for c in df.columns}
    if wanted.lower() in lowered:
        return lowered[wanted.lower()]
    for must_contain in keyword_sets:
        for c in df.columns:
            cl = c.lower()
            if all(kw in cl for kw in must_contain):
                return c
    raise KeyError(
        f"Could not find a column matching '{wanted}' (looked for columns "
        f"containing any of {keyword_sets}). Available columns: {list(df.columns)}"
    )


def load_data(
    path: str,
    mass_col: str = "Monoisotopic Mass",
    intensity_col: str = "Sum Intensity",
    rt_col: str = "Apex RT",
    sheet_name=0,
) -> pd.DataFrame:
    """Step 1: read (M, I, T) from a deconvoluted Excel export.
    Step 2: sort by M ascending.
    """
    raw = pd.read_excel(path, sheet_name=sheet_name)

    m_col = _resolve_column(raw, mass_col, [["mono", "mass"], ["monoisotopic"], ["mass"]])
    i_col = _resolve_column(raw, intensity_col, [["sum", "intensity"], ["intensity"], ["area"], ["abundance"]])
    t_col = _resolve_column(raw, rt_col, [["apex", "rt"], ["retention"], ["rt"]])

    df = raw[[m_col, i_col, t_col]].rename(columns={m_col: "M", i_col: "I", t_col: "T"})
    df = df.dropna(subset=["M", "I", "T"]).astype({"M": float, "I": float, "T": float})
    df = df.sort_values("M", ascending=True, ignore_index=True)  # Step 2
    return df


# ---------------------------------------------------------------------------
# Step 3: assign blocks
# ---------------------------------------------------------------------------

def assign_blocks(df: pd.DataFrame, block_width: float = BLOCK_WIDTH_DA, n_blocks: int = N_BLOCKS) -> pd.DataFrame:
    """Block k covers M in ((k-0.5)*width, (k+0.5)*width); k clipped to [1, n_blocks]."""
    out = df.copy()
    k = np.round(out["M"] / block_width).astype(int)
    out["block"] = k.clip(1, n_blocks)
    return out


# ---------------------------------------------------------------------------
# Step 4: relative intensity within each block
# ---------------------------------------------------------------------------

def compute_relative_intensity(df: pd.DataFrame) -> pd.DataFrame:
    """Within each block, sort by I descending and rescale to a 0..1 Rel_I.

    Rel_I is ALWAYS the doc's literal linear definition, Rel_I = I / max(I in
    block) -- this is the column every selection rule in Step 6.2 reads, and
    it must never silently swap to something else based on a display
    preference. A second column, log10_intensity_qc (min-max-normalized
    log10(I) within the block), is also always computed as a QC/plotting
    alternative for when block intensities span many orders of magnitude --
    it is read-only by plotting code (`Config.intensity_scale`) and is NEVER
    consulted by any base-calling/selection/pairing logic. The name says
    what it's for: visualization and QC, not a selection rule.
    """
    out = df.sort_values(["block", "I"], ascending=[True, False], ignore_index=True)
    out["Rel_I"] = out.groupby("block")["I"].transform(lambda s: s / s.max())

    log_i = np.log10(out["I"].clip(lower=1e-12))

    def _minmax(s: pd.Series) -> pd.Series:
        lo, hi = s.min(), s.max()
        if hi - lo < 1e-12:
            return pd.Series(1.0, index=s.index)
        return (s - lo) / (hi - lo)

    out["log10_intensity_qc"] = log_i.groupby(out["block"]).transform(_minmax)
    return out


# ---------------------------------------------------------------------------
# Step 5: visualize
# ---------------------------------------------------------------------------

def plot_relative_intensity(
    df: pd.DataFrame,
    out_path: str = "relative_intensity.png",
    title: str = "Relative I",
    y_col: str = "Rel_I",
    ylabel: str = "Relative Intensity",
) -> str:
    fig, ax = plt.subplots(figsize=(13, 6))
    ax.scatter(df["M"], df[y_col], facecolors="none", edgecolors="#1f6f8b", s=18, linewidths=0.8)
    ax.set_title(title)
    ax.set_xlabel("Monoisotopic Mass, M (Da)")
    ax.set_ylabel(ylabel)
    ax.set_ylim(-0.02, 1.05)
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def plot_chains_overlay(
    df: pd.DataFrame,
    chains: List[Dict],
    out_path: str = "chains_overlay.png",
    title: str = "Flagged points by chain (Step 6.5)",
    y_col: str = "Rel_I",
    ylabel: str = "Relative Intensity",
) -> str:
    """Step 6.5: color-code each recovered chain's flagged/used points so
    they're visually separated from each other, overlaid on the full Rel_I
    vs M scatter (mirrors the doc's example chart of colored ladder traces
    drawn over the raw background scatter).
    """
    fig, ax = plt.subplots(figsize=(13, 6))

    # Background: every point, faint, so unflagged/unused points stay visible.
    ax.scatter(df["M"], df[y_col], facecolors="none", edgecolors="#bbbbbb", s=14, linewidths=0.6, zorder=1)

    cmap = plt.get_cmap("tab20" if len(chains) > 10 else "tab10")
    for i, chain in enumerate(chains):
        rows = df.loc[chain["indices"]].sort_values("M")
        color = cmap(i % cmap.N)
        label = f"chain {i} ({chain.get('ladder_type', '?')})"
        ax.plot(rows["M"], rows[y_col], "-o", color=color, markersize=4, linewidth=1.4, label=label, zorder=2)
        reuse_idx = chain.get("_reused_peak_indices", [])
        if reuse_idx:
            r_rows = df.loc[reuse_idx]
            ax.scatter(r_rows["M"], r_rows[y_col], facecolors="none", edgecolors=color, s=90,
                       linewidths=1.6, marker="o", zorder=3)

    ax.set_title(title)
    ax.set_xlabel("Monoisotopic Mass, M (Da)")
    ax.set_ylabel(ylabel)
    ax.set_ylim(-0.02, 1.05)
    ax.grid(True, alpha=0.3)
    if 0 < len(chains) <= 20:
        ax.legend(loc="upper right", fontsize=7, ncol=2, framealpha=0.9)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


# ---------------------------------------------------------------------------
# Step 6: base calling
# ---------------------------------------------------------------------------

MassDiffEntry = Tuple[float, Tuple[str, ...]]  # (summed mass, (residue labels...))


class MassDiffTable:
    """Sorted-by-mass lookup of allowed residue-jump sums, built once per
    run. On real LC-MS data a single ladder step can have thousands of
    mass-valid candidate points within tolerance to check (see
    `_find_candidates`), so `match()` uses binary search (bisect) instead
    of scanning every one of the table's entries for every candidate --
    that scan-every-entry approach is what made the first real-workbook
    run unusably slow. Behaves like the plain list of (total, labels)
    tuples it replaces wherever that's still convenient (`len()`,
    iteration, `max_total`).
    """

    def __init__(self, entries: List[MassDiffEntry]):
        entries = sorted(entries, key=lambda e: e[0])
        self._entries = entries
        self._totals = [e[0] for e in entries]

    def match(self, delta: float, tol: float) -> List[MassDiffEntry]:
        target = abs(delta)
        lo = bisect.bisect_left(self._totals, target - tol)
        hi = bisect.bisect_right(self._totals, target + tol)
        return self._entries[lo:hi]

    @property
    def max_total(self) -> float:
        return self._totals[-1] if self._totals else 0.0

    def __len__(self) -> int:
        return len(self._entries)

    def __iter__(self):
        return iter(self._entries)


def build_mass_diff_table(allowed_masses: Dict[str, float], max_residues_per_step: int = 3) -> MassDiffTable:
    """All sums of 1..max_residues_per_step consecutive residues, used to
    recognize ladder steps that jump over one or more missing rungs.
    `max_residues_per_step` is the number of residues summed in one jump;
    the number of *missing rungs* that jump skips is one less than that
    (the rung actually landed on doesn't count as missing) -- see
    `Config.max_missing_positions`.
    """
    table: List[MassDiffEntry] = []
    items = list(allowed_masses.items())
    for n in range(1, max_residues_per_step + 1):
        for combo in itertools.product(items, repeat=n):
            labels = tuple(label for label, _ in combo)
            total = sum(mass for _, mass in combo)
            table.append((total, labels))
    return MassDiffTable(table)


def _match_mass_diff(delta: float, diff_table: MassDiffTable, tol: float) -> List[MassDiffEntry]:
    return diff_table.match(delta, tol)


def _find_candidates(
    df: pd.DataFrame,
    sorted_idx: np.ndarray,
    pos_of: Dict[int, int],
    current_idx: int,
    used: np.ndarray,
    diff_table: List[MassDiffEntry],
    tol: float,
    max_mass_diff: float,
) -> List[Tuple[int, List[MassDiffEntry]]]:
    """Return [(candidate_row_index, matches)] for rows whose mass differs
    from the current row by an allowed amount, in either direction.
    Uses the mass-sorted order to restrict the search window for speed.

    On real LC-MS data this window can hold thousands of candidates per
    call, so the inner loop reads `M` off a plain numpy array (`m_values`)
    instead of `df.at[idx, "M"]` -- repeated `.at[]` scalar lookups were
    the other major cost alongside the linear diff-table scan fixed in
    `MassDiffTable`. `m_values` is positional (`m_values[idx]`), which is
    safe here because `idx` values come from `sorted_idx`, itself built
    from `df.index`, and every index label this pipeline ever produces is
    a contiguous 0..len(df)-1 RangeIndex (see `load_data`/
    `compute_relative_intensity`, both of which sort with
    `ignore_index=True`).
    """
    m_values = df["M"].values
    used_values = used.values if hasattr(used, "values") else used
    cur_M = m_values[current_idx]

    lo = np.searchsorted(m_values, cur_M - max_mass_diff - tol, side="left")
    hi = np.searchsorted(m_values, cur_M + max_mass_diff + tol, side="right")

    out = []
    for pos in range(lo, hi):
        idx = sorted_idx[pos]
        if idx == current_idx or used_values[idx]:
            continue
        delta = m_values[idx] - cur_M
        matches = _match_mass_diff(delta, diff_table, tol)
        if matches:
            out.append((idx, matches))
    return out


def _select_next(
    df: pd.DataFrame,
    candidates: List[Tuple[int, List[MassDiffEntry]]],
    rt_history: List[float],
    cfg: Config,
) -> Optional[Tuple[int, List[MassDiffEntry]]]:
    """Step 6.2, criteria 2 & 3: rank mass-valid candidates by Rel_I, then
    accept the highest-intensity one whose RT still fits the chain's trend.
    """
    if not candidates:
        return None

    # Criterion 2: highest relative intensity first.
    candidates = sorted(candidates, key=lambda c: df.at[c[0], "Rel_I"], reverse=True)

    # Need >= 2 RT points to even compute a delta, regardless of how low
    # min_rt_history was set.
    if len(rt_history) < max(cfg.min_rt_history, 2):
        return candidates[0]

    window = rt_history[-cfg.rt_window:]
    deltas = np.diff(window)
    mean_d = float(deltas.mean())
    std_d = max(float(deltas.std()), cfg.rt_std_floor)
    predicted_rt = rt_history[-1] + mean_d

    # Criterion 3: walk candidates in intensity order, take the first whose
    # RT is consistent with the running trend.
    for idx, matches in candidates:
        if abs(df.at[idx, "T"] - predicted_rt) <= cfg.rt_zscore * std_d:
            return idx, matches
    return None  # nothing satisfies all three criteria -> chain ends (6.3)


def _flatten_labels(edge_labels: List[Tuple[str, ...]]) -> List[str]:
    """Expand a chain's per-edge label tuples (which may carry several
    residues at once, for gap jumps) into one flat ordered list of single
    residue codes, for reference alignment purposes.
    """
    flat: List[str] = []
    for labels in edge_labels:
        flat.extend(labels)
    return flat


def _best_reference_alignment(
    calls: Sequence[str], reference: Sequence[str]
) -> Optional[Tuple[Sequence[str], int, int]]:
    """Find the offset (and orientation) of `reference` that best matches
    `calls`, by brute-force scoring every offset against both the reference
    and its reverse (since we don't know a priori whether a given chain is
    reading the reference 5'->3' or 3'->5'). Returns
    (oriented_reference, offset, score), or None if either input is empty.
    """
    if not calls or not reference:
        return None
    best = None
    for oriented in (list(reference), list(reversed(reference))):
        n, m = len(oriented), len(calls)
        for offset in range(-m + 1, n):
            score = 0
            for i, call in enumerate(calls):
                j = offset + i
                if 0 <= j < n and oriented[j] == call:
                    score += 1
            if best is None or score > best[2]:
                best = (oriented, offset, score)
    return best


def _chain_ordered_calls(chain: Dict) -> List[str]:
    """Flatten a chain's edge labels into one ordered list of single-residue
    calls, left-to-right in the same order as chain['indices'] (already
    assembled low-mass-end-first by `build_chain`).
    """
    return _flatten_labels(chain["edge_labels"])


def _reference_guided_step(
    df: pd.DataFrame,
    sorted_idx: np.ndarray,
    used: pd.Series,
    diff_table: List[MassDiffEntry],
    cfg: Config,
    max_mass_diff: float,
    chains: List[Dict],
    ci: int,
    end: str,  # "append" (extend toward higher mass) or "prepend" (toward lower mass)
    audit_log: Optional[List[Dict]] = None,
) -> Optional[Tuple[int, List[MassDiffEntry], Dict]]:
    """Step 6.6, STAGE 2 only -- called strictly after Stage 1 (primary
    generation + Step 6.4 classification + Step 6.5 flag-and-exclude) is
    finished (see `apply_reference_guided_extension`). Realigns the read's
    current calls against the reference fresh each time (reads are short,
    this is cheap), and checks whether an already-flagged point elsewhere
    would extend this read to a residue the reference confidently predicts
    there. Returns (index, mass_diff_matches, evidence_dict) or None; the
    evidence dict documents exactly why the point was eligible for reuse, so
    nothing about a reuse decision is silent. No-op when `cfg.reference` is
    unset.

    `audit_log`, when given, gets exactly one entry appended per call for
    every outcome past the "is there even a reference to test against"
    guard -- accepted, or rejected with the specific reason -- so
    `apply_reference_guided_extension`'s caller can tell "reuse was tested
    and found nothing" apart from "reuse was never tested" (review-and-
    correction pass requirement 3).
    """
    if not cfg.reference:
        return None
    chain = chains[ci]
    ordered_calls = _chain_ordered_calls(chain)
    if not ordered_calls:
        return None
    read_rank = chain["read_rank"]

    def _reject(reason: str, **extra) -> None:
        if audit_log is not None:
            audit_log.append({
                "read_rank": read_rank, "end": end, "outcome": "rejected", "reason": reason, **extra
            })

    alignment = _best_reference_alignment(ordered_calls, cfg.reference)
    if alignment is None:
        _reject("no reference alignment could be computed for this read")
        return None
    oriented, offset, score = alignment
    min_score = max(2, 0.6 * len(ordered_calls))
    if score < min_score:
        _reject(f"reference alignment too weak to trust ({score}/{len(ordered_calls)} positions match, "
                f"need >= {min_score:.1f})")
        return None  # alignment too weak to trust for borrowing a point

    if end == "append":
        target_pos = offset + len(ordered_calls)
        cur = chain["indices"][-1]
        increasing = True
    else:
        target_pos = offset - 1
        cur = chain["indices"][0]
        increasing = False
    if not (0 <= target_pos < len(oriented)):
        _reject("reference position for the next residue falls outside the reference sequence bounds")
        return None
    expected_residue = oriented[target_pos]

    chain_idx_set = set(chain["indices"])
    cur_M = df.at[cur, "M"]
    lo = np.searchsorted(df["M"].values, cur_M - max_mass_diff - cfg.mass_tol, side="left")
    hi = np.searchsorted(df["M"].values, cur_M + max_mass_diff + cfg.mass_tol, side="right")

    best = None
    for pos in range(lo, hi):
        idx = sorted_idx[pos]
        if idx == cur or idx in chain_idx_set or not used[idx]:
            continue  # only consider points already flagged elsewhere, not new/unflagged ones
        if (df.at[idx, "M"] > cur_M) != increasing:
            continue
        delta = df.at[idx, "M"] - cur_M
        matches = [m for m in _match_mass_diff(delta, diff_table, cfg.mass_tol) if expected_residue in m[1]]
        if not matches:
            continue
        if best is None or df.at[idx, "Rel_I"] > df.at[best[0], "Rel_I"]:
            best = (idx, matches)
    if best is None:
        _reject(f"no already-flagged peak in the mass window matches the expected residue "
                f"({expected_residue}) at reference position {target_pos}",
                reference_position=target_pos, expected_residue=expected_residue)
        return None
    idx, matches = best
    observed_mass_delta = float(df.at[idx, "M"] - cur_M)
    best_match = min(matches, key=lambda m: abs(m[0] - abs(observed_mass_delta)))
    mass_error = abs(abs(observed_mass_delta) - best_match[0])

    rt_history = [df.at[i, "T"] for i in chain["indices"]]
    if end == "prepend":
        rt_history = list(reversed(rt_history))
    if len(rt_history) >= max(cfg.min_rt_history, 2):
        window = rt_history[-cfg.rt_window:]
        deltas = np.diff(window)
        mean_d = float(deltas.mean())
        std_d = max(float(deltas.std()), cfg.rt_std_floor)
        predicted_rt = rt_history[-1] + mean_d
        observed_rt = float(df.at[idx, "T"])
        if abs(observed_rt - predicted_rt) > cfg.rt_zscore * std_d:
            _reject(f"candidate peak matched on mass but failed the RT trend check: observed RT "
                    f"{observed_rt:.3f} vs predicted {predicted_rt:.3f} (tolerance "
                    f"{cfg.rt_zscore * std_d:.3f})",
                    reference_position=target_pos, expected_residue=expected_residue,
                    candidate_peak_mass=round(float(df.at[idx, "M"]), 4),
                    observed_mass_delta=round(observed_mass_delta, 4))
            return None  # still has to respect criterion 3 (RT trend)
        rt_support = (f"pass: observed RT {observed_rt:.3f} within "
                       f"{cfg.rt_zscore:g}x trend std ({abs(observed_rt - predicted_rt):.3f} / "
                       f"{cfg.rt_zscore * std_d:.3f} min of predicted {predicted_rt:.3f})")
    else:
        rt_support = "not checked: not enough RT history yet on this read to fit a trend"

    donor_ranks = sorted(
        chains[j]["read_rank"] for j in range(len(chains)) if j != ci and idx in chains[j]["indices"]
    )
    reused_from_read_rank = donor_ranks[0] if donor_ranks else None

    alignment_confidence = round(score / max(len(ordered_calls), 1), 3)
    reuse_evidence_notes = [
        f"reference alignment: {score}/{len(ordered_calls)} positions match (confidence {alignment_confidence:.2f})",
        f"expected residue at reference position {target_pos}: {expected_residue}",
        f"observed mass delta {observed_mass_delta:+.4f} Da matches {'/'.join(best_match[1])} "
        f"(table value {best_match[0]:.4f} Da, mass_error {mass_error:.4f} Da)",
        f"RT support: {rt_support}",
    ]
    reuse_warning = []
    if alignment_confidence < 0.8:
        reuse_warning.append(f"alignment confidence {alignment_confidence:.2f} is modest -- recommend manual review")
    if mass_error > cfg.mass_tol * 0.75:
        reuse_warning.append(f"mass_error {mass_error:.4f} Da is close to the {cfg.mass_tol} Da tolerance")
    if len(donor_ranks) > 1:
        reuse_warning.append(f"this peak is already shared by reads {donor_ranks} -- reused by more than one other read")

    evidence = {
        "reused_from_read_rank": reused_from_read_rank,
        "reused_in_read_rank": chain["read_rank"],
        "reference_position": target_pos,
        "expected_residue": expected_residue,
        "observed_mass_delta": round(observed_mass_delta, 4),
        "mass_error": round(mass_error, 4),
        "RT_support": rt_support,
        "reuse_evidence": reuse_evidence_notes,
        "reuse_warning": reuse_warning,
        # Internal-only, not part of the public schema above -- lets
        # assign_peak_status find this point without re-deriving it.
        "_peak_index": int(idx),
    }
    if audit_log is not None:
        audit_log.append({
            "read_rank": read_rank, "end": end, "outcome": "accepted",
            "reason": "accepted -- mass match and RT trend both support reuse (see reuse_evidence on the read)",
            "reference_position": target_pos, "expected_residue": expected_residue,
            "candidate_peak_mass": round(float(df.at[idx, "M"]), 4),
            "observed_mass_delta": round(observed_mass_delta, 4),
            "reused_from_read_rank": reused_from_read_rank,
        })
    return idx, matches, evidence


def apply_reference_guided_extension(
    df: pd.DataFrame,
    chains: List[Dict],
    used: pd.Series,
    diff_table: List[MassDiffEntry],
    cfg: Config,
) -> List[Dict]:
    """Step 6.6, STAGE 2: runs strictly AFTER Stage 1 (every primary read
    has already been built by strict greedy walking, and ranked into
    read_rank). For each read, in read_rank order, try to extend its ends
    further by reusing points already flagged by other reads -- but only
    when the reference sequence strongly predicts the residue there. This
    never changes which points seeded or built a Stage-1 read, and never
    runs interleaved with Stage-1 generation; it only adds extra,
    explicitly-evidenced points on top. No-op when `cfg.reference` is unset
    (every chain still gets an empty `reuse_evidence` list for a uniform
    report schema).

    Returns the reuse candidate audit log (review-and-correction pass
    requirement 3): one entry per tested candidate, accepted or rejected
    with a reason, across every chain and end direction. Always an empty
    list when `cfg.reference` is unset -- that emptiness means "reuse was
    never tested (no reference provided)", which is a different fact than
    "reuse was tested and zero candidates qualified" (a non-empty log of
    all-rejected entries). `build_reference_reuse_audit_table` turns this
    into the three-way candidates/accepted/rejected split for review.
    """
    for chain in chains:
        chain.setdefault("reuse_evidence", [])
        chain.setdefault("_reused_peak_indices", [])
    audit_log: List[Dict] = []
    if not cfg.reference or not chains:
        return audit_log

    max_mass_diff = diff_table.max_total
    sorted_idx = df.sort_values("M").index.to_numpy()
    order = sorted(range(len(chains)), key=lambda i: chains[i]["read_rank"])

    for ci in order:
        chain = chains[ci]
        for end in ("append", "prepend"):
            while True:
                step = _reference_guided_step(
                    df, sorted_idx, used, diff_table, cfg, max_mass_diff, chains, ci, end, audit_log
                )
                if step is None:
                    break
                idx, matches, evidence = step
                cur = chain["indices"][-1] if end == "append" else chain["indices"][0]
                delta = abs(df.at[idx, "M"] - df.at[cur, "M"])
                best_label = min(matches, key=lambda m: abs(m[0] - delta))[1]
                chain["_reused_peak_indices"].append(evidence.pop("_peak_index"))
                chain["reuse_evidence"].append(evidence)
                if end == "append":
                    chain["indices"].append(idx)
                    chain["edge_labels"].append(best_label)
                else:
                    chain["indices"].insert(0, idx)
                    chain["edge_labels"].insert(0, best_label)
    return audit_log


def _extend_direction(
    df: pd.DataFrame,
    sorted_idx: np.ndarray,
    pos_of: Dict[int, int],
    seed_idx: int,
    used: np.ndarray,
    diff_table: List[MassDiffEntry],
    cfg: Config,
    max_mass_diff: float,
    increasing: bool,
) -> Tuple[List[int], List[Tuple[str, ...]]]:
    """Strict primary walk (Steps 6.1-6.3): walk away from seed_idx in one
    mass direction using ONLY unused points, marking each as used. This is
    Phase 1 of base-calling and never reuses an already-flagged point --
    reference-guided reuse is a separate, strictly second-stage pass (see
    `apply_reference_guided_extension`) that runs only after every primary
    read like this one has already been built.
    """
    chain: List[int] = []
    edge_labels: List[Tuple[str, ...]] = []
    cur = seed_idx
    rt_history = [df.at[seed_idx, "T"]]

    while True:
        cands = _find_candidates(df, sorted_idx, pos_of, cur, used, diff_table, cfg.mass_tol, max_mass_diff)
        cands = [
            (idx, m) for idx, m in cands
            if (df.at[idx, "M"] > df.at[cur, "M"]) == increasing
        ]
        choice = _select_next(df, cands, rt_history, cfg)
        if choice is None:
            break  # Step 6.3: chain ends here
        idx, matches = choice
        # Prefer the match whose summed mass is closest to the observed delta.
        delta = abs(df.at[idx, "M"] - df.at[cur, "M"])
        best_label = min(matches, key=lambda m: abs(m[0] - delta))[1]
        chain.append(idx)
        edge_labels.append(best_label)
        used[idx] = True
        rt_history.append(df.at[idx, "T"])
        cur = idx

    return chain, edge_labels


def build_chain(
    df: pd.DataFrame,
    sorted_idx: np.ndarray,
    pos_of: Dict[int, int],
    seed_idx: int,
    used: np.ndarray,
    diff_table: List[MassDiffEntry],
    cfg: Config,
) -> Dict:
    """Steps 6.1-6.3 (Phase 1, strict/primary): grow a ladder outward from a
    seed point in both mass directions using ONLY unused points, until no
    further point satisfies all three criteria. `reuse_evidence` starts
    empty and is only ever populated later, by the separate Phase 2 pass
    (`apply_reference_guided_extension`) -- never during this primary build.
    """
    max_mass_diff = diff_table.max_total
    used[seed_idx] = True

    down_chain, down_labels = _extend_direction(
        df, sorted_idx, pos_of, seed_idx, used, diff_table, cfg, max_mass_diff, increasing=False
    )
    up_chain, up_labels = _extend_direction(
        df, sorted_idx, pos_of, seed_idx, used, diff_table, cfg, max_mass_diff, increasing=True
    )

    full_chain = list(reversed(down_chain)) + [seed_idx] + up_chain
    full_labels = list(reversed(down_labels)) + up_labels  # one fewer than points
    return {
        "indices": full_chain,
        "edge_labels": full_labels,
        "seed_index": seed_idx,
        "reuse_evidence": [],
    }


# ---------------------------------------------------------------------------
# Step 6.4: 5'/3' ladder classification via paired-read evidence
# ---------------------------------------------------------------------------

def _chain_seed_stats(df: pd.DataFrame, chain: Dict) -> Dict:
    seed_idx = chain["seed_index"]
    m = float(df.at[seed_idx, "M"])
    return {
        "seed_mass": m,
        "seed_mass_int": int(m // 1),
        "seed_mass_frac": float(m % 1.0),
        "seed_block": int(df.at[seed_idx, "block"]) if "block" in df.columns else None,
        "seed_rt": float(df.at[seed_idx, "T"]),
        "seed_rel_i": float(df.at[seed_idx, "Rel_I"]),
    }


def _pairwise_scores_vectorized(
    chains: List[Dict],
    cfg: Config,
    seed_stats: List[Dict],
    mass_ranges: List[Tuple[float, float]],
    t_std: float,
    ranks_arr: np.ndarray,
) -> Dict[str, np.ndarray]:
    """Score every (i, j) candidate chain pair (i < j) in one vectorized pass.

    Same scoring rule as before (see module docstring / Dr. Jiang's pairing
    rule: the 5' read has a SMALLER seed-mass decimal AND a LONGER RT than
    its corresponding 3' read), just computed for all ~n^2/2 pairs at once
    with numpy instead of one Python function call per pair. On real data
    n is in the thousands, so a plain Python loop over n^2/2 pairs -- even
    after removing string formatting from the per-pair cost -- was still
    too slow (tens of millions of Python-level function calls and dict
    constructions). Operating on flat 1-D arrays of length n*(n-1)/2
    (rather than full n x n matrices) keeps memory to roughly half of the
    square-matrix approach, which matters at n ~ 4500 (n^2/2 ~ 10M pairs).

    Returns a dict of 1-D numpy arrays (one entry per pair, aligned with
    `ii`/`jj`) plus the `ii`/`jj` index arrays themselves.
    """
    n = len(chains)
    ii, jj = np.triu_indices(n, k=1)

    frac = np.array([s["seed_mass_frac"] for s in seed_stats], dtype=np.float64)
    rt = np.array([s["seed_rt"] for s in seed_stats], dtype=np.float64)
    block = np.array(
        [s["seed_block"] if s["seed_block"] is not None else np.nan for s in seed_stats],
        dtype=np.float64,
    )
    mass_min = np.array([mr[0] for mr in mass_ranges], dtype=np.float64)
    mass_max = np.array([mr[1] for mr in mass_ranges], dtype=np.float64)

    decimal_diff = frac[jj] - frac[ii]   # >0 => j's decimal is larger than i's
    rt_diff = rt[jj] - rt[ii]            # >0 => j's RT is longer than i's

    # A clean pair has these point the SAME direction: whichever read has
    # the smaller decimal must also be the one with the longer RT.
    sign_consistent = ((decimal_diff > 0) & (rt_diff < 0)) | ((decimal_diff < 0) & (rt_diff > 0))

    decimal_strength = np.minimum(np.abs(decimal_diff) / 0.5, 1.0)
    rt_strength = np.minimum(np.abs(rt_diff) / max(t_std, 1e-6), 1.0)

    rank_diff = np.abs(ranks_arr[ii] - ranks_arr[jj])
    rank_score = np.maximum(0.0, 1.0 - rank_diff / max(n, 1))

    a_min, a_max = mass_min[ii], mass_max[ii]
    b_min, b_max = mass_min[jj], mass_max[jj]
    overlap = np.minimum(a_max, b_max) - np.maximum(a_min, b_min)
    span = np.maximum(a_max, b_max) - np.minimum(a_min, b_min)
    safe_span = np.where(span > 0, span, 1.0)
    overlap_score = np.where(span > 0, np.maximum(0.0, overlap / safe_span), 0.0)
    del overlap, span, safe_span

    same_start_block = (block[ii] == block[jj]) & ~np.isnan(block[ii]) & ~np.isnan(block[jj])

    weights = dict(decimal=0.30, rt=0.25, rank=0.10, overlap=0.20, block=0.05, closure=0.10)
    total = (
        weights["decimal"] * decimal_strength
        + weights["rt"] * rt_strength
        + weights["rank"] * rank_score
        + weights["overlap"] * overlap_score
        + weights["block"] * same_start_block.astype(np.float64)
    )
    del decimal_strength, rt_strength, rank_score

    closure_diff = None
    if cfg.precursor_mass is not None:
        closure_diff = np.abs((a_max + b_max) - cfg.precursor_mass)
        closure_score = np.maximum(0.0, 1.0 - closure_diff / max(cfg.precursor_mass_tol, 1e-6))
        total = total + weights["closure"] * closure_score
        del closure_score

    # Mild discount only -- this is a tie-breaking preference so that, when
    # two candidate partners are otherwise similar, the sign-consistent one
    # wins the greedy 1:1 assignment. It must NOT be strong enough to push
    # a well-matched-but-contradictory pair below `min_pair_score`, because
    # that would silently route a genuine conflict into the fallback path
    # instead of surfacing it as "conflict" -- which defeats the purpose of
    # having that category.
    total = np.where(sign_consistent, total, total * 0.85)

    return {
        "ii": ii, "jj": jj, "score": total, "sign_consistent": sign_consistent,
        "decimal_diff": decimal_diff, "rt_diff": rt_diff, "overlap_score": overlap_score,
        "rank_diff": rank_diff, "same_start_block": same_start_block, "closure_diff": closure_diff,
    }


def _pairwise_evidence_text(seed_stats: List[Dict], pair: Dict) -> Tuple[List[str], List[str]]:
    """Build the human-readable evidence/warnings lists for one already-
    scored pair (see `_pairwise_score`). Only called for the handful of
    pairs the greedy matcher in `find_best_pairs` actually keeps.
    """
    i, j = pair["i"], pair["j"]
    sa, sb = seed_stats[i], seed_stats[j]
    decimal_diff, rt_diff = pair["decimal_diff"], pair["rt_diff"]
    sign_consistent = pair["sign_consistent"]
    overlap_score = pair["overlap_score"]
    rank_diff = pair["rank_diff"]
    same_start_block = pair["same_start_block"]
    closure_diff = pair["closure_diff"]
    closure_note = (
        "not evaluated (no precursor_mass configured)" if closure_diff is None
        else f"|sum of terminal masses - precursor| = {closure_diff:.3f} Da"
    )

    evidence = [
        f"seed decimal: chain {i}={sa['seed_mass_frac']:.3f} vs chain {j}={sb['seed_mass_frac']:.3f} (diff={decimal_diff:+.3f})",
        f"seed RT: chain {i}={sa['seed_rt']:.3f} vs chain {j}={sb['seed_rt']:.3f} (diff={rt_diff:+.3f})",
        f"decimal/RT sign consistency: {'consistent with Jiang rule' if sign_consistent else 'CONTRADICTORY'}",
        f"intensity-rank gap between the two reads: {rank_diff}",
        f"mass-range overlap fraction: {overlap_score:.2f}",
        f"same starting block: {same_start_block}",
        f"precursor mass closure: {closure_note}",
    ]
    warnings = []
    if not sign_consistent:
        warnings.append("decimal-part and RT evidence disagree on which read is 5' vs 3' for this candidate pair")
    if overlap_score < 0.05:
        warnings.append("mass ranges barely overlap; pair may not originate from the same molecule")
    return evidence, warnings


def find_best_pairs(df: pd.DataFrame, chains: List[Dict], cfg: Config) -> Dict[int, Optional[Dict]]:
    """Greedy 1:1 matching: score every candidate chain pair, then assign
    partners in descending score order so each chain is paired at most once.
    Returns {chain_index: pair_evidence_dict_or_None}.

    This is an O(n^2) all-pairs comparison, which is unavoidable for a true
    "best partner out of all reads" search -- but on real data n can be in
    the thousands (most short-lived noise chains never make it past
    `min_chain_len`, but plenty do), so the per-pair scoring is computed for
    ALL pairs at once with numpy (`_pairwise_scores_vectorized`) rather than
    with a Python-level loop -- at n ~ 4500 that's ~10M pairs, which a plain
    Python loop cannot get through fast enough even with no string
    formatting in the hot path. The expensive human-readable evidence text
    is only built afterwards, for the <=n/2 pairs the greedy matcher below
    actually keeps (see `_pairwise_evidence_text`).
    """
    n = len(chains)
    if n < 2:
        return {i: None for i in range(n)}

    rank_order = sorted(range(n), key=lambda i: -df.at[chains[i]["seed_index"], "Rel_I"])
    ranks_arr = np.empty(n, dtype=np.float64)
    for r, i in enumerate(rank_order):
        ranks_arr[i] = r

    seed_stats = [_chain_seed_stats(df, c) for c in chains]
    m_values = df["M"].values
    mass_ranges = [
        (float(np.min(m_values[c["indices"]])), float(np.max(m_values[c["indices"]])))
        for c in chains
    ]
    t_std = float(df["T"].std())

    scored = _pairwise_scores_vectorized(chains, cfg, seed_stats, mass_ranges, t_std, ranks_arr)
    order = np.argsort(-scored["score"], kind="stable")

    # Reorder once into plain numpy arrays (cheap: each is a single
    # vectorized fancy-index pass), but deliberately do NOT call .tolist()
    # on these -- at real-workbook scale n_chains can be in the thousands,
    # so n*(n-1)/2 is in the tens of millions, and converting that many
    # elements to Python lists (x10 fields) costs real wall-clock time
    # (tens of seconds) for a loop that, sorted descending by score, almost
    # always breaks out after only a few thousand iterations at most. Index
    # the reordered numpy arrays directly inside the loop instead, so cost
    # scales with iterations actually executed, not with n^2.
    ii_sorted = scored["ii"][order]
    jj_sorted = scored["jj"][order]
    score_sorted = scored["score"][order]
    sign_sorted = scored["sign_consistent"][order]
    decimal_diff_sorted = scored["decimal_diff"][order]
    rt_diff_sorted = scored["rt_diff"][order]
    overlap_sorted = scored["overlap_score"][order]
    rank_diff_sorted = scored["rank_diff"][order]
    same_block_sorted = scored["same_start_block"][order]
    closure_diff_sorted = scored["closure_diff"][order] if scored["closure_diff"] is not None else None
    del scored, order

    assigned: Dict[int, Optional[Dict]] = {i: None for i in range(n)}
    taken: set = set()
    for k in range(len(ii_sorted)):
        score = float(score_sorted[k])
        if score < cfg.min_pair_score:
            # Sorted descending by score -- nothing from here on can pass
            # the threshold either, so the rest of the loop is dead work.
            break
        i, j = int(ii_sorted[k]), int(jj_sorted[k])
        if i in taken or j in taken:
            continue
        pair = {
            "i": i, "j": j, "score": score, "sign_consistent": bool(sign_sorted[k]),
            "decimal_diff": float(decimal_diff_sorted[k]), "rt_diff": float(rt_diff_sorted[k]),
            "overlap_score": float(overlap_sorted[k]), "rank_diff": int(rank_diff_sorted[k]),
            "same_start_block": bool(same_block_sorted[k]),
            "closure_diff": (float(closure_diff_sorted[k]) if closure_diff_sorted is not None else None),
        }
        evidence, warnings = _pairwise_evidence_text(seed_stats, pair)
        pair["evidence"], pair["warnings"] = evidence, warnings
        assigned[i] = pair
        assigned[j] = pair
        taken.add(i)
        taken.add(j)
        if len(taken) == n:
            break
    return assigned


# How many top reads `Top_Parallel_Reads` shows side by side. Review tiering
# (`build_read_summary`'s `top_parallel_group` column) keys off this same
# number so the two outputs always describe the same set of reads.
TOP_PARALLEL_N = 4


def _human_ladder_label(ladder_type: Optional[str]) -> str:
    """Human-facing call for Dr. Jiang's review table and the dashboard. The
    internal likely_5prime/likely_3prime/ambiguous/conflict labels still
    exist unchanged elsewhere (`ladder_classification`) -- this is just what
    a reviewer should read at a glance: 5' / 3' / ambiguous / conflict.
    (Earlier drafts of this table used '?' for the ambiguous case; the
    dashboard design spec calls for the literal word 'ambiguous' instead, so
    this is the single place that changed.)"""
    return {
        "likely_5prime": "5'",
        "likely_3prime": "3'",
        "ambiguous": "ambiguous",
        "conflict": "conflict",
    }.get(ladder_type, "ambiguous")


def _confidence_tier(confidence: Optional[float]) -> str:
    """3-tier human confidence label (high/medium/low) shown alongside the
    human ladder call, kept separate from the raw numeric `ladder_confidence`
    score. By construction in `classify_ladders`: ambiguous reads always
    have confidence exactly 0.0, and conflicts are always capped at <= 0.4 --
    so both land in 'low' here every time, never 'medium'/'high'. Only a
    paired likely_5prime/likely_3prime call (real, consistent paired-read
    evidence) can reach 'medium' or 'high'."""
    if confidence is None:
        return "low"
    if confidence >= 0.75:
        return "high"
    if confidence >= 0.45:
        return "medium"
    return "low"


def _short_text(text: Optional[str], max_chars: int = 100) -> str:
    """Compact single-line version of an already-joined evidence/warnings
    string (sequencing-assist dashboard data contract, Tasks 2-3). The full
    text stays available verbatim in `ladder_evidence`/`ladder_warnings`
    (Classification_Evidence.xlsx, read_summary.csv) -- this is only the
    short version for dense table cells where a multi-sentence string
    would overflow the row."""
    if not text:
        return ""
    return text if len(text) <= max_chars else text[: max_chars - 1].rstrip() + "…"


def _suggested_action(ladder_call: str, confidence_tier: Optional[str], low_confidence_noise: bool) -> str:
    """Plain-language next step for a read, keyed off the same human-facing
    call/tier/tiering fields the dashboard already shows (sequencing-assist
    dashboard data contract, Task 3). This is the one column that turns the
    decision table from "here are the rows" into "here is what to do" --
    no new evidence is computed, it is purely a mapping over existing
    fields:
      - ambiguous            -> review partner/reference before trusting it
      - conflict              -> decimal and RT evidence disagree, needs a human
      - 5'/3' + high/medium   -> safe to use as-is
      - 5'/3' + low confidence, or flagged low_confidence_noise -> deprioritize
    """
    if ladder_call == "ambiguous":
        return "Review partner read or reference sequence before using"
    if ladder_call == "conflict":
        return "Manual review required: decimal and RT evidence disagree"
    if ladder_call in ("5'", "3'"):
        if not low_confidence_noise and confidence_tier in ("high", "medium"):
            return "Use as primary sequencing evidence"
        return "Low priority unless needed for coverage"
    return "Low priority unless needed for coverage"


def classify_ladders(df: pd.DataFrame, chains: List[Dict], cfg: Config) -> None:
    """Step 6.4: classify every chain as likely_5prime / likely_3prime /
    ambiguous / conflict, with confidence + evidence + warnings, mutating
    each chain dict in place.

    Rule (per Dr. Jiang): find each chain's best candidate partner read (the
    complementary 5'/3' read of the same molecule) and compare them
    directly -- smaller seed-mass decimal AND longer RT than the partner =>
    5'-ladder. There is no other primary rule:
      - a partner found, and decimal/RT agree  -> likely_5prime / likely_3prime
      - a partner found, but decimal/RT disagree -> conflict (contradictory
        evidence -- this case genuinely has a partner, it just disagrees)
      - no acceptable partner found             -> ambiguous (no evidence
        either way -- NOT a guess)
    The across-chain seed-decimal median is never the determining rule. For
    an unpaired chain it is reported as a low-weight, clearly-labeled
    supplementary note only, with confidence 0.0, so a reader never mistakes
    "ambiguous" for a confident-but-weak answer.
    """
    pairs = find_best_pairs(df, chains, cfg)

    if chains:
        seed_fracs = [float(df.at[c["seed_index"], "M"] % 1.0) for c in chains]
        med_frac = float(np.median(seed_fracs))
    else:
        med_frac = 0.5

    for idx, chain in enumerate(chains):
        pair = pairs.get(idx)
        chain["candidate_partner_rank"] = None
        chain["paired_decimal_difference"] = None
        chain["paired_rt_difference"] = None
        chain["paired_length_overlap"] = None
        chain["mass_difference_consistency"] = None

        if pair is None:
            chain["paired_with"] = None
            chain["pair_score"] = None
            seed_frac = float(df.at[chain["seed_index"], "M"] % 1.0)
            chain["ladder_type"] = "ambiguous"
            chain["ladder_confidence"] = 0.0
            chain["ladder_evidence"] = [
                "no candidate partner read met min_pair_score -- this chain has no 5'/3' partner "
                "evidence either way, so it cannot be called 5' or 3'",
                f"supplementary, non-determining note: seed decimal {seed_frac:.3f} vs "
                f"across-chain median {med_frac:.3f} (this is NOT used to decide the label)",
            ]
            chain["ladder_warnings"] = ["no paired read found -- recommend manual review or a reference sequence to disambiguate"]
            continue

        partner_idx = pair["j"] if pair["i"] == idx else pair["i"]
        chain["paired_with"] = partner_idx
        chain["pair_score"] = round(pair["score"], 3)
        chain["candidate_partner_rank"] = chains[partner_idx].get("read_rank")

        # Express the pair's signed diffs from THIS chain's point of view
        # (pair stores them as j-minus-i, so flip sign when this chain is j).
        sign = 1.0 if pair["i"] == idx else -1.0
        chain["paired_decimal_difference"] = round(sign * pair["decimal_diff"], 4)
        chain["paired_rt_difference"] = round(sign * pair["rt_diff"], 4)
        chain["paired_length_overlap"] = round(pair["overlap_score"], 3)
        chain["mass_difference_consistency"] = bool(pair["sign_consistent"])

        # Read length / overlap both feed confidence -- a pair built on two
        # short, barely-overlapping reads should never score as high as one
        # built on two long, well-overlapping reads, even with an identical
        # decimal/RT signal.
        len_self = len(chain["indices"])
        len_partner = len(chains[partner_idx]["indices"])
        length_factor = min(min(len_self, len_partner) / 6.0, 1.0)

        if not pair["sign_consistent"]:
            chain["ladder_type"] = "conflict"
            chain["ladder_confidence"] = round(min(pair["score"], 0.4) * (0.6 + 0.4 * length_factor), 3)
            chain["ladder_evidence"] = pair["evidence"]
            chain["ladder_warnings"] = pair["warnings"] + [f"paired with chain {partner_idx}, but decimal/RT evidence conflicts -- needs manual review"]
            continue

        i, j = pair["i"], pair["j"]
        idx_has_smaller_decimal = (idx == i and pair["decimal_diff"] > 0) or (idx == j and pair["decimal_diff"] < 0)
        ladder = "likely_5prime" if idx_has_smaller_decimal else "likely_3prime"
        confidence = round(min(0.5 + 0.5 * pair["score"], 0.97) * (0.6 + 0.4 * length_factor), 3)

        chain["ladder_type"] = ladder
        chain["ladder_confidence"] = confidence
        chain["ladder_evidence"] = pair["evidence"] + [
            f"paired with chain {partner_idx} (pair score={pair['score']:.2f}, "
            f"read lengths {len_self}/{len_partner})"
        ]
        chain["ladder_warnings"] = pair["warnings"]


def assign_peak_status(df: pd.DataFrame, chains: List[Dict]) -> pd.Series:
    """Step 6.5/6.7: one peak_status label per data point, for the final
    annotated table. Five possible values:

      primary_used       -- point belongs to a kept read (Phase 1) that
                             Step 6.4 confidently called likely_5prime or
                             likely_3prime.
      ambiguous_retained  -- point belongs to a kept read Step 6.4 could not
                             call (no partner found). Reported, NOT dropped
                             ("not aggressively excluded" per Step 6.5).
      conflict_retained   -- point belongs to a kept read whose partner
                             evidence was contradictory. Also reported, not
                             dropped.
      reference_reused    -- point was borrowed by Stage 2 (Step 6.6) to
                             extend a DIFFERENT read than the one it
                             originally belonged to (see `reuse_evidence`
                             on the read that borrowed it for full detail).
      unused              -- point never joined any kept read (either it was
                             never picked at all, or it was an unextendable
                             single-peak "chain" of length < min_chain_len
                             that got dropped).

    A point can legitimately play two roles at once (e.g. it anchors one
    primary read AND is also reused to extend a second read) -- when that
    happens this single summary column reports 'reference_reused', since
    that is the more notable fact; the complete picture for every role a
    point played is still in each read's own `indices` / `reuse_evidence`.
    """
    status = pd.Series("unused", index=df.index, dtype=object)
    label_to_status = {
        "likely_5prime": "primary_used",
        "likely_3prime": "primary_used",
        "ambiguous": "ambiguous_retained",
        "conflict": "conflict_retained",
    }
    for chain in chains:
        s = label_to_status.get(chain.get("ladder_type"), "ambiguous_retained")
        for i in chain["indices"]:
            status.at[i] = s
    for chain in chains:
        for reused_idx in chain.get("_reused_peak_indices", []):
            status.at[reused_idx] = "reference_reused"
    return status


def run_base_calling(df: pd.DataFrame, cfg: Config) -> Tuple[List[Dict], List[Dict]]:
    """Steps 6.1-6.7, run as 3 explicit, non-interleaved phases:

    Phase 1 (primary, strict): repeatedly seed a new "parallel short read"
    from the highest block-wise-Rel_I unused point (Jiang-style global
    flag-and-exclude), walk it out via the strict Step 6.2 criteria only,
    flag every point it uses, and repeat from the next-highest unflagged
    seed. This is the ONLY phase that decides which points belong to which
    primary read. Reads are then ranked by descending seed Rel_I and given a
    stable `read_rank` (1 = highest intensity) used everywhere downstream
    (the Top_Parallel_Reads table, Stage-2 reuse evidence, the read summary).

    Phase 2 (second-stage, optional): only after every primary read from
    Phase 1 exists and has a read_rank, optionally extend read ends using
    reference-guided reuse of already-flagged points (Step 6.6) -- see
    `apply_reference_guided_extension`. No-op without `cfg.reference`.

    Phase 3 (classification): apply Step 6.4's paired-read 5'/3'
    classification to the final read set -- see `classify_ladders`. Once
    every chain has a ladder_type, every data point gets an explicit
    `peak_status` (Step 6.5/6.7) -- see `assign_peak_status`.

    Returns (chains, reuse_audit_log) -- the second element is the Step 6.6
    candidate audit log from Phase 2 (see `apply_reference_guided_extension`),
    always `[]` when no reference was provided.
    """
    diff_table = build_mass_diff_table(cfg.allowed_masses, cfg.effective_max_residues_per_step)
    sorted_idx = df.sort_values("M").index.to_numpy()
    pos_of = {idx: pos for pos, idx in enumerate(sorted_idx)}

    # df.index may not be 0..n-1 contiguous after upstream filtering, so a
    # pandas Series (not a plain numpy array) is used as the "used" flag map.
    used = pd.Series(False, index=df.index)

    # Phase 1: sort ALL peaks by block-wise (linear) Rel_I descending, seed
    # from the highest unflagged point, strict walk, flag used peaks,
    # repeat from the next highest unflagged seed. Ambiguous/conflict reads
    # are NOT decided here -- they fall out of Phase 3 -- so they cannot
    # aggressively exclude peaks beyond what their own walk legitimately uses.
    intensity_order = df.sort_values("Rel_I", ascending=False).index.tolist()

    chains: List[Dict] = []
    for seed in intensity_order:
        if used[seed]:
            continue
        chain = build_chain(df, sorted_idx, pos_of, seed, used, diff_table, cfg)
        if len(chain["indices"]) >= cfg.min_chain_len:
            chains.append(chain)
        # length-1 "chains" (unextendable single peaks) are dropped from the
        # report but remain flagged as used, per 6.5.

    # Stable read_rank (1 = highest seed Rel_I), assigned right after Phase 1
    # so both Stage 2 (Phase 2) and the read summary can refer to reads by
    # this same number.
    for rank, chain in enumerate(
        sorted(chains, key=lambda c: -df.at[c["seed_index"], "Rel_I"]), start=1
    ):
        chain["read_rank"] = rank

    # Phase 2: second-stage only, strictly after Phase 1 is fully done.
    reuse_audit_log = apply_reference_guided_extension(df, chains, used, diff_table, cfg)

    # Phase 3: paired-read 5'/3' classification of the final read set.
    classify_ladders(df, chains, cfg)

    for chain in chains:
        chain["sequence_calls"] = [labels[0] if len(labels) == 1 else "/".join(labels) for labels in chain["edge_labels"]]
        chain["masses"] = df.loc[chain["indices"], "M"].tolist()
        chain["rel_intensities"] = df.loc[chain["indices"], "Rel_I"].tolist()
        chain["rts"] = df.loc[chain["indices"], "T"].tolist()

    df["peak_status"] = assign_peak_status(df, chains)  # Step 6.5/6.7

    return chains, reuse_audit_log


def build_read_summary(df: pd.DataFrame, chains: List[Dict], cfg: Config) -> pd.DataFrame:
    """Step 6.4 output table: one row per recovered read, with the full set
    of paired-read evidence fields the classification decision was actually
    based on -- this is the flat, manual-review-friendly counterpart to the
    nested chain dicts. Rows are sorted by read_rank (highest intensity
    first, matching the Top_Parallel_Reads table).

    Column groups (left to right):
      - human-facing call: ladder_call (5'/3'/?/conflict), ladder_confidence_tier
        (high/medium/low) -- read these first; they are what Dr. Jiang should
        actually read, not the internal labels further right.
      - review tiering (Stage-2 review-and-correction pass): review_priority,
        primary_review, top_parallel_group, low_confidence_noise -- which of
        the (often thousands of) chains are actually worth a human look.
      - the full paired-read evidence Task A asked for (seed/partner masses,
        decimal/RT/overlap diffs, ...).
      - decision_basis / fallback_median_used -- an explicit, auditable
        confirmation that a call was decided by real paired-read evidence
        and never by the across-chain median fallback (see classify_ladders).
      - internal labels last: ladder_classification (likely_5prime/
        likely_3prime/ambiguous/conflict) and the raw numeric ladder_confidence
        score -- preserved unchanged, just no longer the only thing shown.
    """
    rows = []
    for chain in chains:
        seed_idx = chain["seed_index"]
        seed_mass = float(df.at[seed_idx, "M"])
        partner_idx = chain.get("paired_with")
        partner = chains[partner_idx] if partner_idx is not None else None
        partner_seed_mass = float(df.at[partner["seed_index"], "M"]) if partner is not None else None

        closure = None
        if cfg.precursor_mass is not None and partner is not None:
            closure = abs((max(chain["masses"]) + max(partner["masses"])) - cfg.precursor_mass)

        ladder_type = chain.get("ladder_type")
        confidence = chain.get("ladder_confidence")
        tier = _confidence_tier(confidence)
        read_rank = chain["read_rank"]
        read_length = len(chain["indices"])

        # Decision basis: was this call actually decided from a paired
        # partner read, or is there no partner at all (-> ambiguous, by
        # design never decided by the across-chain median -- see
        # classify_ladders). fallback_median_used is always False under the
        # current rule and exists purely so that fact is auditable per row,
        # not just asserted in a docstring.
        decision_basis = "paired_evidence" if partner_idx is not None else "no_partner_found"
        fallback_median_used = False

        top_parallel_group = read_rank <= TOP_PARALLEL_N
        low_confidence_noise = (
            not top_parallel_group and read_length <= cfg.min_chain_len and tier == "low"
        )
        primary_review = top_parallel_group or (tier in ("high", "medium") and not low_confidence_noise)
        if top_parallel_group:
            review_priority = "1_top_parallel"
        elif primary_review:
            review_priority = "2_primary_review"
        elif low_confidence_noise:
            review_priority = "4_low_confidence_noise"
        else:
            review_priority = "3_secondary"

        rows.append({
            "read_rank": read_rank,
            "ladder_call": _human_ladder_label(ladder_type),
            "ladder_confidence_tier": tier,
            "review_priority": review_priority,
            "primary_review": primary_review,
            "top_parallel_group": top_parallel_group,
            "low_confidence_noise": low_confidence_noise,
            "seed_index": seed_idx,
            "seed_mass": round(seed_mass, 4),
            "seed_mass_integer": int(seed_mass // 1),
            "seed_mass_decimal": round(seed_mass % 1.0, 4),
            "seed_rt": round(float(df.at[seed_idx, "T"]), 3),
            "start_block": int(df.at[seed_idx, "block"]) if "block" in df.columns else None,
            "read_length": read_length,
            "mean_rel_i": round(float(np.mean(chain["rel_intensities"])), 4),
            "candidate_partner_rank": chain.get("candidate_partner_rank"),
            "partner_seed_mass": round(partner_seed_mass, 4) if partner_seed_mass is not None else None,
            "partner_seed_mass_integer": int(partner_seed_mass // 1) if partner_seed_mass is not None else None,
            "partner_seed_mass_decimal": round(partner_seed_mass % 1.0, 4) if partner_seed_mass is not None else None,
            "paired_decimal_difference": chain.get("paired_decimal_difference"),
            "paired_rt_difference": chain.get("paired_rt_difference"),
            "paired_length_overlap": chain.get("paired_length_overlap"),
            "mass_difference_consistency": chain.get("mass_difference_consistency"),
            "precursor_or_intact_mass_closure": round(closure, 4) if closure is not None else None,
            "decision_basis": decision_basis,
            "fallback_median_used": fallback_median_used,
            "ladder_classification": ladder_type,
            "ladder_confidence": confidence,
            "ladder_evidence": "; ".join(chain.get("ladder_evidence") or []),
            "ladder_warnings": "; ".join(chain.get("ladder_warnings") or []),
        })
    cols = ["read_rank", "ladder_call", "ladder_confidence_tier",
            "review_priority", "primary_review", "top_parallel_group", "low_confidence_noise",
            "seed_index", "seed_mass", "seed_mass_integer", "seed_mass_decimal", "seed_rt",
            "start_block", "read_length", "mean_rel_i", "candidate_partner_rank", "partner_seed_mass",
            "partner_seed_mass_integer", "partner_seed_mass_decimal", "paired_decimal_difference",
            "paired_rt_difference", "paired_length_overlap", "mass_difference_consistency",
            "precursor_or_intact_mass_closure", "decision_basis", "fallback_median_used",
            "ladder_classification", "ladder_confidence", "ladder_evidence", "ladder_warnings"]
    out = pd.DataFrame(rows, columns=cols)
    # Nullable integer dtype so "no partner" prints as a blank cell, not an
    # ugly "7.0" float -- purely cosmetic, makes the CSV nicer to read by eye.
    for int_col in ("candidate_partner_rank", "partner_seed_mass_integer"):
        out[int_col] = out[int_col].astype("Int64")
    return out.sort_values("read_rank", ignore_index=True) if not out.empty else out


def build_classification_evidence_table(read_summary: pd.DataFrame) -> pd.DataFrame:
    """Step 6.4 audit export (review-and-correction pass requirement 2): a
    focused subset of `build_read_summary`'s columns, holding exactly the
    per-call audit trail Dr. Jiang asked for -- paired read rank, seed mass
    integer/decimal, partner seed mass integer/decimal, paired decimal/RT
    difference, read length overlap, evidence summary, and an explicit
    fallback/median warning -- without the rest of the review-tiering
    columns that `Read_Summary_Review.xlsx` also carries.
    """
    cols = [
        "read_rank", "ladder_call", "ladder_confidence_tier",
        "candidate_partner_rank", "seed_mass_integer", "seed_mass_decimal",
        "partner_seed_mass_integer", "partner_seed_mass_decimal",
        "paired_decimal_difference", "paired_rt_difference", "paired_length_overlap",
        "decision_basis", "fallback_median_used", "ladder_evidence", "ladder_warnings",
    ]
    if read_summary.empty:
        return pd.DataFrame(columns=cols)
    return read_summary.loc[:, cols].copy()


def build_sequencing_decision_summary(read_summary: pd.DataFrame) -> pd.DataFrame:
    """sequencing_decision_summary.csv -- sequencing-assist dashboard data
    contract, Task 3. One row per read. This is the dashboard's primary
    decision-support table: every column already exists in
    `build_read_summary`'s output except `suggested_action`, which is a
    plain-language next step derived from the same ladder_call/
    confidence_tier/low_confidence_noise fields (see `_suggested_action`).
    The point of this table is that the dashboard guides what a researcher
    should do with a read, not just what was found about it.
    """
    cols = ["read_rank", "ladder_call", "confidence_tier", "review_priority",
            "read_length", "mean_rel_i", "candidate_partner_rank",
            "decision_basis", "evidence_short", "warning_short", "suggested_action"]
    if read_summary.empty:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(index=read_summary.index)
    out["read_rank"] = read_summary["read_rank"]
    out["ladder_call"] = read_summary["ladder_call"]
    out["confidence_tier"] = read_summary["ladder_confidence_tier"]
    out["review_priority"] = read_summary["review_priority"]
    out["read_length"] = read_summary["read_length"]
    out["mean_rel_i"] = read_summary["mean_rel_i"]
    out["candidate_partner_rank"] = read_summary["candidate_partner_rank"]
    out["decision_basis"] = read_summary["decision_basis"]
    out["evidence_short"] = read_summary["ladder_evidence"].map(_short_text)
    out["warning_short"] = read_summary["ladder_warnings"].map(_short_text)
    out["suggested_action"] = [
        _suggested_action(call, tier, noise)
        for call, tier, noise in zip(
            read_summary["ladder_call"], read_summary["ladder_confidence_tier"], read_summary["low_confidence_noise"]
        )
    ]
    return out.loc[:, cols]


def build_peak_status_table(df: pd.DataFrame, chains: List[Dict]) -> pd.DataFrame:
    """Peak_Status.xlsx content (review-and-correction pass requirement 5):
    one row per input data point with its mass/intensity/RT, `peak_status`
    (Step 6.5/6.7's primary_used / ambiguous_retained / conflict_retained /
    reference_reused / unused taxonomy), and -- when the point belongs to a
    kept read -- that read's rank, human-facing ladder call, and role
    (primary vs. reused-by-a-different-read), so a reviewer can see at a
    glance which read a given peak supports and why, not just its bare
    status label.
    """
    read_rank_of: Dict[int, int] = {}
    ladder_call_of: Dict[int, str] = {}
    role_of: Dict[int, str] = {}
    for chain in chains:
        call = _human_ladder_label(chain.get("ladder_type"))
        for i in chain["indices"]:
            read_rank_of[i] = chain["read_rank"]
            ladder_call_of[i] = call
            role_of.setdefault(i, "primary_used")
        for i in chain.get("_reused_peak_indices", []):
            role_of[i] = "reused_by_other_read"

    out = pd.DataFrame(index=df.index)
    out["mass"] = df["M"]
    out["rel_intensity"] = df["Rel_I"]
    out["rt"] = df["T"]
    if "block" in df.columns:
        out["block"] = df["block"]
    out["peak_status"] = df["peak_status"] if "peak_status" in df.columns else ""
    out["read_rank"] = pd.Series(read_rank_of, dtype="Int64").reindex(df.index)
    out["ladder_call"] = pd.Series(ladder_call_of, dtype=object).reindex(df.index).fillna("")
    out["role"] = pd.Series(role_of, dtype=object).reindex(df.index).fillna("")
    return out.sort_values("mass", ignore_index=True)


def build_reference_reuse_audit_table(reuse_audit_log: List[Dict]) -> pd.DataFrame:
    """Reference_Reuse_Audit.xlsx content (review-and-correction pass
    requirement 3): a flat, one-row-per-tested-candidate table built
    straight from the Step 6.6 audit log (`apply_reference_guided_extension`'s
    second return value). Every row is a candidate reuse point the
    algorithm actually evaluated, tagged `outcome` accepted/rejected with a
    plain-language `reason` -- so "reuse was tested and nothing qualified"
    reads completely differently from an empty table ("reuse was never
    tested because no reference was provided"). Only produced by the
    pipeline when `cfg.reference` is set; otherwise this is legitimately
    empty (not missing).
    """
    cols = ["read_rank", "end", "outcome", "reason", "reference_position",
            "expected_residue", "candidate_peak_mass", "observed_mass_delta", "reused_from_read_rank"]
    if not reuse_audit_log:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(reuse_audit_log)
    for c in cols:
        if c not in out.columns:
            out[c] = None
    return out.loc[:, cols].sort_values(["read_rank", "end"], ignore_index=True)


# ---------------------------------------------------------------------------
# Jiang-style "top parallel reads" side-by-side table (Step 6.5/6.7 output)
# ---------------------------------------------------------------------------

def build_top_parallel_reads_grid(df: pd.DataFrame, chains: List[Dict], n: int = TOP_PARALLEL_N) -> List[List[str]]:
    """Step 6.5/6.7 output named "Top_Parallel_Reads": the top-n highest-
    (linear)-intensity short reads laid out side by side, one row per ladder
    position and one column block per read -- deliberately shaped like Dr.
    Jiang's own "Example of 4 top intensity sequences in the low mass
    region" chart so it is readable by eye, not just a long machine table.

    Each read gets a 2-row header (read rank + ladder classification +
    confidence, then column labels) so those per-read facts are stated once
    instead of repeated on every row. Ranking is ALWAYS by linear seed Rel_I
    (never log10_intensity_qc).

    Returns a plain grid (list of rows of strings) -- write it with
    `write_grid_to_csv` / `write_grid_to_xlsx`.
    """
    ranked = sorted(chains, key=lambda c: -df.at[c["seed_index"], "Rel_I"])[:n]
    if not ranked:
        return [["No reads recovered -- nothing to show."]]

    max_len = max(len(c["indices"]) for c in ranked)
    n_cols_per_read = 5  # Mass, Rel.I, RT, Call, Peak Status
    blank = ""

    header1: List[str] = []
    header2: List[str] = []
    for chain in ranked:
        call = _human_ladder_label(chain.get("ladder_type"))
        tier = _confidence_tier(chain.get("ladder_confidence"))
        label = f"Read (rank {chain['read_rank']}) -- {call} (confidence: {tier})"
        header1.extend([label] + [blank] * (n_cols_per_read - 1) + [blank])  # + spacer
        header2.extend(["Mass (Da)", "Rel.I", "RT (min)", "Call", "Peak Status", blank])

    rows = [header1, header2]
    for pos in range(max_len):
        row: List[str] = []
        for chain in ranked:
            if pos < len(chain["indices"]):
                idx = chain["indices"][pos]
                call = chain["sequence_calls"][pos - 1] if pos > 0 else "(seed)"
                status = df.at[idx, "peak_status"] if "peak_status" in df.columns else ""
                row.extend([
                    f"{float(df.at[idx, 'M']):.4f}",
                    f"{float(df.at[idx, 'Rel_I']):.3f}",
                    f"{float(df.at[idx, 'T']):.3f}",
                    str(call),
                    str(status),
                    blank,
                ])
            else:
                row.extend([blank] * n_cols_per_read + [blank])
        rows.append(row)
    return rows


def build_top_parallel_reads_long(
    df: pd.DataFrame, chains: List[Dict], read_summary: pd.DataFrame, n: int = TOP_PARALLEL_N
) -> pd.DataFrame:
    """top_parallel_reads_long.csv -- sequencing-assist dashboard data
    contract, Task 2. Same top-n-by-intensity reads as
    `build_top_parallel_reads_grid`, but reshaped to one row per (read,
    peak) instead of a side-by-side grid with merged headers and blank
    spacer columns. The grid/Excel table is for human eyes and is kept
    unchanged; the dashboard should read this long table instead of
    parsing the grid's header text.

    Per-read fields (ladder_call, confidence_tier, ...) are pulled from
    `read_summary` (already built earlier in `run_pipeline`) rather than
    recomputed, so this table can never disagree with
    Read_Summary_Review.xlsx / Classification_Evidence.xlsx for the same
    read.
    """
    cols = ["read_rank", "row_position", "mass", "rel_i", "rt", "call", "peak_status",
            "ladder_call", "confidence_tier", "ladder_classification", "ladder_confidence",
            "candidate_partner_rank", "evidence_short", "warning_short"]
    ranked = sorted(chains, key=lambda c: -df.at[c["seed_index"], "Rel_I"])[:n]
    if not ranked:
        return pd.DataFrame(columns=cols)

    summary_by_rank = read_summary.set_index("read_rank").to_dict("index") if not read_summary.empty else {}

    rows = []
    for chain in ranked:
        read_rank = chain["read_rank"]
        srow = summary_by_rank.get(read_rank, {})
        ladder_call = srow.get("ladder_call", _human_ladder_label(chain.get("ladder_type")))
        confidence_tier = srow.get("ladder_confidence_tier", _confidence_tier(chain.get("ladder_confidence")))
        ladder_classification = srow.get("ladder_classification", chain.get("ladder_type"))
        ladder_confidence = srow.get("ladder_confidence", chain.get("ladder_confidence"))
        candidate_partner_rank = srow.get("candidate_partner_rank", chain.get("candidate_partner_rank"))
        evidence_short = _short_text(srow.get("ladder_evidence") or "; ".join(chain.get("ladder_evidence") or []))
        warning_short = _short_text(srow.get("ladder_warnings") or "; ".join(chain.get("ladder_warnings") or []))

        for pos, idx in enumerate(chain["indices"]):
            call = chain["sequence_calls"][pos - 1] if pos > 0 else "(seed)"
            rows.append({
                "read_rank": read_rank,
                "row_position": pos,
                "mass": round(float(df.at[idx, "M"]), 4),
                "rel_i": round(float(df.at[idx, "Rel_I"]), 4),
                "rt": round(float(df.at[idx, "T"]), 3),
                "call": str(call),
                "peak_status": df.at[idx, "peak_status"] if "peak_status" in df.columns else "",
                "ladder_call": ladder_call,
                "confidence_tier": confidence_tier,
                "ladder_classification": ladder_classification,
                "ladder_confidence": ladder_confidence,
                "candidate_partner_rank": candidate_partner_rank,
                "evidence_short": evidence_short,
                "warning_short": warning_short,
            })
    return pd.DataFrame(rows, columns=cols)


def write_grid_to_csv(grid: List[List[str]], path: str) -> str:
    import csv
    with open(path, "w", newline="") as f:
        csv.writer(f).writerows(grid)
    return str(path)


def write_grid_to_xlsx(grid: List[List[str]], path: str, header_rows: int = 2) -> Optional[str]:
    """Excel export of `Top_Parallel_Reads`. Bolds the header rows, freezes
    them in place, and widens columns to fit their content -- purely
    cosmetic, but this table only does its job ("readable by eye") if it
    doesn't need horizontal scrolling and re-reading the header to make
    sense of each row. Returns None (and skips writing) if openpyxl isn't
    installed; the .csv counterpart is always written regardless."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "Top_Parallel_Reads"
    for row in grid:
        ws.append(row)
    bold = Font(bold=True)
    for r in range(1, min(header_rows, len(grid)) + 1):
        for cell in ws[r]:
            cell.font = bold
    if len(grid) > header_rows:
        ws.freeze_panes = f"A{header_rows + 1}"
    n_cols = max((len(row) for row in grid), default=0)
    for col_idx in range(1, n_cols + 1):
        longest = max((len(str(row[col_idx - 1])) for row in grid if col_idx <= len(row)), default=0)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(longest + 2, 8), 40)
    wb.save(path)
    return str(path)


def write_dataframes_to_xlsx(sheets: Dict[str, pd.DataFrame], path: str) -> Optional[str]:
    """Pre-dashboard review package writer (review-and-correction pass
    requirement 5): writes one or more named DataFrames as sheets in a
    single .xlsx, with bolded/frozen headers and content-fit column widths
    on every sheet -- the same readability treatment as
    `write_grid_to_xlsx`, generalized to plain tabular DataFrames. Returns
    None (and skips writing) if openpyxl isn't installed.
    """
    try:
        from openpyxl.styles import Font
    except ImportError:
        return None
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    bold = Font(bold=True)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = bold
        if ws.max_row > 1:
            ws.freeze_panes = "A2"
        sample_rows = min(ws.max_row, 200)
        for col_idx in range(1, ws.max_column + 1):
            longest = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, sample_rows + 1)),
                default=0,
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(longest + 2, 8), 40)
    wb.save(path)
    return str(path)


# ---------------------------------------------------------------------------
# Step 6.6: optional comparison against a reference sequence
# ---------------------------------------------------------------------------

def _global_align(a: Sequence[str], b: Sequence[str], mismatch_penalty: float = 1.0, gap_penalty: float = 1.0):
    """Minimal Needleman-Wunsch global alignment over arbitrary tokens
    (single-letter or modification codes), used to spot mismatches that may
    indicate modifications or isoforms (Step 6.6). No external deps.
    """
    n, m = len(a), len(b)
    score = np.zeros((n + 1, m + 1))
    for i in range(1, n + 1):
        score[i, 0] = -i * gap_penalty
    for j in range(1, m + 1):
        score[0, j] = -j * gap_penalty
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            match = score[i - 1, j - 1] + (1.0 if a[i - 1] == b[j - 1] else -mismatch_penalty)
            score[i, j] = max(match, score[i - 1, j] - gap_penalty, score[i, j - 1] - gap_penalty)

    aligned_a, aligned_b = [], []
    i, j = n, m
    while i > 0 or j > 0:
        if i > 0 and j > 0 and np.isclose(
            score[i, j], score[i - 1, j - 1] + (1.0 if a[i - 1] == b[j - 1] else -mismatch_penalty)
        ):
            aligned_a.append(a[i - 1]); aligned_b.append(b[j - 1]); i -= 1; j -= 1
        elif i > 0 and np.isclose(score[i, j], score[i - 1, j] - gap_penalty):
            aligned_a.append(a[i - 1]); aligned_b.append("-"); i -= 1
        else:
            aligned_a.append("-"); aligned_b.append(b[j - 1]); j -= 1
    return list(reversed(aligned_a)), list(reversed(aligned_b))


def compare_to_reference(read_calls: Sequence[str], reference: Sequence[str]) -> Dict:
    """Step 6.6: align a read's nucleotide calls against a reference and
    report mismatches as candidate modification / isoform sites.
    """
    aligned_read, aligned_ref = _global_align(list(read_calls), list(reference))
    mismatches = [
        {"position": i, "read": r, "reference": ref}
        for i, (r, ref) in enumerate(zip(aligned_read, aligned_ref))
        if r != ref
    ]
    identity = 1.0 - len(mismatches) / max(len(aligned_ref), 1)
    return {
        "aligned_read": aligned_read,
        "aligned_reference": aligned_ref,
        "mismatches": mismatches,
        "identity": identity,
    }


# ---------------------------------------------------------------------------
# Step 6.7: assemble overlapping short reads into longer contigs
# ---------------------------------------------------------------------------

def assemble_contigs(chains: List[Dict], min_overlap: int = 2) -> List[List[int]]:
    """Greedy overlap merge: two reads merge if a suffix of one chain's
    nucleotide calls matches a prefix of another's for >= min_overlap
    residues. Returns lists of chain-indices (into `chains`) that were
    merged together, in merge order. This is a simple heuristic for
    relating short reads (Step 6.7), not a full assembler.
    """
    seqs = [c["sequence_calls"] for c in chains]
    n = len(seqs)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            x = parent[x]
        return x

    def union(x, y):
        parent[find(x)] = find(y)

    for i in range(n):
        for j in range(n):
            if i == j:
                continue
            si, sj = seqs[i], seqs[j]
            max_k = min(len(si), len(sj))
            for k in range(max_k, min_overlap - 1, -1):
                if si[-k:] == sj[:k]:
                    union(i, j)
                    break

    groups: Dict[int, List[int]] = {}
    for i in range(n):
        groups.setdefault(find(i), []).append(i)
    return list(groups.values())


# ---------------------------------------------------------------------------
# Jiang method compliance audit -- machine-readable counterpart to
# Jiang_Method_Audit.md. Written into every run's base_calling_report.json
# as `method_compliance` so a script (or the dashboard) can check which
# steps ran without parsing the markdown audit doc.
# ---------------------------------------------------------------------------

def build_method_compliance(cfg: "Config") -> Dict:
    """Per-step mapping of Dr. Jiang's nested algorithm to this
    implementation. `implemented` is one of "yes" / "partial" /
    "reference-dependent". Step 6.6's status is computed from `cfg` since
    it is the only step whose implemented-ness is run-specific (it is a
    deliberate no-op without a supplied reference, not a missing feature).
    See Jiang_Method_Audit.md for the full prose version with worked
    justification for every limitation listed here.
    """
    step_6_6_implemented = "yes" if cfg.reference is not None else "reference-dependent (not supplied this run)"

    return {
        "method": "Dr. Jiang's Nested Algorithm for tRNA mixtures using block-wise relative intensity",
        "audit_doc": "Jiang_Method_Audit.md",
        "steps": {
            "step_1_input_columns": {
                "description": "Read M/I/T input columns",
                "implemented": "yes",
                "proof": ["load_data()", "annotated_data.csv columns: M, I, T"],
                "limitations": "Column identification is keyword-based (_resolve_column); an export with unrecognized headers fails to load rather than guessing.",
            },
            "step_2_mass_ascending_sort": {
                "description": "Mass-ascending sort",
                "implemented": "yes",
                "proof": ["load_data(): df.sort_values('M', ascending=True)", "annotated_data.csv row order"],
                "limitations": "None -- unconditional sort, no tunable parameter.",
            },
            "step_3_block_definition": {
                "description": "Block definition, k range, block width",
                "implemented": "yes",
                "proof": ["BLOCK_WIDTH_DA=320.0, N_BLOCKS=95 (k=1..95)", "assign_blocks()", "annotated_data.csv['block']", "report['n_blocks']"],
                "limitations": "320.0 Da is a fixed approximation of average single-nucleotide residue mass, not derived per-run. N_BLOCKS=95 is a fixed ceiling, not computed from the input's mass span.",
            },
            "step_4_linear_rel_i": {
                "description": "Linear Rel.I = I / block max I",
                "implemented": "yes",
                "proof": ["compute_relative_intensity()", "annotated_data.csv['Rel_I']"],
                "limitations": "None for the computation itself; see method_notes.linear_rel_i_for_selection for how it's used.",
            },
            "step_5_visualization": {
                "description": "Rel.I vs M visualization",
                "implemented": "yes",
                "proof": ["plot_relative_intensity()", "relative_intensity.png", "report['plot']"],
                "limitations": "Static PNG only; the dashboard's chart is a separate re-rendering of the same Rel_I/M/block columns.",
            },
            "step_6_1_seed_order": {
                "description": "Seed order by highest unflagged Rel.I",
                "implemented": "yes",
                "proof": ["run_base_calling() Phase 1: intensity_order = df.sort_values('Rel_I', ascending=False)", "chains[i].read_rank ordering"],
                "limitations": "None -- unconditional global sort, no tunable parameter.",
            },
            "step_6_2_mass_filter_rt_trend": {
                "description": "Mass hard filter -> highest Rel.I -> RT trend check",
                "implemented": "yes",
                "proof": ["_find_candidates()", "_select_next()", "Config.mass_tol/rt_zscore/rt_window", "report['gap_mode']/['max_residues_per_step_used']"],
                "limitations": "mass_tol and rt_zscore are tunable thresholds standing in for 'mass hard filter' and 'RT trend check', not literal values from the source document. rt_std_floor prevents a degenerate near-zero RT std from blowing up the z-score.",
            },
            "step_6_3_stop_condition": {
                "description": "Stop when no candidate satisfies criteria",
                "implemented": "yes",
                "proof": ["_select_next() returns None -> _extend_direction() breaks", "chains[i].n_points variability"],
                "limitations": "Config.min_chain_len additionally drops length-1 chains from the report (still marked used per 6.5); a reporting threshold, not part of the stopping rule itself.",
            },
            "step_6_4_paired_read_classification": {
                "description": "Paired-read 5'/3' classification using decimal part and RT evidence, with confidence/warnings",
                "implemented": "yes",
                "proof": ["classify_ladders()", "find_best_pairs()", "chains[i].ladder_type/ladder_confidence/ladder_evidence/ladder_warnings", "read_summary.csv & Classification_Evidence.csv evidence columns"],
                "limitations": "Config.min_pair_score (acceptable-partner cutoff) and the length_factor-weighted confidence formula are implementation heuristics, not literal values from the source method. The across-chain seed-decimal median is never the determining rule (see method_notes).",
            },
            "step_6_5_flag_and_exclude": {
                "description": "Flag used peaks and avoid them in the next primary read",
                "implemented": "yes",
                "proof": ["_extend_direction() sets used[idx]=True", "_find_candidates() excludes used points", "annotated_data.csv/Peak_Status.csv['peak_status']", "chains_overlay.png"],
                "limitations": "None for primary reads -- flagging is unconditional and global. The one intentional exception is Step 6.6 reuse (see method_notes).",
            },
            "step_6_6_reference_guided_reuse": {
                "description": "Reference-dependent second-stage comparison/reuse, only if reference is supplied",
                "implemented": step_6_6_implemented,
                "proof": ["apply_reference_guided_extension()", "review_summary.reference_provided", "reference_comparisons (present only if reference supplied)", "Reference_Reuse_Audit.xlsx (present only if reference supplied)"],
                "limitations": "Only exercised on runs that pass a reference; this is by design (deliberate no-op), not a missing feature.",
            },
            "step_6_7_top_parallel_reads": {
                "description": "Generate many short reads in Rel.I order and report their relationships/top parallel reads",
                "implemented": "yes",
                "proof": ["build_top_parallel_reads_grid()", "build_top_parallel_reads_long()", "assemble_contigs()", "Top_Parallel_Reads.csv/.xlsx", "top_parallel_reads_long.csv", "report['contig_groups']"],
                "limitations": "assemble_contigs(min_overlap=2)'s overlap threshold is a tunable heuristic, not a literal value from the source document.",
            },
        },
        "method_notes": {
            "linear_rel_i_for_selection": "Every selection decision in Step 6 (seeding, candidate ranking, pair scoring) reads Rel_I exclusively; Config.intensity_scale only controls which column the plots draw.",
            "log_intensity_qc_display_only": "log10_intensity_qc is computed alongside Rel_I for visual inspection only; no base-calling function reads this column.",
            "step_6_4_evidence_based_not_ground_truth": "5'/3' classification is evidence-based (agreeing partner -> 5'/3', disagreeing -> conflict, no partner -> ambiguous, explicitly not a guess), not an absolute ground truth.",
            "step_6_6_reference_dependent": "Reference-guided reuse only runs when cfg.reference is supplied; every chain still gets an empty reuse_evidence list otherwise for schema uniformity.",
            "no_peak_reuse_in_primary_reads": "Primary reads (Phase 1) never reuse a flagged point; reference-guided reuse (Phase 2) runs strictly afterward and only extends existing read ends, enforced by run_base_calling()'s three-phase structure.",
        },
    }


# ---------------------------------------------------------------------------
# End-to-end pipeline
# ---------------------------------------------------------------------------

def run_pipeline(
    input_path: Optional[str] = None,
    df: Optional[pd.DataFrame] = None,
    mass_col: str = "Monoisotopic Mass",
    intensity_col: str = "Sum Intensity",
    rt_col: str = "Apex RT",
    sheet_name=0,
    out_dir: str = ".",
    reference: Optional[Sequence[str]] = None,
    cfg: Optional[Config] = None,
) -> Dict:
    run_start_time = time.time()  # dashboard data-contract Task 1: runtime_seconds
    # dashboard data-contract Task 1: file_name. run_demo/in-memory callers
    # pass df directly with no path, so that case gets an honest label
    # rather than a fabricated filename.
    file_name = Path(input_path).name if input_path is not None else "in-memory dataframe (no input file path provided)"

    cfg = cfg or Config()
    if reference is not None:
        cfg.reference = list(reference)  # also drives active reuse in Step 6.6, not just post-hoc comparison
    out_dir_p = Path(out_dir)
    out_dir_p.mkdir(parents=True, exist_ok=True)

    if df is None:
        if input_path is None:
            raise ValueError("Provide either input_path or df.")
        df = load_data(input_path, mass_col, intensity_col, rt_col, sheet_name)  # Steps 1-2
    else:
        df = df.sort_values("M", ignore_index=True)

    df = assign_blocks(df, cfg.block_width, cfg.n_blocks)        # Step 3
    df = compute_relative_intensity(df)                          # Step 4 (always computes both Rel_I and log10_intensity_qc)

    # Display-only switch: which column the plots draw. Selection logic
    # inside run_base_calling always uses the linear Rel_I column and never
    # consults this setting.
    plot_col = "Rel_I" if cfg.intensity_scale == "linear" else "log10_intensity_qc"
    rel_i_label = "Relative Intensity (linear)" if cfg.intensity_scale == "linear" else "Relative Intensity (log10-normalized, display only)"
    plot_path = plot_relative_intensity(df, str(out_dir_p / "relative_intensity.png"), y_col=plot_col, ylabel=rel_i_label)  # Step 5

    chains, reuse_audit_log = run_base_calling(df, cfg)            # Step 6 (Phases 1-3)

    overlay_path = None
    if chains:
        overlay_path = plot_chains_overlay(
            df, chains, str(out_dir_p / "chains_overlay.png"), y_col=plot_col, ylabel=rel_i_label
        )  # Step 6.5

    report = {
        "n_points": len(df),
        "n_chains": len(chains),
        "plot": plot_path,
        "chains_overlay_plot": overlay_path,
        "gap_mode": "exploratory" if cfg.exploratory_mode else "jiang_strict",
        "max_residues_per_step_used": cfg.effective_max_residues_per_step,
        "max_missing_positions_used": cfg.max_missing_positions,
        "chains": [
            {
                "ladder_type": c["ladder_type"],
                "ladder_confidence": c.get("ladder_confidence"),
                "ladder_evidence": c.get("ladder_evidence"),
                "ladder_warnings": c.get("ladder_warnings"),
                "paired_with": c.get("paired_with"),
                "pair_score": c.get("pair_score"),
                "n_points": len(c["indices"]),
                "n_reused": len(c.get("reuse_evidence", [])),
                "reuse_evidence": c.get("reuse_evidence", []),
                "mass_range": [round(min(c["masses"]), 4), round(max(c["masses"]), 4)],
                "sequence_calls": c["sequence_calls"],
            }
            for c in chains
        ],
    }

    if cfg.reference is not None:
        report["reference_comparisons"] = [
            {"chain_index": i, **compare_to_reference(c["sequence_calls"], cfg.reference)}
            for i, c in enumerate(chains)
        ]

    contig_groups = assemble_contigs(chains) if chains else []
    report["contig_groups"] = contig_groups

    # Step 6.4 manual-review table: one row per read, full paired-read
    # classification evidence (Task A field set) plus the review-and-
    # correction pass's human-facing labels and review tiers.
    read_summary = build_read_summary(df, chains, cfg)
    read_summary_csv = out_dir_p / "read_summary.csv"
    read_summary.to_csv(read_summary_csv, index=False)
    report["read_summary_csv"] = str(read_summary_csv)

    # "Top_Parallel_Reads": top-4 highest-intensity short reads side by
    # side, eye-readable layout mirroring Dr. Jiang's chart (Task C).
    top_grid = build_top_parallel_reads_grid(df, chains, n=TOP_PARALLEL_N)
    top_csv = out_dir_p / "Top_Parallel_Reads.csv"
    write_grid_to_csv(top_grid, str(top_csv))
    report["Top_Parallel_Reads_csv"] = str(top_csv)
    top_xlsx_path = write_grid_to_xlsx(top_grid, str(out_dir_p / "Top_Parallel_Reads.xlsx"))
    if top_xlsx_path:
        report["Top_Parallel_Reads_xlsx"] = top_xlsx_path
    else:
        report["Top_Parallel_Reads_xlsx_warning"] = (
            "openpyxl not installed -- skipped .xlsx export, Top_Parallel_Reads.csv is still available."
        )

    # Sequencing-assist dashboard data contract, Task 2: long-format
    # counterpart to the side-by-side grid above -- the dashboard reads
    # this, not the grid's merged header text.
    top_parallel_long = build_top_parallel_reads_long(df, chains, read_summary, n=TOP_PARALLEL_N)
    top_parallel_long_csv = out_dir_p / "top_parallel_reads_long.csv"
    top_parallel_long.to_csv(top_parallel_long_csv, index=False)
    report["top_parallel_reads_long_csv"] = str(top_parallel_long_csv)

    # Sequencing-assist dashboard data contract, Task 3: the dashboard's
    # primary decision-support table -- one row per read with a plain
    # suggested_action, not just raw evidence columns.
    decision_summary = build_sequencing_decision_summary(read_summary)
    decision_summary_csv = out_dir_p / "sequencing_decision_summary.csv"
    decision_summary.to_csv(decision_summary_csv, index=False)
    report["sequencing_decision_summary_csv"] = str(decision_summary_csv)

    # ---- Pre-dashboard review package (review-and-correction pass req. 5) ----
    # Read_Summary_Review.xlsx, Peak_Status.xlsx, Classification_Evidence.xlsx,
    # and (only when a reference was actually provided) Reference_Reuse_Audit.xlsx.
    classification_evidence = build_classification_evidence_table(read_summary)
    peak_status_table = build_peak_status_table(df, chains)
    reuse_audit_table = build_reference_reuse_audit_table(reuse_audit_log)

    # Sequencing-assist dashboard data contract: Classification_Evidence and
    # Peak_Status were previously Excel-only, which would force the
    # dashboard to parse .xlsx client-side. Add CSV counterparts so every
    # dashboard panel has a CSV/JSON source, matching the rule already
    # applied to read_summary/top_parallel/decision_summary -- Excel stays
    # the human-review export, CSV/JSON is what the dashboard reads.
    classification_evidence_csv = out_dir_p / "Classification_Evidence.csv"
    classification_evidence.to_csv(classification_evidence_csv, index=False)
    report["classification_evidence_csv"] = str(classification_evidence_csv)

    peak_status_csv = out_dir_p / "Peak_Status.csv"
    peak_status_table.to_csv(peak_status_csv, index=False)
    report["peak_status_csv"] = str(peak_status_csv)

    def _write_review_xlsx(filename: str, sheets: Dict[str, pd.DataFrame]) -> Optional[str]:
        path = write_dataframes_to_xlsx(sheets, str(out_dir_p / filename))
        key = filename.rsplit(".", 1)[0]
        if path:
            report[f"{key}_xlsx"] = path
        else:
            report[f"{key}_xlsx_warning"] = f"openpyxl not installed -- skipped {filename}."
        return path

    _write_review_xlsx("Read_Summary_Review.xlsx", {"Read_Summary_Review": read_summary})
    _write_review_xlsx("Peak_Status.xlsx", {"Peak_Status": peak_status_table})
    _write_review_xlsx("Classification_Evidence.xlsx", {"Classification_Evidence": classification_evidence})
    if cfg.reference is not None:
        # No "Summary" sheet here on purpose: aggregate counts belong in the
        # web dashboard / report["review_summary"] below, not duplicated
        # inside the Excel review package. Excel stays a row-level data
        # package -- Candidates is every tested reuse point, Accepted/
        # Rejected are the same rows split by outcome for quick filtering.
        _write_review_xlsx("Reference_Reuse_Audit.xlsx", {
            "Candidates": reuse_audit_table,
            "Accepted": reuse_audit_table[reuse_audit_table["outcome"] == "accepted"] if not reuse_audit_table.empty else reuse_audit_table,
            "Rejected": reuse_audit_table[reuse_audit_table["outcome"] == "rejected"] if not reuse_audit_table.empty else reuse_audit_table,
        })

    # Review-and-correction pass requirement 6: the exact metrics the
    # reviewer needs to sanity-check this run at a glance, without opening
    # any of the files above.
    call_counts = read_summary["ladder_call"].value_counts().to_dict() if not read_summary.empty else {}
    report["review_summary"] = {
        "reads_5prime": call_counts.get("5'", 0),
        "reads_3prime": call_counts.get("3'", 0),
        "reads_ambiguous": call_counts.get("ambiguous", 0),
        "reads_conflict": call_counts.get("conflict", 0),
        "primary_used_peaks": int((df["peak_status"] == "primary_used").sum()) if "peak_status" in df.columns else None,
        "reference_provided": cfg.reference is not None,
        "reference_reuse_candidates_tested": len(reuse_audit_table),
        "reference_reuse_accepted": int((reuse_audit_table["outcome"] == "accepted").sum()) if not reuse_audit_table.empty else 0,
        "reference_reuse_rejected": int((reuse_audit_table["outcome"] == "rejected").sum()) if not reuse_audit_table.empty else 0,
    }

    # Sequencing-assist dashboard data contract, Task 1: run metadata that
    # was previously missing from the report entirely (file identity,
    # timing, block count) plus the full peak_status / read-call
    # breakdowns (review_summary above only ever had primary_used_peaks).
    # Added as new top-level fields, additive -- nothing that already
    # reads `review_summary` has to change.
    report["file_name"] = file_name
    report["runtime_seconds"] = round(time.time() - run_start_time, 3)
    report["n_blocks"] = int(df["block"].nunique()) if "block" in df.columns else None

    _peak_status_keys = ["primary_used", "ambiguous_retained", "conflict_retained", "reference_reused", "unused"]
    if "peak_status" in df.columns:
        _vc = df["peak_status"].value_counts()
        report["peak_status_counts"] = {k: int(_vc.get(k, 0)) for k in _peak_status_keys}
    else:
        report["peak_status_counts"] = {k: 0 for k in _peak_status_keys}

    report["read_call_counts"] = {
        "5prime": call_counts.get("5'", 0),
        "3prime": call_counts.get("3'", 0),
        "ambiguous": call_counts.get("ambiguous", 0),
        "conflict": call_counts.get("conflict", 0),
    }

    # "At least 3 of the top 4 reads are ambiguous/conflict" -- the same
    # condition DESIGN.md's Run Overview warning banner is built around,
    # now computed once here so the dashboard doesn't have to re-derive it.
    if not read_summary.empty:
        _top4 = read_summary[read_summary["read_rank"] <= TOP_PARALLEL_N]
        _n_top4_problem = int(_top4["ladder_call"].isin(["ambiguous", "conflict"]).sum())
    else:
        _n_top4_problem = 0
    report["top_parallel_warning"] = _n_top4_problem >= 3

    annotated_csv = out_dir_p / "annotated_data.csv"
    df.to_csv(annotated_csv, index=False)
    report["annotated_csv"] = str(annotated_csv)

    # Jiang method compliance audit -- machine-readable counterpart to
    # Jiang_Method_Audit.md, mapping every step of the source method to
    # the code/output that implements it. See build_method_compliance().
    report["method_compliance"] = build_method_compliance(cfg)

    report_json = out_dir_p / "base_calling_report.json"
    with open(report_json, "w") as f:
        json.dump(report, f, indent=2, default=float)
    report["report_json"] = str(report_json)

    return {
        "data": df,
        "chains": chains,
        "report": report,
        "read_summary": read_summary,
        "Top_Parallel_Reads": top_grid,
        "top_parallel_reads_long": top_parallel_long,
        "decision_summary": decision_summary,
        "classification_evidence": classification_evidence,
        "peak_status_table": peak_status_table,
        "reuse_audit_table": reuse_audit_table,
    }


# ---------------------------------------------------------------------------
# Built-in demo, reproducing the worked example from the source document
# ---------------------------------------------------------------------------

def _demo_dataframe(overlapping: bool = False) -> pd.DataFrame:
    """The 4 worked-example ladders ('low mass region') from the document,
    flattened into long (M, I, T) form with synthetic but plausible I and T
    values (the original doc only tabulates M and Rel_I for this example).

    By default the 4 ladders are shifted apart in mass (`overlapping=False`)
    so this demo is a clean illustration that the mechanics work end-to-end.
    Pass `overlapping=True` to reproduce the document's literal example,
    where all 4 ladders occupy the *same* low-mass window -- this is
    explicitly the hardest case (per Step 6.1, starting from a crowded
    region) because the very first hop of each chain has no RT trend yet to
    disambiguate it (see Step 6.2's "RT" criterion only working once a
    trend exists), so chains can legitimately cross between true sequences
    on that first step. That is a property of the documented heuristic
    itself, not just this implementation -- the doc's own Steps 6.6/6.7
    (reference comparison, reusing flagged points) exist to resolve exactly
    this kind of ambiguity by hand.
    """
    seqs = {
        "seq1": [976.2489, 1305.3026, 1610.3440, 1939.3960, 2268.4506, 2597.5032,
                  2942.5528, 3287.6004, 3592.6438, 3937.6910, 4282.7413, 4627.7899,
                  4932.8300, 5237.8742, 5566.9350, 5895.9849],
        "seq2": [1206.2352, 1511.2767, 1856.3242, 2162.3491, 2491.4024, 2796.4439,
                  3141.4936, 3447.5317, 3792.5729, 4137.6215, 4442.6648, 4748.6873,
                  5053.7268, 5359.7551, 5688.8042, 6031.8729, 6360.9285, 6665.9675,
                  6971.9931],
        "seq3": [1816.3185, 2122.3434, 2427.3849, 2733.4127, 3062.4644, 3367.5065,
                  3712.5554, 4041.6079, 4347.6371, 4652.6799, 4957.7238, 5263.7456,
                  5592.7997, 5935.8741, 6280.9168, 6585.9625],
        "seq4": [1625.1543, 1984.2177, 2290.2439, 2596.2695, 2939.3383, 3284.3865,
                  3590.4123, 3919.4679, 4225.4943, 4554.5473, 4899.5961, 5205.6205,
                  5550.6643, 5895.7125, 6203.7551, 6548.8088, 6877.8589],
    }
    # Each ladder gets its own well-separated RT trend (a tRNA ladder series
    # elutes along a characteristic, gradually-shifting RT band under fixed
    # LC-MS conditions) with only mild noise, so the RT criterion in Step
    # 6.2 has something real to discriminate on.
    rng = np.random.default_rng(42)
    base_rts = {"seq1": 2.0, "seq2": 5.0, "seq3": 8.0, "seq4": 11.0}
    drifts = {"seq1": 0.04, "seq2": -0.03, "seq3": 0.025, "seq4": -0.02}
    mass_offsets = (
        {"seq1": 0.0, "seq2": 0.0, "seq3": 0.0, "seq4": 0.0}
        if overlapping
        else {"seq1": 0.0, "seq2": 9000.0, "seq3": 18000.0, "seq4": 27000.0}
    )
    rows = []
    for name, masses in seqs.items():
        base_rt = base_rts[name]
        rt_drift = drifts[name]
        offset = mass_offsets[name]
        for i, m in enumerate(masses):
            intensity = 1.0e6 * (0.97 ** i) * rng.uniform(0.85, 1.0)
            rt = base_rt + rt_drift * i + rng.normal(0, 0.003)
            rows.append({"M": m + offset, "I": intensity, "T": round(rt, 4), "_truth": name})
    return pd.DataFrame(rows)


def run_demo(
    out_dir: str = "./demo_output",
    overlapping: bool = False,
    reference: Optional[Sequence[str]] = None,
    intensity_scale: str = "linear",
    exploratory_mode: bool = False,
) -> Dict:
    demo_df = _demo_dataframe(overlapping=overlapping)
    cfg = Config(
        mass_tol=0.05, rt_zscore=2.0, min_rt_history=2, rt_std_floor=0.01,
        intensity_scale=intensity_scale, exploratory_mode=exploratory_mode,
    )
    result = run_pipeline(df=demo_df[["M", "I", "T"]], out_dir=out_dir, reference=reference, cfg=cfg)

    # Compare recovered chains back to the known ground-truth groups for a
    # quick sanity check (only meaningful for this synthetic demo). NOTE:
    # this demo's 4 ladders are 4 INDEPENDENT example sequences from the
    # doc, not true 5'/3' pairs of the same molecule -- so Step 6.4's
    # pairing logic has no real pair to find here. This demo is a
    # mechanical smoke test of the pairing code path (does it run, does it
    # produce confidence/evidence/warnings), NOT a validation that real
    # biological 5'/3' pairs get correctly identified. That validation can
    # only come from a real Dr. Jiang workbook with known paired reads.
    truth_lookup = demo_df.set_index(demo_df.index)["_truth"]
    print(f"Demo: {len(demo_df)} points -> {len(result['chains'])} recovered chains "
          f"(gap_mode={result['report']['gap_mode']})")
    for i, chain in enumerate(result["chains"]):
        truths = {truth_lookup.get(idx, "?") for idx in chain["indices"]}
        print(f"  chain {i}: {len(chain['indices'])} pts, ladder={chain['ladder_type']} "
              f"(confidence={chain.get('ladder_confidence')}), truth_groups={truths}, "
              f"calls={''.join(chain['sequence_calls'])}")
    print(f"  Top_Parallel_Reads: {result['report']['Top_Parallel_Reads_csv']}")
    rs = result["report"]["review_summary"]
    print(f"  review_summary: 5'={rs['reads_5prime']} 3'={rs['reads_3prime']} "
          f"ambiguous={rs['reads_ambiguous']} conflict={rs['reads_conflict']} "
          f"primary_used_peaks={rs['primary_used_peaks']} "
          f"reuse(tested/accepted/rejected)={rs['reference_reuse_candidates_tested']}/"
          f"{rs['reference_reuse_accepted']}/{rs['reference_reuse_rejected']}")
    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", help="Path to the deconvoluted Excel file (Step 1).")
    parser.add_argument("--sheet", default=0, help="Sheet name or index (default: first sheet).")
    parser.add_argument("--mass-col", default="Monoisotopic Mass")
    parser.add_argument("--intensity-col", default="Sum Intensity")
    parser.add_argument("--rt-col", default="Apex RT")
    parser.add_argument("--out-dir", default="./trna_output")
    parser.add_argument("--reference", help="Reference sequence as a plain string of A/C/G/U letters. Used both "
                                              "to flag mismatches (Step 6.6 comparison) and to actively guide "
                                              "gap-filling / point reuse during chain extension.")
    parser.add_argument("--mass-tol", type=float, default=Config.mass_tol)
    parser.add_argument("--max-residues-per-step", type=int, default=Config.max_residues_per_step,
                         help="Jiang-strict cap: how many residues may be summed together to test one "
                              "mass-difference jump (default 3, i.e. up to 2 missing ladder rungs per jump).")
    parser.add_argument("--exploratory", action="store_true",
                         help="Opt into a looser cap (--exploratory-max-residues-per-step) instead of "
                              "the conservative Jiang-strict default. Off by default.")
    parser.add_argument("--exploratory-max-residues-per-step", type=int,
                         default=Config.exploratory_max_residues_per_step,
                         help="Residues-per-step cap used only when --exploratory is set (default 4).")
    parser.add_argument("--rt-zscore", type=float, default=Config.rt_zscore)
    parser.add_argument("--intensity-scale", choices=["linear", "log"], default=Config.intensity_scale,
                         help="Display-only: which Rel_I column the plots draw. Primary selection logic "
                              "always uses linear Rel_I regardless of this setting.")
    parser.add_argument("--min-pair-score", type=float, default=Config.min_pair_score,
                         help="Step 6.4: minimum composite score for two reads to be accepted as a 5'/3' pair.")
    parser.add_argument("--precursor-mass", type=float, default=None,
                         help="Optional intact/precursor mass (Da), used as extra Step 6.4 pairing evidence.")
    parser.add_argument("--demo", action="store_true", help="Run on the built-in worked example instead of --input.")
    parser.add_argument("--demo-hard", action="store_true",
                         help="With --demo, reproduce the document's literal low-mass-region example "
                              "where all 4 ladders overlap in mass (harder / more ambiguous case).")
    args = parser.parse_args()

    reference = list(args.reference.strip().upper()) if args.reference else None

    if args.demo:
        run_demo(args.out_dir, overlapping=args.demo_hard, reference=reference,
                  intensity_scale=args.intensity_scale, exploratory_mode=args.exploratory)
        return

    if not args.input:
        parser.error("Provide --input <file.xlsx> or use --demo.")

    cfg = Config(
        mass_tol=args.mass_tol, max_residues_per_step=args.max_residues_per_step,
        exploratory_mode=args.exploratory, exploratory_max_residues_per_step=args.exploratory_max_residues_per_step,
        rt_zscore=args.rt_zscore, intensity_scale=args.intensity_scale,
        min_pair_score=args.min_pair_score, precursor_mass=args.precursor_mass,
    )

    result = run_pipeline(
        input_path=args.input,
        mass_col=args.mass_col,
        intensity_col=args.intensity_col,
        rt_col=args.rt_col,
        sheet_name=args.sheet,
        out_dir=args.out_dir,
        reference=reference,
        cfg=cfg,
    )

    report = result["report"]
    print(f"Loaded {report['n_points']} points, recovered {report['n_chains']} chains. (gap_mode={report['gap_mode']})")
    print(f"Annotated data:        {report['annotated_csv']}")
    print(f"Read summary (6.4):    {report['read_summary_csv']}")
    print(f"Rel-I scatter:         {report['plot']}")
    print(f"Chains overlay:        {report['chains_overlay_plot']}")
    print(f"Top_Parallel_Reads:    {report['Top_Parallel_Reads_csv']}")
    if "Top_Parallel_Reads_xlsx" in report:
        print(f"Top_Parallel_Reads xl: {report['Top_Parallel_Reads_xlsx']}")
    elif "Top_Parallel_Reads_xlsx_warning" in report:
        print(f"  (note: {report['Top_Parallel_Reads_xlsx_warning']})")
    print("Pre-dashboard review package:")
    for key in ("Read_Summary_Review", "Peak_Status", "Classification_Evidence", "Reference_Reuse_Audit"):
        if f"{key}_xlsx" in report:
            print(f"  {key}.xlsx: {report[f'{key}_xlsx']}")
        elif f"{key}_xlsx_warning" in report:
            print(f"  (note: {report[f'{key}_xlsx_warning']})")
    rs = report["review_summary"]
    print(f"Review summary: 5'={rs['reads_5prime']} 3'={rs['reads_3prime']} "
          f"ambiguous={rs['reads_ambiguous']} conflict={rs['reads_conflict']} "
          f"primary_used_peaks={rs['primary_used_peaks']}")
    if rs["reference_provided"]:
        print(f"  reference reuse: {rs['reference_reuse_candidates_tested']} candidates tested, "
              f"{rs['reference_reuse_accepted']} accepted, {rs['reference_reuse_rejected']} rejected")
    else:
        print("  reference reuse: not tested (no --reference provided)")
    print(f"Full report:           {report['report_json']}")


if __name__ == "__main__":
    main()
