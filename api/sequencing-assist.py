"""
Combined single-phase Vercel Python function for RNA LC-MS Sequencing Assist.

POST /api/sequencing-assist
  form fields: file (.xlsx), reference_sequence (optional)
  returns: JSON with scatter/sigmoid preview + pipeline results + base64 Excel
"""

import base64
import io
import json
import math
import os
import sys
import tempfile
import time

import numpy as np
import pandas as pd
from flask import Flask, request, jsonify

# Add the function's own directory and its parent to sys.path.
# Vercel may place includeFiles alongside the function OR at the bundle root,
# so we search both locations for trna_nested_algorithm.
_FUNC_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_FUNC_DIR)
for _p in [_FUNC_DIR, _REPO_ROOT]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from trna_nested_algorithm import (  # noqa: E402
    load_data,
    assign_blocks,
    compute_relative_intensity,
    run_pipeline as _run_nested_pipeline,
    Config as _AlgoConfig,
    BLOCK_WIDTH_DA,
    N_BLOCKS,
    build_read_summary as _build_read_summary,
    build_top_parallel_reads_long as _build_top_parallel_long,
    compare_to_reference as _compare_to_reference,
    _flatten_labels,
)

# ── Constants ──────────────────────────────────────────────────────────────────
PRE_SUB_LIMIT = 20_000
PIPELINE_POINT_LIMIT = 8_000

# ── tRNA reference mapping ─────────────────────────────────────────────────────
# Theoretical RNA ladder mass terminus constants validated against Dr. Jiang's
# JavaScript implementation (tRNA_suite_v1_update.html).
# 5' ladder fragments (5'-phosphate terminus): mass = _MASS_START_5P + Σresidues
# 3' ladder fragments (5'-OH terminus):        mass = _MASS_START_3P + Σresidues
_MASS_START_5P = 97.9769
_MASS_START_3P = -61.9558

# Symbol-to-mass for every modification symbol found in 46_tRNA_iso.xlsx Sheet1.
# Isobaric modifications map to the same mass (indistinguishable by mass alone).
_REF_SYMBOL_MASS: dict[str, float] = {
    # Canonical
    "A": 329.05252, "C": 305.04129, "G": 345.04743, "U": 306.02530,
    # Uridine modifications
    "D": 308.04095,       # dihydrouridine
    "T": 320.04095,       # ribothymidine (5-methyluridine, m5U)
    "Um": 320.04095,      # 2'-O-methyluridine (isobaric with T/m5U)
    "m3U": 320.04095,     # 3-methyluridine (isobaric with Um)
    "m5Um": 334.05660,    # 5,2'-O-dimethyluridine
    "mnm5U": 349.06750,   # 5-methylaminomethyluridine
    "ncm5U": 363.04677,   # 5-carbamoylmethyluridine
    "mcm5U": 378.04643,   # 5-methoxycarbonylmethyluridine
    "mchm5U": 394.04135,  # 5-(carboxyhydroxymethyl)uridine methyl ester
    "acp3U": 407.07298,   # 3-(3-amino-3-carboxypropyl)uridine
    # Cytidine modifications
    "m5C": 319.05694,     # 5-methylcytidine
    "Cm": 319.05694,      # 2'-O-methylcytidine (isobaric with m5C)
    "m3C": 319.05694,     # 3-methylcytidine (isobaric with m5C)
    "ac4C": 347.05185,    # N4-acetylcytidine
    "f5C": 333.03620,     # 5-formylcytidine
    "f5Cm": 347.05185,    # 5-formyl-2'-O-methylcytidine (isobaric with ac4C)
    "hm5C": 335.05238,    # 5-hydroxymethylcytidine
    # Adenosine modifications
    "m1A": 343.06817,     # 1-methyladenosine (same mass class as mA/Am)
    "m6Am": 357.08382,    # N6,2'-O-dimethyladenosine
    "I": 330.03654,       # inosine
    "m1I": 344.05200,     # 1-methylinosine
    "i6A": 397.11512,     # N6-isopentenyladenosine
    "ms2i6A": 443.10284,  # 2-methylthio-N6-isopentenyladenosine
    "t6A": 474.09003,     # N6-threonylcarbamoyladenosine
    # Guanosine modifications
    "m1G": 359.06308,     # 1-methylguanosine
    "m2G": 359.06308,     # N2-methylguanosine (isobaric with m1G)
    "m7G": 359.06308,     # 7-methylguanosine (isobaric with m1G)
    "Gm": 359.06308,      # 2'-O-methylguanosine (isobaric with m1G)
    "m22G": 373.07873,    # N2,N2-dimethylguanosine
    "Q": 471.11551,       # queuosine
}

# Canonical path search order for the reference file (local dev vs Vercel bundle)
_TRNA_REF_SEARCH_PATHS = [
    os.path.join(_FUNC_DIR, "data", "46_tRNA_iso.xlsx"),
    os.path.join(_REPO_ROOT, "data", "46_tRNA_iso.xlsx"),
    # Developer local path (populated only on the original developer machine)
    os.path.join(
        os.path.expanduser("~"),
        "Downloads", "01_UAlbany_RNA_LCMS",
        "01_Active_Ladder_LCMS_Sequencing",
        "SUNY_from_Downloads", "TRNA", "46_tRNA_iso.xlsx",
    ),
]

# Full modification dictionary (from project reference, dictonary.csv).
# Isobaric pairs are merged under the most biologically common label; the
# decode step picks the closest mass match so both are effectively covered.
_RESIDUE_MASS: dict[str, float] = {
    # Canonical bases
    "A":          329.05252,
    "U":          306.02530,
    "G":          345.04743,
    "C":          305.04129,
    # Uridine modifications
    "D":          308.04095,  # dihydrouridine
    "Um/m1Ψ":     320.04095,  # 2'-O-methyluridine / 1-methylpseudouridine / Ψm / mU
    "s2U/s4U":    322.00246,  # 2-thiouridine / 4-thiouridine
    "mo5U":       336.03587,  # 5-methoxyuridine
    "m5s2U":      336.01811,  # 5-methyl-2-thiouridine
    "m5Um":       334.05660,  # 5,2'-O-dimethyluridine
    "mnm5U":      349.06750,  # 5-methylaminomethyluridine
    "ncm5U":      363.04677,  # 5-carbamoylmethyluridine
    "mnm5s2U":    365.04466,  # 5-methylaminomethyl-2-thiouridine
    "mcm5U":      378.04643,  # 5-methoxycarbonylmethyluridine
    "cmo5U":      380.02570,  # uridine 5-oxyacetic acid
    "cmnm5U":     393.05733,  # 5-carboxymethylaminomethyluridine
    "mcm5s2U":    394.02359,  # 5-methoxycarbonylmethyl-2-thiouridine
    "mchm5U":     394.04135,  # 5-(carboxyhydroxymethyl)uridine methyl ester
    "ncm5Um":     377.06242,  # 5-carbamoylmethyl-2'-O-methyluridine
    "acp3U":      407.07298,  # 3-(3-amino-3-carboxypropyl)uridine
    "cmnm5s2U":   409.03449,  # 5-carboxymethylaminomethyl-2-thiouridine
    "tm5U":       443.04000,  # 5-taurinomethyluridine
    "tm5s2U":     459.01710,  # 5-taurinomethyl-2-thiouridine
    # Cytidine modifications
    "s2C":        321.01844,  # 2-thiocytidine
    "m5C/Cm":     319.05694,  # 5-methylcytidine / 2'-O-methylcytidine
    "ac4C":       347.05185,  # N4-acetylcytidine / 5-formyl-2'-O-methylcytidine
    "f5C":        333.03620,  # 5-formylcytidine
    "k2C":        433.13625,  # lysidine
    # Adenosine modifications
    "I":          330.03654,  # inosine
    "m1I":        344.05200,  # 1-methylinosine
    "mA/Am":      343.06817,  # N6-methyladenosine / 2'-O-methyladenosine
    "i6A":        397.11512,  # N6-isopentenyladenosine
    "io6A":       413.11003,  # N6-(cis-hydroxyisopentenyl)adenosine
    "ms2i6A":     443.10284,  # 2-methylthio-N6-isopentenyladenosine
    "t6A":        474.09003,  # N6-threonylcarbamoyladenosine
    "m6t6A":      488.10568,  # N6-methyl-N6-threonylcarbamoyladenosine
    "ms2t6A":     520.07775,  # 2-methylthio-N6-threonylcarbamoyladenosine
    "Ar(p)":      541.06111,  # 2'-O-ribosyladenosine (phosphate)
    "yW":         469.09866,  # wybutosine (C17H21N5O7; residue = NMP − H2O = 487.110 − 18.011)
    "o2yW":       485.09395,  # peroxywybutosine (yW + O; +15.995 Da oxidation)
    # Guanosine modifications
    "mG/Gm":      359.06308,  # 7-methylguanosine / N2-methylguanosine / 2'-O-methylguanosine
    "G'":         346.05740,  # di-guanosine derivative
    "m22G":       373.07873,  # N2,N2-dimethylguanosine
    "m22Gm":      387.09438,  # N2,N2,2'-O-trimethylguanosine
    "archaeosine":386.07398,  # archaeosine (fa7d7G)
    "Q":          471.11551,  # queuosine
    "manQ/galQ":  633.16834,  # mannosyl- / galactosyl-queuosine
}
# _CHAIN_TOL: mass tolerance used by the chain-building algorithm (must match
# the mass_tol passed to _AlgoConfig below).  Used in the FDR null model so
# the null step-match rate reflects the same rule as the chain builder.
_CHAIN_TOL = 0.05
# _DECODE_TOL: slightly looser tolerance for the post-hoc sequence decoder.
# Chain building is already done; 0.07 Da gives a small margin for rounding
# when labeling a step that was already accepted at 0.05 Da.
_DECODE_TOL = 0.07

# ── tRNA reference pre-mapping helpers ────────────────────────────────────────

def _trna_sequences_from_module() -> dict[str, list[str]]:
    """Load tRNA sequences from the embedded Python module (primary source, always available)."""
    try:
        import trna_reference as _ref
        return _ref.TRNA_SEQUENCES
    except ImportError:
        return {}


def _trna_sequences_from_xlsx() -> dict[str, list[str]]:
    """Load tRNA sequences from 46_tRNA_iso.xlsx (fallback for local development)."""
    ref_path = None
    for candidate in _TRNA_REF_SEARCH_PATHS:
        if os.path.isfile(candidate):
            ref_path = candidate
            break
    if ref_path is None:
        return {}
    try:
        df_ref = pd.read_excel(ref_path, sheet_name=0, header=0, index_col=0)
    except Exception:
        return {}
    seqs: dict[str, list[str]] = {}
    for _, row in df_ref.iterrows():
        name = row.get("Anticodon", "")
        if not isinstance(name, str):
            continue
        if any(name.startswith(pfx) for pfx in ("covered", "not covered")):
            continue
        seq = [str(v) for v in row[1:] if isinstance(v, str) and str(v) in _REF_SYMBOL_MASS]
        if len(seq) >= 5:
            seqs[name] = seq
    return seqs


def _load_tRNA_theoretical_masses() -> list[float]:
    """
    Compute all theoretical 5' and 3' ladder fragment masses for the 46 human
    tRNA families using Dr. Jiang's validated terminus constants.  Returns a
    sorted list in the 2,000–23,000 Da analysis window.

    Uses the embedded trna_reference.py module (deployed to Vercel), falling
    back to the 46_tRNA_iso.xlsx file for local development.  Returns an empty
    list if neither source is available.
    """
    sequences = _trna_sequences_from_module()
    if not sequences:
        sequences = _trna_sequences_from_xlsx()
    if not sequences:
        return []

    masses: list[float] = []
    for seq in sequences.values():
        residues = [_REF_SYMBOL_MASS[s] for s in seq if s in _REF_SYMBOL_MASS]
        if len(residues) < 5:
            continue
        n = len(residues)
        # 5' ladder: START_5P + Σ residues[0:k]  for k=1..n
        cumsum = 0.0
        for r in residues:
            cumsum += r
            m = _MASS_START_5P + cumsum
            if 2_000.0 <= m <= 23_000.0:
                masses.append(m)
        # 3' ladder: START_3P + Σ residues[n-k:n]  for k=1..n (3'→5' walk)
        cumsum = 0.0
        for r in reversed(residues):
            cumsum += r
            m = _MASS_START_3P + cumsum
            if 2_000.0 <= m <= 23_000.0:
                masses.append(m)

    return sorted(set(round(m, 4) for m in masses))


def _annotate_reference_matches(
    df_pipeline: "pd.DataFrame",
    theoretical_masses: list[float],
) -> tuple["pd.DataFrame", int]:
    """
    Add a boolean column 'ref_matched' to df_pipeline.  A peak is marked matched
    if its mass is within 10 PPM (minimum 0.05 Da) of any theoretical ladder mass.
    Returns (annotated_df, n_matched).
    """
    if not theoretical_masses or len(df_pipeline) == 0:
        df_pipeline = df_pipeline.copy()
        df_pipeline["ref_matched"] = False
        return df_pipeline, 0

    theory_arr = np.array(theoretical_masses)
    data_masses = df_pipeline["M"].values
    matched = np.zeros(len(data_masses), dtype=bool)

    for i, dm in enumerate(data_masses):
        tol = max(0.05, dm * 10e-6)  # 10 PPM, minimum 0.05 Da
        lo = np.searchsorted(theory_arr, dm - tol)
        hi = np.searchsorted(theory_arr, dm + tol, side="right")
        if hi > lo:
            matched[i] = True

    df_out = df_pipeline.copy()
    df_out["ref_matched"] = matched
    return df_out, int(matched.sum())


# ── Flask app ──────────────────────────────────────────────────────────────────
app = Flask(__name__)


@app.after_request
def _add_cors(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return response


# Vercel passes the full URL path as PATH_INFO (e.g. /api/sequencing-assist),
# so register both the root and the full path to handle either mounting style.
@app.route("/", methods=["OPTIONS"])
@app.route("/api/sequencing-assist", methods=["OPTIONS"])
def _preflight():
    return ("", 204)


@app.route("/", methods=["POST"])
@app.route("/api/sequencing-assist", methods=["POST"])
def analyze():
    _t0 = time.time()
    file = request.files.get("file")
    if file is None:
        return jsonify({"detail": "No file provided"}), 400

    ref_seq = request.form.get("reference_sequence", "").strip()
    # User-adjustable pipeline parameters
    try:
        min_chain_len_param = max(3, min(int(request.form.get("min_chain_len", "10") or "10"), 50))
    except (ValueError, TypeError):
        min_chain_len_param = 10
    try:
        top_n_chains_param = max(4, min(int(request.form.get("top_n_chains", "10") or "10"), 25))
    except (ValueError, TypeError):
        top_n_chains_param = 10
    # Signal threshold: minimum Rel_I (% of block max) for a peak to be used as a chain seed
    try:
        min_rel_i_param = max(0.0, min(float(request.form.get("min_rel_i", "5") or "5"), 30.0))
    except (ValueError, TypeError):
        min_rel_i_param = 5.0
    # Intact/precursor mass for closure scoring (optional). When provided the
    # algorithm adds a +0.10 pairing-score bonus when the 5' + 3' terminal masses
    # sum to the precursor within precursor_mass_tol (default 1.0 Da).
    precursor_mass_param = None
    try:
        pm_raw = request.form.get("precursor_mass", "").strip()
        if pm_raw:
            pm_val = float(pm_raw)
            if 1_000.0 <= pm_val <= 200_000.0:
                precursor_mass_param = pm_val
    except (ValueError, TypeError):
        pass
    fname = file.filename or "upload.xlsx"
    if not fname.lower().endswith((".xlsx", ".xls")):
        return jsonify({"detail": "Expected an Excel (.xlsx) file"}), 400

    with tempfile.TemporaryDirectory() as tmpdir:
        # ── 1. Save + parse ───────────────────────────────────────────────
        fpath = os.path.join(tmpdir, fname)
        file.save(fpath)
        try:
            df = load_data(fpath)
            df = assign_blocks(df)
            df = compute_relative_intensity(df)
        except Exception as exc:
            return jsonify({"detail": f"Cannot parse Excel file: {exc}"}), 422

        n_original = len(df)

        # ── 2. Pre-subsample (Charge 1 files can be 30-50k points) ────────
        was_pre_subsampled = False
        if n_original > PRE_SUB_LIMIT:
            df = (
                df.groupby("block")
                .apply(lambda g: g.nlargest(50, "I"), include_groups=False)
                .reset_index(drop=True)
                .sort_values("M", ignore_index=True)
            )
            # pandas 3.x drops group-key column in groupby-apply; recompute.
            df["block"] = (df["M"] / BLOCK_WIDTH_DA).round().astype(int).clip(1, N_BLOCKS)
            was_pre_subsampled = True

        df_stored = df.copy()

        # ── 3. Intact detection ───────────────────────────────────────────
        data_type_warning = _detect_data_type(df)

        # ── 4. Scatter / sigmoid preview ──────────────────────────────────
        scatter_points = _build_scatter(df)
        sigmoid_points = _build_sigmoid(df)
        rt_vals = df["T"]
        rt_spread = round(float(rt_vals.max() - rt_vals.min()), 2)
        preview_rows = df.head(5)[["M", "I", "T", "block", "Rel_I"]].round(4).to_dict(orient="records")

        # ── 5. Pipeline-level subsample ───────────────────────────────────
        was_subsampled = False
        df_pipeline = df
        if len(df) > PIPELINE_POINT_LIMIT:
            df_pipeline = (
                df.groupby("block")
                .apply(lambda g: g.nlargest(25, "I"), include_groups=False)
                .reset_index(drop=True)
                .sort_values("M", ignore_index=True)
            )
            df_pipeline["block"] = (
                (df_pipeline["M"] / BLOCK_WIDTH_DA).round().astype(int).clip(1, N_BLOCKS)
            )
            was_subsampled = True

        # ── 5a. RT quality metric ─────────────────────────────────────────
        # R² of RT vs log(mass) measures how well the sigmoidal elution is
        # developed.  Well-resolved ion-pairing gradient data should have
        # R² ≥ 0.70 over the full 2–23 kDa range.
        rt_quality = None
        try:
            rq_df = df_pipeline[
                (df_pipeline["M"] >= 2000) & (df_pipeline["M"] <= 23000) &
                df_pipeline["M"].notna() & df_pipeline["T"].notna()
            ]
            if len(rq_df) >= 20:
                log_m = np.log(rq_df["M"].values)
                rt_v  = rq_df["T"].values
                lm_mean, rt_mean = log_m.mean(), rt_v.mean()
                ss_tot = float(((rt_v - rt_mean) ** 2).sum())
                beta   = float(((log_m - lm_mean) * (rt_v - rt_mean)).sum() /
                               ((log_m - lm_mean) ** 2).sum())
                alpha  = float(rt_mean - beta * lm_mean)
                ss_res = float(((rt_v - (alpha + beta * log_m)) ** 2).sum())
                r2     = max(0.0, 1.0 - ss_res / ss_tot) if ss_tot > 0 else 0.0
                rt_quality = {
                    "r2": round(r2, 3),
                    "slope": round(beta, 4),
                    "n_points": len(rq_df),
                    "grade": (
                        "excellent" if r2 >= 0.85
                        else "good" if r2 >= 0.70
                        else "marginal" if r2 >= 0.40
                        else "flat/poor"
                    ),
                }
        except Exception:
            pass

        # ── 5b. Mass-range + signal-threshold pre-filter ──────────────────
        # The hydrolysis ladder range is 2,000–23,000 Da. The 2 kDa floor
        # retains 6–7-mer fragments (short end of 5'/3' ladders) which carry
        # meaningful coverage in the 2–5 kDa bin that the meeting breakdown
        # table explicitly tracks. Above 23 kDa the algorithm "DID NOT ATTEMPT"
        # per the reference spreadsheet.
        # Only keep peaks with Rel_I ≥ min_rel_i_param% of their block max.
        HYDRO_MASS_MIN = 2_000.0
        HYDRO_MASS_MAX = 23_000.0
        df_pipeline = df_pipeline[
            (df_pipeline["M"] >= HYDRO_MASS_MIN) &
            (df_pipeline["M"] <= HYDRO_MASS_MAX)
        ].copy()
        if min_rel_i_param > 0 and len(df_pipeline) > 0:
            high_conf = df_pipeline[df_pipeline["Rel_I"] >= min_rel_i_param / 100.0]
            # Only apply the Rel_I gate if it leaves enough points for the algorithm
            if len(high_conf) >= 30:
                df_pipeline = high_conf
        if len(df_pipeline) == 0:
            # No peaks survived filtering — likely intact data; run on unfiltered mass range
            df_pipeline = df

        # ── 6. Reference tokens (user-provided single-sequence alignment) ────
        ref_tokens = None
        if ref_seq:
            toks = [c for c in ref_seq.upper().replace("T", "U") if c in "AUGC"]
            if toks:
                ref_tokens = toks

        # ── 6b. tRNA reference library pre-mapping ────────────────────────
        # Load theoretical 5'/3' ladder masses for all 46 human tRNA families
        # and annotate each peak.  Reported in the response so the user can judge
        # how much of the data is accounted for by known tRNA sequences.
        tRNA_theoretical_masses = _load_tRNA_theoretical_masses()
        n_ref_theoretical = len(tRNA_theoretical_masses)
        df_pipeline, n_ref_matched = _annotate_reference_matches(
            df_pipeline, tRNA_theoretical_masses
        )
        ref_library_loaded = n_ref_theoretical > 0

        # ── 7. Run pipeline ───────────────────────────────────────────────
        out_dir = os.path.join(tmpdir, "output")
        os.makedirs(out_dir, exist_ok=True)
        try:
            # rt_std_floor=0.10: per-step RT tolerance ±0.25 min.
            # The previous value of 0.30 (±0.75 min/step) allowed 40-step chains
            # to drift 15+ min across the full HPLC run, producing false cross-tRNA
            # connections in bulk data.  0.10 still tolerates natural ladder RT
            # variation (typically 0.05–0.20 min/step) while preventing artifacts.
            # max_residues_per_step=3 is Dr. Jiang's strict default; never set to 1.
            algo_cfg = _AlgoConfig(
                mass_tol=_CHAIN_TOL,
                rt_std_floor=0.10,
                min_rt_history=3,
                precursor_mass=precursor_mass_param,
                allowed_masses=dict(_RESIDUE_MASS),
                max_residues_per_step=3,
            )
            result = _run_nested_pipeline(
                df=df_pipeline, out_dir=out_dir, reference=ref_tokens, cfg=algo_cfg
            )
        except Exception as exc:
            return jsonify({"detail": f"Pipeline failed: {exc}"}), 500

        report = result.get("report", {})
        data_df = result["data"]
        chains_list = result["chains"]

        # ── 8. Chain overlay data ─────────────────────────────────────────
        n_chains_total = len(chains_list)
        n_chains_min_10 = sum(1 for c in chains_list if len(c["indices"]) >= min_chain_len_param)
        min_len = min_chain_len_param

        top_chains = []
        for chain_idx, chain in sorted(
            enumerate(chains_list),
            key=lambda x: len(x[1]["indices"]),
            reverse=True,
        )[:top_n_chains_param]:
            if len(chain["indices"]) < min_len:
                continue
            rows = data_df.loc[chain["indices"]].sort_values("M")
            for _, row in rows.iterrows():
                top_chains.append({
                    "chain_index": chain_idx,
                    "read_rank": int(chain.get("read_rank", chain_idx + 1)),
                    "ladder_type": chain["ladder_type"],
                    "mass": float(row["M"]),
                    "rel_i": float(row["Rel_I"]),
                    "rt": float(row["T"]),
                    "n_points": len(chain["indices"]),
                })

        # ── 9. Coverage by intensity threshold ────────────────────────────
        # Each bin answers: "of peaks with I ≥ X% of max, what fraction are
        # explained by candidate reads?" — matches the breakdown-table format.
        max_I_global = float(data_df["I"].max()) if len(data_df) > 0 else 1.0
        _used_status = {
            "primary_used", "ambiguous_retained",
            "conflict_retained", "reference_reused",
        }
        is_used_series = (
            data_df["peak_status"].isin(_used_status)
            if "peak_status" in data_df.columns
            else pd.Series(False, index=data_df.index)
        )
        coverage_bins = []
        for threshold, label in [
            (0.10, ">10% of max intensity"),
            (0.05, ">5% of max intensity"),
            (0.02, ">2% of max intensity"),
            (0.00, "All peaks"),
        ]:
            mask = (
                data_df["I"] >= threshold * max_I_global
                if threshold > 0
                else pd.Series(True, index=data_df.index)
            )
            total = int(mask.sum())
            matched = int((mask & is_used_series).sum())
            coverage_bins.append({
                "label": label,
                "total": total,
                "matched": matched,
                "pct": round(matched / total * 100, 1) if total else 0,
            })

        # Coverage by mass range — matches the meeting breakdown table:
        # for each mass bin (2-5K, 5-10K, 10-15K, 15-23K), count peaks above
        # each intensity threshold and how many are explained by chains.
        # Thresholds are relative to the global max across the full 2-23K range.
        coverage_by_mass_range = []
        for mass_lo, mass_hi, mass_label in [
            (2_000, 5_000, "2–5 kDa"),
            (5_000, 10_000, "5–10 kDa"),
            (10_000, 15_000, "10–15 kDa"),
            (15_000, 23_000, "15–23 kDa"),
        ]:
            sub = data_df[(data_df["M"] >= mass_lo) & (data_df["M"] < mass_hi)]
            if len(sub) == 0:
                continue
            sub_used = (
                sub["peak_status"].isin(_used_status)
                if "peak_status" in sub.columns
                else pd.Series(False, index=sub.index)
            )
            thresholds_row = []
            for pct, lbl in [(0.10, ">10%"), (0.05, ">5%"), (0.02, ">2%")]:
                mask_sub = sub["I"] >= pct * max_I_global
                total_sub = int(mask_sub.sum())
                matched_sub = int((mask_sub & sub_used).sum())
                thresholds_row.append({
                    "threshold": lbl,
                    "total": total_sub,
                    "matched": matched_sub,
                    "pct": round(matched_sub / total_sub * 100, 1) if total_sub else 0,
                })
            coverage_by_mass_range.append({
                "mass_range": mass_label,
                "thresholds": thresholds_row,
            })

        # ── 9b. Empirical FDR estimate ────────────────────────────────────
        try:
            empirical_fdr = _compute_step_match_null(df_pipeline)
        except Exception:
            empirical_fdr = None

        # ── 10. Sigmoid post-pipeline ─────────────────────────────────────
        sp_df = data_df[["M", "T", "Rel_I"]].copy()
        sp_df["status"] = (
            data_df["peak_status"]
            if "peak_status" in data_df.columns
            else "unknown"
        )
        if len(sp_df) > 8000:
            high = sp_df[sp_df["Rel_I"] >= 0.3]
            low_pool = sp_df[sp_df["Rel_I"] < 0.3]
            low = low_pool.sample(
                n=min(8000 - len(high), len(low_pool)), random_state=42
            )
            sp_df = pd.concat([high, low])
        sigmoid_post = sp_df.round(4).to_dict(orient="records")

        # ── 11. Read CSV outputs ──────────────────────────────────────────
        def _read_csv(key):
            path = report.get(key)
            if not path or not os.path.exists(path):
                return None
            try:
                df_tmp = pd.read_csv(path)
                df_tmp = df_tmp.where(df_tmp.notna(), None)
                return _sanitize(df_tmp.to_dict(orient="records"))
            except Exception:
                return None

        sequencing_decision = _read_csv("sequencing_decision_summary_csv")
        classification_ev = _read_csv("classification_evidence_csv")
        peak_status_data = _read_csv("peak_status_csv")

        # Build read_summary and top_parallel directly from algorithm objects
        # so we avoid the TOP_PARALLEL_N=4 CSV ceiling and can include partner
        # chains that fall outside the user's top_n_chains_param window.
        try:
            read_summary_df = _build_read_summary(data_df, chains_list, algo_cfg)
            read_summary_data = _sanitize(
                read_summary_df.where(read_summary_df.notna(), None).to_dict(orient="records")
            )
        except Exception:
            read_summary_data = _read_csv("read_summary_csv")
            read_summary_df = pd.DataFrame(read_summary_data or [])

        try:
            top_par_df = _build_top_parallel_long(
                data_df, chains_list, read_summary_df, n=top_n_chains_param
            )
            # Add partner chains not already in the top N so SequenceAssembly
            # can always display the true paired 5′/3′ chains together.
            top_ranks = {int(r) for r in top_par_df["read_rank"].dropna().unique()}
            partner_ranks_needed: set[int] = set()
            for row in top_par_df.drop_duplicates("read_rank").to_dict("records"):
                pr = row.get("candidate_partner_rank")
                if pr is not None and not pd.isna(pr):
                    pr_int = int(pr)
                    if pr_int not in top_ranks:
                        partner_ranks_needed.add(pr_int)
            if partner_ranks_needed:
                partner_chains = [
                    c for c in chains_list if c.get("read_rank") in partner_ranks_needed
                ]
                if partner_chains:
                    extra_df = _build_top_parallel_long(
                        data_df, partner_chains, read_summary_df, n=len(partner_chains)
                    )
                    top_par_df = pd.concat([top_par_df, extra_df], ignore_index=True)
            top_parallel = _sanitize(
                top_par_df.where(top_par_df.notna(), None).to_dict(orient="records")
            )
        except Exception:
            top_parallel = _read_csv("top_parallel_reads_long_csv")

        # ── 11c. Reference comparisons for top reads ─────────────────────
        # When the user provided a reference, the algorithm aligns every
        # chain against it.  Extract per-rank comparison data only for
        # reads that appear in the top_parallel table so the payload stays
        # small (342 full alignments would be expensive on large runs).
        top_read_ranks = {
            int(r["read_rank"]) for r in (top_parallel or [])
            if r.get("read_rank") is not None
        }
        ref_comparison_map: dict = {}
        if ref_tokens:
            for comp in report.get("reference_comparisons", []) or []:
                ci = comp.get("chain_index")
                if ci is not None and 0 <= ci < len(chains_list):
                    chain_meta = chains_list[ci]
                    rk = chain_meta.get("read_rank")
                    if rk is not None and int(rk) in top_read_ranks:
                        # Use _flatten_labels so gap steps ("A/G/C" in sequence_calls)
                        # are expanded into individual residues ["A","G","C"] before
                        # global alignment — otherwise "/" tokens never match the reference.
                        # Isobaric pairs like ("Um/m1Ψ",) remain single tokens (correct).
                        flat_calls = _flatten_labels(chain_meta.get("edge_labels", []))
                        # 3' chains walk 3'→5'; reverse before aligning against the
                        # 5'→3' reference so identity and mismatch positions are correct.
                        ladder_type = str(chain_meta.get("ladder_type", "")).lower()
                        is_3prime = "3prime" in ladder_type
                        if is_3prime:
                            flat_calls = list(reversed(flat_calls))
                        corrected = _compare_to_reference(flat_calls, ref_tokens)
                        ref_comparison_map[str(int(rk))] = _sanitize({
                            "aligned_read": list(corrected.get("aligned_read", [])),
                            "aligned_reference": list(corrected.get("aligned_reference", [])),
                            "mismatches": list(corrected.get("mismatches", [])),
                            "identity": round(float(corrected.get("identity", 0.0)), 4),
                            "orientation_corrected": is_3prime,
                        })

        # ── 11b. Build report summary + modification counts ──────────────
        # Derive read-call and peak-usage counts from the CSV data so the
        # RunOverview component has structured stats without needing a separate
        # base_calling_report.json file.
        read_call_counts = {"5prime": 0, "3prime": 0, "ambiguous": 0, "conflict": 0}
        peak_status_counts = {
            "primary_used": 0, "ambiguous_retained": 0,
            "conflict_retained": 0, "reference_reused": 0, "unused": 0,
        }
        if read_summary_data:
            for row in read_summary_data:
                call = str(row.get("ladder_call", "")).lower()
                if "5" in call:
                    read_call_counts["5prime"] += 1
                elif "3" in call:
                    read_call_counts["3prime"] += 1
                elif "ambig" in call:
                    read_call_counts["ambiguous"] += 1
                else:
                    read_call_counts["conflict"] += 1

        if peak_status_data:
            for row in peak_status_data:
                s = str(row.get("peak_status", "unused"))
                if s in peak_status_counts:
                    peak_status_counts[s] += 1

        used_peaks = peak_status_counts["primary_used"]
        ambig_conflict = (
            peak_status_counts["ambiguous_retained"] + peak_status_counts["conflict_retained"]
        )
        n_pipeline_total = len(df_pipeline)
        ref_match_pct = round(n_ref_matched / n_pipeline_total * 100, 1) if n_pipeline_total else 0.0
        report_summary = {
            "file_name": fname,
            "runtime_seconds": round(time.time() - _t0, 1),
            "n_points": n_pipeline_total,
            "n_chains": n_chains_total,
            "n_blocks": int(df_pipeline["block"].nunique()),
            "read_call_counts": read_call_counts,
            "peak_status_counts": peak_status_counts,
            "top_parallel_warning": ambig_conflict > max(used_peaks, 1),
            "tRNA_ref_library": {
                "loaded": ref_library_loaded,
                "n_theoretical_masses": n_ref_theoretical,
                "n_peaks_matched": n_ref_matched,
                "n_peaks_total": n_pipeline_total,
                "match_pct": ref_match_pct,
            },
            "review_summary": {
                "reads_5prime": read_call_counts["5prime"],
                "reads_3prime": read_call_counts["3prime"],
                "reads_ambiguous": read_call_counts["ambiguous"],
                "reads_conflict": read_call_counts["conflict"],
                "primary_used_peaks": used_peaks,
                "reference_provided": ref_tokens is not None,
                "reference_reuse_candidates_tested": 0,
                "reference_reuse_accepted": 0,
                "reference_reuse_rejected": 0,
            },
        }

        # Modification counts: decode all qualified reads and tally residue types.
        # Returned as a sorted list for the modification frequency panel.
        mod_counter_json: dict = {}
        if peak_status_data and read_summary_data:
            try:
                ps_w = pd.DataFrame(peak_status_data)
                ps_w = ps_w[ps_w["read_rank"].notna()].copy()
                ps_w["read_rank"] = ps_w["read_rank"].astype(int)
                m_by_rk = {
                    int(rk): grp["mass"].tolist()
                    for rk, grp in ps_w.groupby("read_rank")
                }
                rs_df = pd.DataFrame(read_summary_data)
                for _, row in rs_df.iterrows():
                    rk = int(row.get("read_rank", 0) or 0)
                    length = int(row.get("read_length", 0) or 0)
                    if length < min_len:
                        continue
                    masses = sorted(m for m in m_by_rk.get(rk, []) if m)
                    if len(masses) < 2:
                        continue
                    for nt in _decode_sequence(masses).split("-"):
                        mod_counter_json[nt] = mod_counter_json.get(nt, 0) + 1
            except Exception:
                pass

        canonical_set = {"A", "U", "G", "C"}
        mod_counts_response = [
            {"nt": nt, "count": count, "is_canonical": nt in canonical_set,
             "is_unknown": nt.startswith("?")}
            for nt, count in sorted(mod_counter_json.items(), key=lambda x: -x[1])
        ]

        # ── 12. Build Excel ───────────────────────────────────────────────
        sidecar = {
            "coverage_by_intensity": coverage_bins,
            "coverage_by_mass_range": coverage_by_mass_range,
            "reference_sequence": "".join(ref_tokens) if ref_tokens else None,
            "n_pipeline_points": len(df_pipeline),
            "n_original_points": n_original,
            "was_subsampled": was_subsampled,
            "n_chains_total": n_chains_total,
            "n_chains_min_10": n_chains_min_10,
            "min_chain_len_shown": min_len,
            "rt_quality": rt_quality,
            "precursor_mass": precursor_mass_param,
        }
        try:
            excel_bytes = _build_excel(out_dir, sidecar)
            excel_b64 = base64.b64encode(excel_bytes).decode("utf-8")
        except Exception:
            excel_b64 = ""

        # ── 13. Compose JSON response ──────────────────────────────────────
        resp = {
            # Upload-raw fields (needed immediately for scatter/sigmoid display)
            "session_id": "vercel",
            "filename": fname,
            "n_points": n_original,
            "n_points_stored": len(df_stored),
            "was_pre_subsampled": was_pre_subsampled,
            "mass_range": [
                round(float(df_stored["M"].min()), 2),
                round(float(df_stored["M"].max()), 2),
            ],
            "rt_range": [
                round(float(rt_vals.min()), 2),
                round(float(rt_vals.max()), 2),
            ],
            "rt_spread_minutes": rt_spread,
            "n_blocks": int(df_stored["block"].nunique()),
            "preview_rows": preview_rows,
            "scatter_points": scatter_points,
            "sigmoid_points": sigmoid_points,
            "data_type_warning": data_type_warning,
            # Pipeline fields
            "report": report_summary,
            "mod_counts": mod_counts_response,
            "top_parallel_reads_long": top_parallel,
            "sequencing_decision_summary": sequencing_decision,
            "classification_evidence": classification_ev,
            "peak_status": peak_status_data,
            "read_summary": read_summary_data,
            "top_chains_for_plot": _sanitize(top_chains),
            "n_chains_total": n_chains_total,
            "n_chains_min_10": n_chains_min_10,
            "min_chain_len_shown": min_len,
            "coverage_by_intensity": coverage_bins,
            "coverage_by_mass_range": coverage_by_mass_range,
            "empirical_fdr": empirical_fdr,
            "rt_quality": rt_quality,
            "precursor_mass_used": precursor_mass_param,
            "sigmoid_post_pipeline": _sanitize(sigmoid_post),
            "was_subsampled": was_subsampled,
            "n_original_points": n_original,
            "n_pipeline_points": len(df_pipeline),
            "reference_comparisons": ref_comparison_map or None,
            "reference_sequence_used": "".join(ref_tokens) if ref_tokens else None,
            # tRNA reference library statistics
            "tRNA_ref_library": {
                "loaded": ref_library_loaded,
                "n_theoretical_masses": n_ref_theoretical,
                "n_peaks_matched": n_ref_matched,
                "n_peaks_total": n_pipeline_total,
                "match_pct": ref_match_pct,
            },
            # Excel download (base64 encoded)
            "excel_b64": excel_b64,
        }
        return jsonify(resp)


# ── Helper functions ───────────────────────────────────────────────────────────

def _detect_data_type(df):
    n = len(df)
    reasons = []
    raw_k = (df["M"] / BLOCK_WIDTH_DA).round()
    n_over = int((raw_k > N_BLOCKS).sum())
    frac_over = n_over / n if n else 0.0
    max_mass = float(df["M"].max()) if n > 0 else 0.0

    # Strong indicator: mass range extends well beyond the hydrolysis ladder range
    if max_mass > 25_000:
        reasons.append(
            f"Maximum mass {max_mass:,.0f} Da exceeds the hydrolysis range (2,000–23,000 Da). "
            "This file likely contains intact RNA mass data — the ladder algorithm "
            "only uses the 2,000–23,000 Da region."
        )
    elif frac_over > 0.01:
        reasons.append(
            f"{n_over} point(s) ({frac_over * 100:.1f}%) have mass above the "
            f"~{N_BLOCKS * BLOCK_WIDTH_DA:.0f} Da ladder range."
        )
    if n < 600:
        reasons.append(f"Only {n} data points — hydrolysis ladders are typically 1,000+.")
    if "block" in df.columns and n > 0:
        used = int(df["block"].nunique())
        span = max(int(df["block"].max() - df["block"].min()) + 1, 1)
        if used / span < 0.5 and span > 10:
            reasons.append(
                f"Only {used}/{span} blocks in range are populated — "
                "points look clustered rather than forming a continuous ladder."
            )
    likely_intact = len(reasons) >= 2 or frac_over > 0.03 or max_mass > 30_000
    return {"likely_intact": likely_intact, "reasons": reasons}


def _build_scatter(df):
    sdf = df[["M", "Rel_I", "T", "block"]].copy()
    if len(sdf) > 10000:
        high = sdf[sdf["Rel_I"] >= 0.3]
        low_pool = sdf[sdf["Rel_I"] < 0.3]
        low = low_pool.sample(n=min(10000 - len(high), len(low_pool)), random_state=42)
        sdf = pd.concat([high, low]).sort_values("M", ignore_index=True)
    return sdf.round(4).to_dict(orient="records")


def _build_sigmoid(df):
    sdf = df[["M", "T", "Rel_I"]].copy()
    if len(sdf) > 8000:
        high = sdf[sdf["Rel_I"] >= 0.3]
        low_pool = sdf[sdf["Rel_I"] < 0.3]
        low = low_pool.sample(n=min(8000 - len(high), len(low_pool)), random_state=42)
        sdf = pd.concat([high, low])
    return sdf.round(4).to_dict(orient="records")


def _compute_step_match_null(df, n_sample=3000):
    """
    Estimate per-step false-match probability by sampling consecutive mass-pair
    differences across the dataset and checking what fraction fall within the
    decode tolerance of any residue mass.

    For a chain of length L (L+1 peaks), P(all L steps match by chance) = p^L.
    This gives an upper bound on the per-chain false-discovery rate.
    """
    masses = np.sort(df["M"].values)
    if len(masses) < 4:
        return None

    if len(masses) > n_sample:
        rng = np.random.default_rng(42)
        idx = np.sort(rng.choice(len(masses), n_sample, replace=False))
        masses = masses[idx]

    residue_arr = np.array(sorted(_RESIDUE_MASS.values()))
    max_residue = float(residue_arr[-1]) * 1.5  # upper limit for a valid residue step
    min_residue = 150.0                           # below smallest canonical base (C ≈ 305 Da)

    null_matches = 0
    null_total = 0
    n = len(masses)
    for i in range(n):
        for j in range(i + 1, n):
            diff = masses[j] - masses[i]
            if diff > max_residue:
                break
            if diff < min_residue:
                continue
            null_total += 1
            if float(np.min(np.abs(residue_arr - diff))) <= _CHAIN_TOL:
                null_matches += 1

    if null_total == 0:
        return None

    p_step = null_matches / null_total

    # P(chain of length L passes entirely by chance) = p_step^(L−1)
    fdr_by_length = {
        str(L): round(p_step ** (L - 1) * 100, 8)
        for L in [5, 8, 10, 15, 20]
    }

    return {
        "p_step_null_pct": round(p_step * 100, 3),
        "n_pairs_tested": null_total,
        "fdr_by_chain_length_pct": fdr_by_length,
        "interpretation": (
            f"Random step-match rate: {p_step * 100:.2f}%. "
            f"FDR for ≥10-nt chain: {fdr_by_length['10']:.2e}%. "
            "Chains ≥10 positions are effectively not explainable by chance mass coincidence."
        ),
    }


def _decode_sequence(masses):
    if len(masses) < 2:
        return ""
    calls = []
    for i in range(1, len(masses)):
        delta = masses[i] - masses[i - 1]
        best_nt, best_diff = None, _DECODE_TOL
        for nt, m in _RESIDUE_MASS.items():
            if abs(delta - m) < best_diff:
                best_diff = abs(delta - m)
                best_nt = nt
        calls.append(best_nt if best_nt else f"?{delta:.0f}")
    return "-".join(calls)


def _sanitize(obj):
    if isinstance(obj, (np.floating,)):
        v = float(obj)
        return None if (math.isnan(v) or math.isinf(v)) else v
    if isinstance(obj, float):
        return None if (math.isnan(obj) or math.isinf(obj)) else obj
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.bool_,)):
        return bool(obj)
    if isinstance(obj, dict):
        return {k: _sanitize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize(v) for v in obj]
    return obj


def _build_excel(out_dir_str, sidecar):
    from pathlib import Path
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out_dir = Path(out_dir_str)

    def read_csv_safe(name):
        p = out_dir / name
        try:
            return pd.read_csv(p) if p.exists() else None
        except Exception:
            return None

    annotated = read_csv_safe("annotated_data.csv")
    peak_status = read_csv_safe("Peak_Status.csv")
    read_summary = read_csv_safe("read_summary.csv")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    PRIME5_FILL = PatternFill("solid", fgColor="DBEAFE")
    PRIME3_FILL = PatternFill("solid", fgColor="EDE9FE")
    AMB_FILL = PatternFill("solid", fgColor="FEF3C7")
    CONF_FILL = PatternFill("solid", fgColor="FEE2E2")
    ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header(ws, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

    def auto_width(ws, max_w=45):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, 10), max_w
            )

    def row_fill(call):
        if not call:
            return None
        v = str(call).lower()
        if "5" in v:
            return PRIME5_FILL
        if "3" in v:
            return PRIME3_FILL
        if "conflict" in v:
            return CONF_FILL
        return AMB_FILL

    # Sheet 1: Candidate Short Reads
    ws1 = wb.create_sheet("Candidate Short Reads")
    cols1 = [
        "Read #", "5′/3′ Call", "Confidence", "Length (nt)",
        "Decoded Sequence (mass differences → nucleotides)",
        "Mass Start (Da)", "Mass End (Da)", "RT Range (min)",
        "Mean Rel_I", "Evidence (brief)", "Warnings",
    ]
    set_header(ws1, cols1)
    min_len = int(sidecar.get("min_chain_len_shown", 10))
    rows_written = 0
    if peak_status is not None and read_summary is not None:
        ps_wr = peak_status[peak_status["read_rank"].notna()].copy()
        ps_wr["read_rank"] = ps_wr["read_rank"].astype(int)
        mass_by_rank = {
            int(rk): grp["mass"].tolist()
            for rk, grp in ps_wr.groupby("read_rank")
        }
        for _, row in read_summary.iterrows():
            rk = int(row.get("read_rank", 0) or 0)
            call = str(row.get("ladder_call", ""))
            length = int(row.get("read_length", 0) or 0)
            if length < min_len:
                continue
            masses = sorted(mass_by_rank.get(rk, []))
            seq = _decode_sequence(masses)
            m_s = round(min(masses), 2) if masses else ""
            m_e = round(max(masses), 2) if masses else ""
            rt_sub = ps_wr[ps_wr["read_rank"] == rk]["rt"].dropna()
            rt_str = f"{rt_sub.min():.1f} – {rt_sub.max():.1f}" if len(rt_sub) else ""
            mean_ri = round(float(row.get("mean_rel_i", 0) or 0), 4)
            evid = str(row.get("ladder_evidence", ""))[:150]
            warn = str(row.get("ladder_warnings", ""))[:150]
            conf = str(row.get("ladder_confidence_tier", ""))
            r = rows_written + 2
            fill = row_fill(call)
            for c, val in enumerate(
                [rk, call, conf, length, seq, m_s, m_e, rt_str, mean_ri, evid, warn], 1
            ):
                cell = ws1.cell(row=r, column=c, value=val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=(c in (5, 10, 11)))
                if fill:
                    cell.fill = fill
                elif r % 2 == 0:
                    cell.fill = ALT_FILL
            rows_written += 1
    ws1.freeze_panes = "A2"
    auto_width(ws1)
    ws1.column_dimensions["E"].width = 55

    # Sheet 2: Coverage Analysis
    ws2 = wb.create_sheet("Coverage Analysis")
    set_header(ws2, ["Intensity Bin", "Total Peaks", "Matched Peaks", "Coverage (%)"])
    cov_bins = sidecar.get("coverage_by_intensity") or []
    for i, b in enumerate(cov_bins):
        r = i + 2
        clr = (
            "D1FAE5" if b["pct"] >= 80
            else "FEF3C7" if b["pct"] >= 50
            else "FFEDD5" if b["pct"] >= 20
            else "FEE2E2"
        )
        for c, val in enumerate(
            [b["label"], b["total"], b["matched"], f'{b["pct"]}%'], 1
        ):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=clr)
    ws2.freeze_panes = "A2"

    # Mass-range breakdown (matches the UI table: rows = mass bins, cols = thresholds)
    cov_mass = sidecar.get("coverage_by_mass_range") or []
    if cov_mass:
        ws2.append([])
        r_section = ws2.max_row + 1
        sec = ws2.cell(row=r_section, column=1, value="Coverage by Mass Range")
        sec.font = Font(bold=True, size=11)
        r_hdr = r_section + 1
        for ci, h in enumerate(["Mass Range", ">10% of max", ">5% of max", ">2% of max"], 1):
            cell = ws2.cell(row=r_hdr, column=ci, value=h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row_data in cov_mass:
            r = ws2.max_row + 1
            ws2.cell(row=r, column=1, value=row_data["mass_range"]).border = border
            ws2.cell(row=r, column=1).font = Font(bold=True)
            for ci, t in enumerate(row_data.get("thresholds", []), 2):
                clr = (
                    "D1FAE5" if t["pct"] >= 80
                    else "FEF3C7" if t["pct"] >= 50
                    else "FFEDD5" if t["pct"] >= 20
                    else "FEE2E2"
                )
                cell = ws2.cell(
                    row=r, column=ci,
                    value=f'{t["pct"]}% ({t["matched"]}/{t["total"]})'
                )
                cell.border = border
                cell.fill = PatternFill("solid", fgColor=clr)

    auto_width(ws2)

    # Sheet 3: Annotated Peak Table
    if annotated is not None:
        ws3 = wb.create_sheet("Annotated Peak Table")
        display_cols = [
            c for c in ["M", "I", "T", "block", "Rel_I", "peak_status"]
            if c in annotated.columns
        ]
        if peak_status is not None:
            ps_mini = (
                peak_status[["mass", "read_rank", "ladder_call"]]
                .rename(columns={"mass": "M_ps"})
                .copy()
            )
            ann2 = annotated.sort_values("M")
            ps_mini = ps_mini.sort_values("M_ps")
            ann2 = pd.merge_asof(
                ann2, ps_mini, left_on="M", right_on="M_ps",
                tolerance=0.005, direction="nearest",
            )
            for col in ["read_rank", "ladder_call"]:
                if col in ann2.columns and col not in display_cols:
                    display_cols.append(col)
            annotated = ann2
        labels = {
            "M": "Mass (Da)", "I": "Sum Intensity", "T": "Apex RT (min)",
            "block": "Block", "Rel_I": "Rel_I", "peak_status": "Status",
            "read_rank": "Read #", "ladder_call": "Call",
        }
        set_header(ws3, [labels.get(c, c) for c in display_cols])
        for ri, (_, row3) in enumerate(annotated[display_cols].head(5000).iterrows(), 2):
            fill = (
                row_fill(row3.get("ladder_call"))
                if not pd.isna(row3.get("read_rank", float("nan")))
                else None
            )
            for ci, col in enumerate(display_cols, 1):
                val = row3[col]
                if pd.isna(val):
                    val = ""
                elif col in ("M", "Rel_I"):
                    val = round(float(val), 4)
                elif col == "I":
                    val = round(float(val), 1)
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.border = border
                if fill:
                    cell.fill = fill
                elif ri % 2 == 0:
                    cell.fill = ALT_FILL
        ws3.freeze_panes = "A2"
        auto_width(ws3)

    # Sheet 4: Run Summary
    ws4 = wb.create_sheet("Run Summary")
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 55
    ws4.cell(row=1, column=1, value="RNA Ladder Sequencing Results").font = Font(
        bold=True, size=14
    )
    ws4.cell(
        row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws4.append([])

    def kv(ws, k, v):
        r = ws.max_row + 1
        c = ws.cell(row=r, column=1, value=k)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="F1F5F9")
        ws.cell(row=r, column=2, value=str(v) if v is not None else "—")

    kv(ws4, "Pipeline points processed", sidecar.get("n_pipeline_points", "—"))
    kv(ws4, "Original points (before subsampling)", sidecar.get("n_original_points", "—"))
    kv(ws4, "Auto-subsampled?", "Yes" if sidecar.get("was_subsampled") else "No")
    kv(ws4, "Total chains recovered", sidecar.get("n_chains_total", "—"))
    kv(
        ws4,
        f"Chains with ≥{min_len} ladder positions",
        sidecar.get("n_chains_min_10", "—"),
    )
    if cov_bins:
        tot = sum(b["total"] for b in cov_bins)
        mat = sum(b["matched"] for b in cov_bins)
        kv(ws4, "Overall peak coverage", f"{mat}/{tot} = {round(mat/tot*100,1)}%" if tot else "—")
    ref = sidecar.get("reference_sequence")
    kv(
        ws4,
        "Reference sequence",
        (f"Yes ({len(ref)} nt): {ref[:40]}…" if ref and len(ref) > 40 else (f"Yes: {ref}" if ref else "No")),
    )
    rtq = sidecar.get("rt_quality")
    if rtq:
        kv(ws4, "RT gradient quality", f"{rtq['grade']} (R² = {rtq['r2']:.3f}, slope = {rtq['slope']:.4f} min/log-Da, n = {rtq['n_points']} peaks)")
    pm = sidecar.get("precursor_mass")
    if pm:
        kv(ws4, "Precursor mass (closure scoring)", f"{pm:.2f} Da")

    # ── Modification profile ───────────────────────────────────────────────────
    # Decode all chains ≥ min_len and count each residue type.
    # Non-canonical entries flag candidate modifications for follow-up.
    ws4.append([])
    r_sec = ws4.max_row + 1
    ws4.cell(row=r_sec, column=1, value="Decoded Modification Profile").font = Font(bold=True, size=11)
    ws4.cell(row=r_sec, column=1).fill = PatternFill("solid", fgColor="EFF6FF")
    mod_counter: dict = {}
    if peak_status is not None and read_summary is not None:
        try:
            ps_w = peak_status[peak_status["read_rank"].notna()].copy()
            ps_w["read_rank"] = ps_w["read_rank"].astype(int)
            m_by_rk = {
                int(rk): grp["mass"].tolist()
                for rk, grp in ps_w.groupby("read_rank")
            }
            for _, row in read_summary.iterrows():
                rk = int(row.get("read_rank", 0) or 0)
                length = int(row.get("read_length", 0) or 0)
                if length < min_len:
                    continue
                masses = sorted(m for m in m_by_rk.get(rk, []) if m)
                if len(masses) < 2:
                    continue
                for nt in _decode_sequence(masses).split("-"):
                    mod_counter[nt] = mod_counter.get(nt, 0) + 1
        except Exception:
            pass

    canonical_set = {"A", "U", "G", "C"}
    total_decoded = sum(mod_counter.values())
    canonical_count = sum(v for k, v in mod_counter.items() if k in canonical_set)
    mod_count = sum(v for k, v in mod_counter.items() if k not in canonical_set and not k.startswith("?"))
    unknown_count = sum(v for k, v in mod_counter.items() if k.startswith("?"))
    kv(ws4, "Total positions decoded", total_decoded if total_decoded else "—")
    kv(ws4, "Canonical (A/U/G/C) positions", canonical_count if total_decoded else "—")
    kv(ws4, "Modified positions identified", mod_count if total_decoded else "—")
    kv(ws4, "Unresolved positions (?mass)", unknown_count if total_decoded else "—")
    mods_found = {k: v for k, v in mod_counter.items()
                  if k not in canonical_set and not k.startswith("?")}
    if mods_found:
        ws4.append([])
        ws4.cell(row=ws4.max_row, column=1, value="  Modification").font = Font(bold=True, italic=True)
        ws4.cell(row=ws4.max_row, column=2, value="Count (positions)").font = Font(bold=True, italic=True)
        for mod, count in sorted(mods_found.items(), key=lambda x: -x[1]):
            r = ws4.max_row + 1
            ws4.cell(row=r, column=1, value=f"  {mod}")
            ws4.cell(row=r, column=2, value=count)
    if unknown_count:
        # List unresolved mass deltas so researcher can investigate
        ws4.append([])
        ws4.cell(row=ws4.max_row + 1, column=1,
                 value="  Unresolved mass deltas (not in modification dictionary):").font = Font(italic=True)
        unknowns = {k: v for k, v in mod_counter.items() if k.startswith("?")}
        for delta, count in sorted(unknowns.items(), key=lambda x: -x[1]):
            r = ws4.max_row + 1
            ws4.cell(row=r, column=1, value=f"  {delta} Da")
            ws4.cell(row=r, column=2, value=f"{count} occurrence(s)")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
