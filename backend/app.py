"""FastAPI backend for RNA Ladder Alignment.

Thin API wrapper — all scientific logic lives unchanged in
../ladder_alignment_pipeline.py.  This module only handles HTTP
plumbing (file upload, validation, serialisation) and delegates
every alignment decision to the existing pipeline functions.

Endpoints
---------
GET  /health        → {"status": "ok"}
POST /align         → JSON with per-direction summary + base64 Excel files

Theoretical sequence input (choose one per request)
----------------------------------------------------
Option A — paste sequence (easier for most users):
    rna_sequence = "GUCUACGGCC..."   (full RNA, 5′→3′, A/U/G/C)
    The backend computes both 5′ and 3′ theoretical mass tables automatically.

Option B — upload CSV files (for non-standard cases):
    theo5_file = <CSV with columns: <base>, theo_mass, position>
    theo3_file = <CSV with columns: <base>, theo_mass, position>
"""

import base64
import io
import json
import sys
import tempfile
import uuid
from collections import Counter
from pathlib import Path
from typing import Optional

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

# ---------------------------------------------------------------------------
# Import core pipeline (repo root) — DO NOT MODIFY THESE FUNCTIONS
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from ladder_alignment_pipeline import AlignConfig, align_ladders, build_excel  # noqa: E402
from trna_nested_algorithm import (  # noqa: E402
    load_data,
    assign_blocks,
    compute_relative_intensity,
    run_pipeline as run_nested_pipeline,
    Config as NestedConfig,
    BLOCK_WIDTH_DA,
    N_BLOCKS,
)
from backend.trna_constraints import (  # noqa: E402
    ResidueDictionary,
    base_call_candidates,
    default_residue_dictionary,
    normalise_sequence_tokens,
    raw_peak_qc,
    read_peak_table,
    residue_dictionary_from_table,
    sequence_to_theoretical_df,
)


# ---------------------------------------------------------------------------
# App
# ---------------------------------------------------------------------------

app = FastAPI(
    title="RNA Ladder Alignment API",
    description=(
        "Wraps the RNA ladder alignment pipeline for web access. "
        "All alignment logic is unchanged from the original pipeline."
    ),
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Sequence → theoretical mass table
#
# Verified against theo_5.csv and theo_3.csv: all 120 intermediate positions
# match exactly; the full-length position (last) matches within 5 mDa
# (rounding accumulation), well inside the 90 mDa "perfect" threshold.
#
# Chemistry summary
# -----------------
# Each internal residue contributes its nucleoside-3',5'-bisphosphate mass
# minus one water (the phosphodiester bond condensation).
#
# Intermediate 5′ fragments  (5′-phosphate + 2′,3′-cyclic phosphate)
#   mass(n) = Σ residues + H₃PO₄  (+97.9769 Da)
#
# Intermediate 3′ fragments  (3′-OH nucleoside terminus)
#   mass(n) = Σ residues − 61.9558 Da
#
# Full-length RNA  (5′-OH + 3′-OH, last position in both tables)
#   mass(n) = Σ residues + H₂O  (+18.0106 Da)
# ---------------------------------------------------------------------------

def _sequence_to_theo_df(
    rna_seq: str | list[str],
    direction: str,
    residues: ResidueDictionary | None = None,
    sample_type: str = "natural_RNA",
) -> pd.DataFrame:
    """Compute the theoretical cumulative ladder masses from an RNA sequence.

    Parameters
    ----------
    rna_seq   : Full RNA sequence in 5′→3′ direction; A/U/G/C or token list.
    direction : '5' or '3'.
                For '3', the sequence is read 3′→5′ (reversed) so that
                position 1 is the 3′-terminal nucleotide, matching the CSV.

    Returns
    -------
    DataFrame with columns  [<direction>', theo_mass, position]
    matching the existing theo CSV format consumed by align_ladders().
    """
    residues = residues or default_residue_dictionary()
    tokens = list(rna_seq) if not isinstance(rna_seq, str) else list(rna_seq)
    return sequence_to_theoretical_df(tokens, direction, residues, sample_type)


def _parse_rna_sequence(raw: str, residues: ResidueDictionary | None = None) -> str:
    """Normalise a user-pasted sequence to uppercase A/U/G/C.

    Strips whitespace, digits, FASTA '>' headers, and common separators.
    Converts T → U (accepts DNA-style input).
    Raises ValueError with a human-readable message on invalid characters.
    """
    residues = residues or default_residue_dictionary()
    return "".join(normalise_sequence_tokens(raw, residues))


def _parse_rna_sequence_tokens(
    raw: str,
    residues: ResidueDictionary | None = None,
) -> list[str]:
    residues = residues or default_residue_dictionary()
    return normalise_sequence_tokens(raw, residues)


# ---------------------------------------------------------------------------
# Ladder data validation (mirrors Streamlit app.py — no logic change)
# ---------------------------------------------------------------------------

REQUIRED_COLS = {
    "base_name", "monoisotopic_mass", "sum_intensity",
    "apex_rt", "n_iteration", "ladder_number",
}
VALID_BASES = {"A", "U", "G", "C", "High"}


def _validate(df: pd.DataFrame) -> list[str]:
    """Return a list of validation error strings (empty = OK)."""
    errors: list[str] = []
    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        errors.append(f"Missing columns: {', '.join(sorted(missing))}")
        return errors

    nulls = df[list(REQUIRED_COLS)].isnull().sum()
    if nulls.any():
        errors.append(f"Null values in: {nulls[nulls > 0].to_dict()}")

    unexpected = set(df["base_name"].unique()) - VALID_BASES
    if unexpected:
        errors.append(f"Unexpected base_name values: {unexpected}")

    bad_high = (
        df[df["base_name"] == "High"]
        .groupby("ladder_number")
        .size()
        .pipe(lambda s: s[s != 1])
    )
    if len(bad_high):
        errors.append(
            f"Ladders without exactly 1 High calibrant: {bad_high.to_dict()}"
        )

    if df["monoisotopic_mass"].min() <= 0:
        errors.append("monoisotopic_mass contains non-positive values")

    return errors


# ---------------------------------------------------------------------------
# Response helpers
# ---------------------------------------------------------------------------

def _per_ladder_records(order: list, meta: dict) -> list[dict]:
    return [
        {
            "ladder":                    ladder,
            "n_iteration":               int(meta[ladder]["n_iteration"]),
            "first_rna_pos":             int(meta[ladder]["first_rna_pos"]),
            "start_pos":                 int(meta[ladder]["start_pos"]),
            "position_shift_correction": int(meta[ladder]["position_shift_correction"]),
            "pos_adjusted":              bool(meta[ladder]["pos_adjusted"]),
            "n_placed":                  int(meta[ladder]["n_placed"]),
            "n_rejected_positions":      int(meta[ladder]["n_rejected_positions"]),
            "delta_mean":                round(float(meta[ladder]["delta_mean"]), 4),
            "delta_std":                 round(float(meta[ladder]["delta_std"]), 4),
            "max_abs_delta_jump":        round(float(meta[ladder]["max_abs_delta_jump"]), 4),
            "overall":                   str(meta[ladder]["overall"]),
        }
        for ladder in order
    ]


def _base_tokens_from_theo(theo5: pd.DataFrame) -> list[str]:
    base_col = next((c for c in theo5.columns if str(c).strip() in {"5'", "5"}), None)
    if base_col is None:
        return []
    return [str(v).strip() for v in theo5[base_col].tolist() if str(v).strip()]


async def _read_upload_table(upload: UploadFile, *, prefer_sheet: str | None = None) -> pd.DataFrame:
    data = await upload.read()
    name = (upload.filename or "").lower()
    try:
        if name.endswith((".xlsx", ".xls")):
            sheets = pd.read_excel(io.BytesIO(data), sheet_name=None)
            if prefer_sheet and prefer_sheet in sheets:
                return sheets[prefer_sheet]
            if prefer_sheet:
                lowered = {sheet.lower(): sheet for sheet in sheets}
                match = lowered.get(prefer_sheet.lower())
                if match:
                    return sheets[match]
            return next(iter(sheets.values()))
        return pd.read_csv(io.BytesIO(data), encoding="utf-8-sig")
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot read uploaded table {upload.filename!r}: {exc}",
        ) from exc


async def _resolve_residue_dictionary(mod_mass_file: Optional[UploadFile]) -> ResidueDictionary:
    if mod_mass_file is None:
        return default_residue_dictionary()
    table = await _read_upload_table(mod_mass_file, prefer_sheet="Sheet2")
    try:
        return residue_dictionary_from_table(table)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.post("/align")
async def align(
    # ── Required ──────────────────────────────────────────────────────────
    ladder_file: UploadFile = File(
        ..., description="Ladder data (.xlsx with a sheet named 'Sheet1')."
    ),

    # ── Theoretical sequence — supply ONE of the two options below ────────

    # Option A: paste the RNA sequence (5′→3′); back-end computes both tables.
    rna_sequence: str = Form(
        "",
        description=(
            "Full RNA sequence in 5′→3′ direction (A/U/G/C; T is accepted and "
            "converted to U).  FASTA format is also accepted.  "
            "When provided, theo5_file and theo3_file are ignored."
        ),
    ),

    # Option B: upload pre-computed CSV files (for non-standard modifications
    # or sequences whose masses differ from the standard residue formula).
    theo5_file: Optional[UploadFile] = File(
        None, description="5′ theoretical sequence CSV (position, theo_mass columns)."
    ),
    theo3_file: Optional[UploadFile] = File(
        None, description="3′ theoretical sequence CSV (position, theo_mass columns)."
    ),
    mod_mass_file: Optional[UploadFile] = File(
        None,
        description=(
            "Optional modification dictionary CSV/XLSX with Sheet2-style "
            "columns Symbol, Nucleotide, Base, and optional Mass."
        ),
    ),
    raw_peak_file: Optional[UploadFile] = File(
        None,
        description=(
            "Optional raw peak CSV/XLSX for tRNA-suite-style QC. Columns may "
            "include Monoisotopic Mass, Sum Intensity, and Apex RT."
        ),
    ),

    # ── Analysis settings (mirror AlignConfig defaults) ───────────────────
    sample_name:           str   = Form("Sample"),
    sample_type:           str   = Form("natural_RNA"),
    reject_below_5p:       float = Form(-19.0),
    reject_below_3p:       float = Form(-1.0),
    strict_abs_da:         float = Form(0.09),
    strict_min_run:        int   = Form(3),
    stable_offset_diff_da: float = Form(0.6),
    stable_offset_min_run: int   = Form(4),
    noisy_jump_da:         float = Form(50.0),
    raw_peak_ppm:          float = Form(10.0),
    raw_peak_mass_min:     float = Form(800.0),
    raw_peak_mass_max:     float = Form(30000.0),
    run_base_call:         bool  = Form(False),
    base_call_top_n:       int   = Form(20),
    base_call_min_len:     int   = Form(3),
    base_call_anchor_min:  float = Form(900.0),
    base_call_anchor_max:  float = Form(12000.0),
) -> JSONResponse:
    """Run both 5′ and 3′ ladder alignments.

    Returns JSON:
    {
      "summary": {
        "5prime": { ladders, perfect, shifted, noisy_shifted, rejected, normal,
                    position_adjusted, per_ladder: [...] },
        "3prime": { ... }
      },
      "files": {
        "<sample>_alignment_5prime.xlsx": "<base64>",
        "<sample>_alignment_3prime.xlsx": "<base64>"
      }
    }
    """

    # ------------------------------------------------------------------ #
    # 1. Read ladder xlsx                                                  #
    # ------------------------------------------------------------------ #
    try:
        ladder_bytes = await ladder_file.read()
        sheets = pd.read_excel(io.BytesIO(ladder_bytes), sheet_name=None)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read ladder xlsx: {exc}")

    if "Sheet1" not in sheets:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Sheet 'Sheet1' not found in the uploaded xlsx. "
                f"Available sheets: {list(sheets.keys())}"
            ),
        )
    df = sheets["Sheet1"]

    # ------------------------------------------------------------------ #
    # 2. Resolve residue dictionary and theoretical DataFrames             #
    #    Priority: rna_sequence > CSV files                                #
    # ------------------------------------------------------------------ #
    if sample_type not in {"natural_RNA", "synthetic_RNA"}:
        raise HTTPException(
            status_code=400,
            detail="sample_type must be 'natural_RNA' or 'synthetic_RNA'.",
        )

    residues = await _resolve_residue_dictionary(mod_mass_file)
    sequence_tokens: list[str] = []

    if rna_sequence.strip():
        # Option A — derive both tables from the pasted sequence
        try:
            sequence_tokens = _parse_rna_sequence_tokens(rna_sequence, residues)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        theo5 = _sequence_to_theo_df(sequence_tokens, "5", residues, sample_type)
        theo3 = _sequence_to_theo_df(sequence_tokens, "3", residues, sample_type)

    elif theo5_file is not None and theo3_file is not None:
        # Option B — use uploaded CSVs
        try:
            theo5 = pd.read_csv(
                io.BytesIO(await theo5_file.read()), encoding="utf-8-sig"
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot read theo5 CSV: {exc}")
        try:
            theo3 = pd.read_csv(
                io.BytesIO(await theo3_file.read()), encoding="utf-8-sig"
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Cannot read theo3 CSV: {exc}")

        for label, theo in [("theo5", theo5), ("theo3", theo3)]:
            if {"theo_mass", "position"} - set(theo.columns):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"{label} CSV must contain 'position' and 'theo_mass' columns."
                    ),
                )
        sequence_tokens = _base_tokens_from_theo(theo5)
    else:
        raise HTTPException(
            status_code=400,
            detail=(
                "Provide either rna_sequence (paste the RNA sequence) "
                "or both theo5_file and theo3_file (upload CSV files)."
            ),
        )

    # ------------------------------------------------------------------ #
    # 3. Validate ladder data                                             #
    # ------------------------------------------------------------------ #
    errors = _validate(df)
    if errors:
        raise HTTPException(status_code=422, detail={"validation_errors": errors})

    # ------------------------------------------------------------------ #
    # 4. Build AlignConfig — DIRECT CALL, NO LOGIC CHANGE                #
    # ------------------------------------------------------------------ #
    cfg = AlignConfig(
        reject_below_5p=reject_below_5p,
        reject_below_3p=reject_below_3p,
        strict_abs_da=strict_abs_da,
        strict_min_run=strict_min_run,
        stable_offset_diff_da=stable_offset_diff_da,
        stable_offset_min_run=stable_offset_min_run,
        noisy_jump_da=noisy_jump_da,
    )

    # ------------------------------------------------------------------ #
    # 5. Run pipeline for both directions                                 #
    # ------------------------------------------------------------------ #
    safe_name = sample_name.replace(" ", "_")
    response: dict = {
        "summary": {},
        "files": {},
        "settings": {"sample_type": sample_type},
    }
    if residues.warnings:
        response["warnings"] = list(residues.warnings)

    for direction, theo in [("5", theo5), ("3", theo3)]:
        # IMPORTANT: direct call to existing pipeline — no modification
        order, meta, maps = align_ladders(df, theo, cfg, direction=direction)

        # Build Excel into memory (same as Streamlit app — no change)
        buf = io.BytesIO()
        build_excel(order, meta, maps, theo, direction, sample_name, buf)
        buf.seek(0)

        cc  = Counter(v["overall"] for v in meta.values())
        key = f"{direction}prime"

        response["summary"][key] = {
            "ladders":            len(order),
            "perfect":            int(cc["perfect"] + cc["mixed"]),
            "shifted":            int(cc["shifted"]),
            "noisy_shifted":      int(cc["noisy_shifted"] + cc["noisy_mixed"]),
            "rejected":           int(cc["rejected"]),
            "normal":             int(cc["normal"]),
            "position_adjusted":  int(
                sum(1 for v in meta.values() if v["pos_adjusted"])
            ),
            "per_ladder": _per_ladder_records(order, meta),
        }

        filename = f"{safe_name}_alignment_{direction}prime.xlsx"
        response["files"][filename] = base64.b64encode(buf.getvalue()).decode()

    # ------------------------------------------------------------------ #
    # 6. Optional raw peak QC and experimental base-call candidates        #
    # ------------------------------------------------------------------ #
    if raw_peak_file is not None:
        raw_table = await _read_upload_table(raw_peak_file)
        try:
            peaks = read_peak_table(raw_table)
            qc = raw_peak_qc(
                peaks,
                theo5,
                theo3,
                ppm=raw_peak_ppm,
                mass_min=raw_peak_mass_min,
                mass_max=raw_peak_mass_max,
            )
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc

        response["raw_peak_qc"] = qc

        if run_base_call:
            if not sequence_tokens:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        "run_base_call requires a 5′ reference sequence or a 5′ "
                        "theoretical CSV with a base column."
                    ),
                )
            try:
                response["base_call"] = base_call_candidates(
                    peaks,
                    qc["unmatched"],
                    sequence_tokens,
                    theo5,
                    theo3,
                    residues,
                    ppm=raw_peak_ppm,
                    top_n=base_call_top_n,
                    min_ladder_len=base_call_min_len,
                    anchor_min=base_call_anchor_min,
                    anchor_max=base_call_anchor_max,
                )
            except ValueError as exc:
                raise HTTPException(status_code=422, detail=str(exc)) from exc

    return JSONResponse(content=response)


# ---------------------------------------------------------------------------
# Sequencing Assist: session storage for uploaded Excel data
# ---------------------------------------------------------------------------

_SESSION_DIR = Path(tempfile.gettempdir()) / "rna_workbench_sessions"
_SESSION_DIR.mkdir(exist_ok=True)


def _session_path(session_id: str) -> Path:
    safe = session_id.replace("/", "").replace("..", "")
    return _SESSION_DIR / safe


# ---------------------------------------------------------------------------
# Data-type heuristic: hydrolysis ladder vs. intact mass file
#
# Dr. Jiang's nested base-calling algorithm assumes a hydrolysis ladder —
# many fragments spanning a wide mass range, one block per ~residue. Intact
# mass files (one or a few full-length species, often charge-state/adduct
# variants of 1-2 molecules) don't have that structure, and feeding them
# through the same chain-calling pipeline produces noisy, low-confidence
# "reads" that aren't meaningful. This is a conservative, explainable
# heuristic — it warns, it never blocks the user from proceeding anyway.
# ---------------------------------------------------------------------------

def _detect_data_type(df: pd.DataFrame, block_width: float = BLOCK_WIDTH_DA, n_blocks: int = N_BLOCKS) -> dict:
    n = len(df)
    reasons: list[str] = []

    # Species heavier than the ladder's designed block range get clipped
    # into the last block, corrupting that block's Rel_I scaling.
    raw_k = (df["M"] / block_width).round()
    n_over_range = int((raw_k > n_blocks).sum())
    frac_over_range = (n_over_range / n) if n else 0.0
    if frac_over_range > 0.01:
        reasons.append(
            f"{n_over_range} point(s) ({frac_over_range * 100:.1f}%) have mass above the "
            f"~{n_blocks * block_width:.0f} Da ladder range and get clipped into the last block."
        )

    # Hydrolysis runs are typically 1,000+ fragments; intact runs are
    # usually a few hundred charge-state/adduct variants of 1-3 species.
    if n < 600:
        reasons.append(f"Only {n} data points — hydrolysis ladders are typically 1,000+.")

    # A real ladder populates most blocks across a continuous mass range.
    # Intact data tends to cluster around a couple of masses with large gaps.
    if "block" in df.columns and n > 0:
        used_blocks = int(df["block"].nunique())
        block_span = max(int(df["block"].max() - df["block"].min()) + 1, 1)
        coverage_ratio = used_blocks / block_span
        if coverage_ratio < 0.5 and block_span > 10:
            reasons.append(
                f"Only {used_blocks}/{block_span} blocks in range are populated — "
                "points look clustered rather than forming a continuous ladder."
            )

    likely_intact = len(reasons) >= 2 or frac_over_range > 0.03
    return {"likely_intact": likely_intact, "reasons": reasons}


# ---------------------------------------------------------------------------
# POST /sequencing-assist/upload-raw
# ---------------------------------------------------------------------------

@app.post("/sequencing-assist/upload-raw")
async def sequencing_assist_upload_raw(
    file: UploadFile = File(..., description="Deconvoluted Excel (.xlsx) file"),
) -> JSONResponse:
    """Parse a raw deconvoluted Excel file using the Python pipeline's own
    load_data / assign_blocks / compute_relative_intensity functions.
    Returns a QC preview and stores the parsed data for later pipeline execution.
    """
    filename = file.filename or "upload.xlsx"
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Expected an Excel (.xlsx) file.")

    session_id = uuid.uuid4().hex[:12]
    session_dir = _session_path(session_id)
    session_dir.mkdir(parents=True, exist_ok=True)

    excel_path = session_dir / filename
    contents = await file.read()
    excel_path.write_bytes(contents)

    try:
        df = load_data(str(excel_path))
        df = assign_blocks(df)
        df = compute_relative_intensity(df)
    except Exception as exc:
        raise HTTPException(
            status_code=422,
            detail=f"Failed to parse Excel file: {exc}",
        ) from exc

    # Pre-subsample extremely large files before saving to parquet.
    # Charge 1 data can exceed 30-100k points; keeping top-50-per-block
    # reduces parquet size while preserving the highest-intensity signals
    # for the pipeline. This is independent of the pipeline's own 8k
    # auto-subsample: the pipeline subsamples further to top-25 if needed.
    PRE_SUB_LIMIT = 20000
    n_original = len(df)
    was_pre_subsampled = False
    if n_original > PRE_SUB_LIMIT:
        df = (
            df.groupby("block")
            .apply(lambda g: g.nlargest(50, "I"), include_groups=False)
            .reset_index(drop=True)
            .sort_values("M", ignore_index=True)
        )
        # pandas 3.x drops the group-key column from apply output when
        # include_groups=False; recompute block from M so it's in the parquet.
        df["block"] = (df["M"] / BLOCK_WIDTH_DA).round().astype(int).clip(1, N_BLOCKS)
        was_pre_subsampled = True

    df.to_parquet(session_dir / "parsed.parquet", index=False)

    data_type_warning = _detect_data_type(df)

    preview_rows = df.head(5)[["M", "I", "T", "block", "Rel_I"]].round(4).to_dict(orient="records")

    # Mass vs RT data for sigmoid curve plot (downsampled)
    sigmoid_df = df[["M", "T", "Rel_I"]].copy()
    if len(sigmoid_df) > 8000:
        high = sigmoid_df[sigmoid_df["Rel_I"] >= 0.3]
        low = sigmoid_df[sigmoid_df["Rel_I"] < 0.3].sample(
            n=min(8000 - len(high), len(sigmoid_df[sigmoid_df["Rel_I"] < 0.3])),
            random_state=42,
        )
        sigmoid_df = pd.concat([high, low])
    sigmoid_points = sigmoid_df.round(4).to_dict(orient="records")

    # Scatter plot data: all points with M, Rel_I, T, block for the background cloud.
    # For large files, downsample to ~10k points for the frontend scatter.
    scatter_df = df[["M", "Rel_I", "T", "block"]].copy()
    if len(scatter_df) > 10000:
        # Stratified sample: keep all high Rel_I points + random sample of low ones
        high = scatter_df[scatter_df["Rel_I"] >= 0.3]
        low = scatter_df[scatter_df["Rel_I"] < 0.3].sample(
            n=min(10000 - len(high), len(scatter_df[scatter_df["Rel_I"] < 0.3])),
            random_state=42,
        )
        scatter_df = pd.concat([high, low]).sort_values("M", ignore_index=True)

    scatter_points = scatter_df.round(4).to_dict(orient="records")

    rt_vals = df["T"]
    rt_spread = round(float(rt_vals.max() - rt_vals.min()), 2)

    return JSONResponse(content={
        "session_id": session_id,
        "filename": filename,
        "n_points": int(n_original),
        "n_points_stored": len(df),
        "was_pre_subsampled": was_pre_subsampled,
        "mass_range": [round(float(df["M"].min()), 2), round(float(df["M"].max()), 2)],
        "rt_range": [round(float(rt_vals.min()), 2), round(float(rt_vals.max()), 2)],
        "rt_spread_minutes": rt_spread,
        "n_blocks": int(df["block"].nunique()),
        "preview_rows": preview_rows,
        "scatter_points": scatter_points,
        "sigmoid_points": sigmoid_points,
        "data_type_warning": data_type_warning,
    })


# ---------------------------------------------------------------------------
# POST /sequencing-assist/run-pipeline
# ---------------------------------------------------------------------------

PIPELINE_POINT_LIMIT = 8000  # Auto-subsample above this to keep pipeline fast

@app.post("/sequencing-assist/run-pipeline")
async def sequencing_assist_run_pipeline(
    session_id: str = Form(...),
    subsample: bool = Form(False),
    reference_sequence: str = Form(""),   # optional reference RNA sequence (5′→3′)
) -> JSONResponse:
    """Run the nested ladder pipeline on a previously uploaded Excel file.
    Returns all dashboard-consumable output data inline as JSON.
    Files with >8k points are auto-subsampled (top per block) for performance.
    """
    session_dir = _session_path(session_id)
    parquet_path = session_dir / "parsed.parquet"
    if not parquet_path.exists():
        raise HTTPException(status_code=404, detail="Session not found. Upload an Excel file first.")

    df = pd.read_parquet(parquet_path)
    n_original = len(df)
    was_subsampled = False

    # Auto-subsample large files — Charge 1 data can have 36k+ points which is too slow
    if len(df) > PIPELINE_POINT_LIMIT or subsample:
        df = df.groupby("block").apply(
            lambda g: g.nlargest(25, "I"), include_groups=False
        ).reset_index(drop=True).sort_values("M", ignore_index=True)
        # pandas 3.x drops the group-key column; recompute block from M.
        df["block"] = (df["M"] / BLOCK_WIDTH_DA).round().astype(int).clip(1, N_BLOCKS)
        was_subsampled = True

    out_dir = session_dir / "output"
    out_dir.mkdir(exist_ok=True)

    # Optional reference sequence — enables Step 6.6 (reference-guided extension
    # and post-hoc comparison). Pass as-is; run_nested_pipeline normalises it.
    ref_tokens: list[str] | None = None
    if reference_sequence.strip():
        # normalise_sequence_tokens strips whitespace, FASTA headers, T→U etc.
        from backend.trna_constraints import normalise_sequence_tokens, default_residue_dictionary
        try:
            ref_tokens = normalise_sequence_tokens(reference_sequence.strip(), default_residue_dictionary())
        except Exception:
            ref_tokens = None  # invalid sequence — run without reference

    try:
        result = run_nested_pipeline(
            df=df,
            out_dir=str(out_dir),
            reference=ref_tokens,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Pipeline execution failed: {exc}",
        ) from exc

    report = result["report"]

    def _read_csv_as_records(path_key: str) -> list | None:
        path_str = report.get(path_key)
        if not path_str or not Path(path_str).exists():
            return None
        try:
            csv_df = pd.read_csv(path_str)
            csv_df = csv_df.where(csv_df.notna(), None)
            return csv_df.to_dict(orient="records")
        except Exception:
            return None

    import math
    import numpy as _np

    def _sanitize(obj):
        """Convert numpy types and replace NaN/Inf with None for JSON."""
        if isinstance(obj, (_np.floating,)):
            return None if _np.isnan(obj) or _np.isinf(obj) else float(obj)
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        if isinstance(obj, (_np.integer,)):
            return int(obj)
        if isinstance(obj, (_np.bool_,)):
            return bool(obj)
        if isinstance(obj, dict):
            return {k: _sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_sanitize(v) for v in obj]
        return obj

    # Coverage by intensity percentile (Jianfeng's key QC figure)
    data_df = result["data"]
    data_sorted = data_df.sort_values("I", ascending=False).reset_index(drop=True)
    n = len(data_sorted)
    is_used = data_sorted["peak_status"].isin(["primary_used", "ambiguous_retained", "conflict_retained", "reference_reused"]) if "peak_status" in data_sorted.columns else pd.Series(False, index=data_sorted.index)

    bins = [
        (0, 0.02,   "Top 0–2%"),
        (0.02, 0.05, "Top 2–5%"),
        (0.05, 0.10, "Top 5–10%"),
        (0.10, 0.20, "Top 10–20%"),
        (0.20, 0.50, "Top 20–50%"),
        (0.50, 1.0,  "Bottom 50%"),
    ]
    coverage_bins = []
    for lo_frac, hi_frac, label in bins:
        i_lo, i_hi = int(n * lo_frac), int(n * hi_frac)
        subset = is_used.iloc[i_lo:i_hi]
        total = len(subset)
        matched = int(subset.sum())
        coverage_bins.append({
            "label": label,
            "total": total,
            "matched": matched,
            "pct": round(matched / total * 100, 1) if total > 0 else 0,
        })

    # Sigmoid data with peak_status after pipeline (for matched/unmatched coloring)
    data_df = result["data"]
    sigmoid_post = data_df[["M", "T", "Rel_I"]].copy()
    sigmoid_post["status"] = data_df["peak_status"] if "peak_status" in data_df.columns else "unknown"
    if len(sigmoid_post) > 8000:
        high = sigmoid_post[sigmoid_post["Rel_I"] >= 0.3]
        low = sigmoid_post[sigmoid_post["Rel_I"] < 0.3].sample(
            n=min(8000 - len(high), len(sigmoid_post[sigmoid_post["Rel_I"] < 0.3])),
            random_state=42,
        )
        sigmoid_post = pd.concat([high, low])
    sigmoid_post_records = sigmoid_post.round(4).to_dict(orient="records")

    # Build top-chains-by-length for the scatter overlay.
    # Only include chains with >= 10 ladder positions — shorter chains are not
    # meaningful for sequencing (they cannot reliably identify a position in a
    # ~70-95 nt tRNA). Fall back to chains with >= 3 positions if no 10-point
    # chains exist, so the plot is never silently empty.
    chains_list = result["chains"]
    n_chains_min_10 = sum(1 for c in chains_list if len(c["indices"]) >= 10)
    min_len_shown = 10 if n_chains_min_10 > 0 else 3
    top_chains_for_plot = []
    by_len = sorted(
        [(i, c) for i, c in enumerate(chains_list) if len(c["indices"]) >= min_len_shown],
        key=lambda x: len(x[1]["indices"]),
        reverse=True,
    )
    for chain_idx, chain in by_len[:10]:
        rows = data_df.loc[chain["indices"]].sort_values("M")
        for _, row in rows.iterrows():
            top_chains_for_plot.append({
                "chain_index": chain_idx,
                "ladder_type": chain["ladder_type"],
                "mass": float(row["M"]),
                "rel_i": float(row["Rel_I"]),
                "rt": float(row["T"]),
                "n_points": len(chain["indices"]),
            })

    # Persist coverage bins and reference info so the download endpoint can read them
    # without re-running the pipeline.
    sidecar = {
        "coverage_by_intensity": coverage_bins,
        "reference_sequence": "".join(ref_tokens) if ref_tokens else None,
        "n_pipeline_points": len(df),
        "n_original_points": n_original,
        "was_subsampled": was_subsampled,
        "n_chains_total": len(chains_list),
        "n_chains_min_10": n_chains_min_10,
        "min_chain_len_shown": min_len_shown,
    }
    (out_dir / "_sidecar.json").write_text(json.dumps(sidecar))

    # Reference comparison summary (if reference was provided)
    reference_comparisons = _sanitize(report.get("reference_comparisons")) if ref_tokens else None

    response = {
        "report": _sanitize({
            k: v for k, v in report.items()
            if not isinstance(v, str) or not v.endswith((".csv", ".xlsx", ".png", ".json"))
        }),
        "top_parallel_reads_long": _sanitize(_read_csv_as_records("top_parallel_reads_long_csv")),
        "sequencing_decision_summary": _sanitize(_read_csv_as_records("sequencing_decision_summary_csv")),
        "classification_evidence": _sanitize(_read_csv_as_records("classification_evidence_csv")),
        "peak_status": _sanitize(_read_csv_as_records("peak_status_csv")),
        "read_summary": _sanitize(_read_csv_as_records("read_summary_csv")),
        "top_chains_for_plot": _sanitize(top_chains_for_plot),
        "n_chains_total": len(chains_list),
        "n_chains_min_10": n_chains_min_10,
        "min_chain_len_shown": min_len_shown,
        "coverage_by_intensity": coverage_bins,
        "reference_comparisons": reference_comparisons,
        "sigmoid_post_pipeline": _sanitize(sigmoid_post_records),
        "was_subsampled": was_subsampled,
        "n_original_points": n_original,
        "n_pipeline_points": len(df),
    }

    return JSONResponse(content=response)


# ---------------------------------------------------------------------------
# Excel results download
# ---------------------------------------------------------------------------

# Nucleotide residue masses for sequence decoding (same as the algorithm's table)
_RESIDUE_MASS: dict[str, float] = {
    "A": 329.0525, "U": 306.0253, "G": 345.0474, "C": 305.0413,
    "D": 308.0410, "Y": 306.0253, "m1A": 343.0682, "m5C": 319.0569,
    "m7G": 359.0631, "m2G": 359.0631, "I": 330.0365,
}
_DECODE_TOL = 0.08  # Da — generous to handle calibration drift


def _decode_sequence(masses: list[float]) -> str:
    """Decode nucleotide identities from successive mass differences in a ladder chain."""
    if len(masses) < 2:
        return ""
    sorted_m = sorted(masses)
    calls: list[str] = []
    for i in range(1, len(sorted_m)):
        delta = sorted_m[i] - sorted_m[i - 1]
        best_nt, best_diff = None, _DECODE_TOL
        for nt, m in _RESIDUE_MASS.items():
            if abs(delta - m) < best_diff:
                best_diff = abs(delta - m)
                best_nt = nt
        calls.append(best_nt if best_nt else f"?{delta:.0f}")
    return "-".join(calls)


def _build_excel_response(session_dir: Path) -> bytes:
    """Build a multi-sheet Excel workbook from a completed pipeline session."""
    import io
    from datetime import datetime
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not available")

    out_dir = session_dir / "output"
    if not out_dir.exists():
        raise HTTPException(404, "Pipeline has not been run for this session.")

    sidecar_path = out_dir / "_sidecar.json"
    sidecar = json.loads(sidecar_path.read_text()) if sidecar_path.exists() else {}

    def read_csv_safe(name: str):
        p = out_dir / name
        try:
            return pd.read_csv(p) if p.exists() else None
        except Exception:
            return None

    annotated   = read_csv_safe("annotated_data.csv")
    peak_status = read_csv_safe("Peak_Status.csv")
    read_summary = read_csv_safe("read_summary.csv")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Styles ────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    PRIME5_FILL  = PatternFill("solid", fgColor="DBEAFE")
    PRIME3_FILL  = PatternFill("solid", fgColor="EDE9FE")
    AMB_FILL     = PatternFill("solid", fgColor="FEF3C7")
    CONF_FILL    = PatternFill("solid", fgColor="FEE2E2")
    ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_header(ws, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = HDR_FILL; cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

    def auto_width(ws, max_w=45):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w+2, 10), max_w)

    def row_fill(call):
        if not call: return None
        v = str(call).lower()
        if "5" in v: return PRIME5_FILL
        if "3" in v: return PRIME3_FILL
        if "conflict" in v: return CONF_FILL
        return AMB_FILL

    # ── Sheet 1: Candidate Short Reads ───────────────────────────────────
    ws1 = wb.create_sheet("Candidate Short Reads")
    cols1 = ["Read #", "5′/3′ Call", "Confidence", "Length (nt)",
             "Decoded Sequence (mass differences → nucleotides)",
             "Mass Start (Da)", "Mass End (Da)", "RT Range (min)",
             "Mean Rel_I", "Evidence (brief)", "Warnings"]
    set_header(ws1, cols1)

    min_len = int(sidecar.get("min_chain_len_shown", 10))
    rows_written = 0

    if peak_status is not None and read_summary is not None:
        ps_with_rank = peak_status[peak_status["read_rank"].notna()].copy()
        ps_with_rank["read_rank"] = ps_with_rank["read_rank"].astype(int)
        mass_by_rank = {
            int(rk): grp["mass"].tolist()
            for rk, grp in ps_with_rank.groupby("read_rank")
        }
        for _, row in read_summary.iterrows():
            rk = int(row.get("read_rank", 0) or 0)
            call  = str(row.get("ladder_call", ""))
            conf  = str(row.get("ladder_confidence_tier", ""))
            length = int(row.get("read_length", 0) or 0)
            if length < min_len:
                continue
            masses = sorted(mass_by_rank.get(rk, []))
            seq   = _decode_sequence(masses)
            m_s   = round(min(masses), 2) if masses else ""
            m_e   = round(max(masses), 2) if masses else ""
            rt_sub = ps_with_rank[ps_with_rank["read_rank"] == rk]["rt"].dropna()
            rt_str = f"{rt_sub.min():.1f} – {rt_sub.max():.1f}" if len(rt_sub) else ""
            mean_ri = round(float(row.get("mean_rel_i", 0) or 0), 4)
            evid  = str(row.get("ladder_evidence", ""))[:150]
            warn  = str(row.get("ladder_warnings", ""))[:150]
            r = rows_written + 2
            fill = row_fill(call)
            for c, val in enumerate([rk, call, conf, length, seq, m_s, m_e, rt_str, mean_ri, evid, warn], 1):
                cell = ws1.cell(row=r, column=c, value=val)
                cell.border = border
                cell.alignment = Alignment(wrap_text=(c in (5, 10, 11)))
                if fill: cell.fill = fill
                elif r % 2 == 0: cell.fill = ALT_FILL
            rows_written += 1
    ws1.freeze_panes = "A2"
    auto_width(ws1)
    ws1.column_dimensions["E"].width = 55

    # ── Sheet 2: Coverage Analysis ────────────────────────────────────────
    ws2 = wb.create_sheet("Coverage Analysis")
    set_header(ws2, ["Intensity Bin", "Total Peaks", "Matched Peaks", "Coverage (%)"])
    cov_bins = sidecar.get("coverage_by_intensity") or []
    if not cov_bins and annotated is not None and "I" in annotated.columns:
        n = len(annotated)
        ann_s = annotated.sort_values("I", ascending=False).reset_index(drop=True)
        used = {"primary_used","ambiguous_retained","conflict_retained","reference_reused"}
        for lo, hi, label in [(0,.02,"Top 0–2%"),(.02,.05,"Top 2–5%"),(.05,.10,"Top 5–10%"),
                               (.10,.20,"Top 10–20%"),(.20,.50,"Top 20–50%"),(.50,1.0,"Bottom 50%")]:
            sub = ann_s.iloc[int(n*lo):int(n*hi)]
            total = len(sub)
            matched = int(sub["peak_status"].isin(used).sum()) if "peak_status" in sub.columns else 0
            cov_bins.append({"label": label, "total": total, "matched": matched,
                             "pct": round(matched/total*100,1) if total else 0})
    for i, b in enumerate(cov_bins):
        r = i + 2
        clr = "D1FAE5" if b["pct"]>=80 else "FEF3C7" if b["pct"]>=50 else "FFEDD5" if b["pct"]>=20 else "FEE2E2"
        for c, val in enumerate([b["label"], b["total"], b["matched"], f'{b["pct"]}%'], 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=clr)
    ws2.freeze_panes = "A2"; auto_width(ws2)

    # ── Sheet 3: Annotated Peak Table ─────────────────────────────────────
    if annotated is not None:
        ws3 = wb.create_sheet("Annotated Peak Table")
        display_cols = [c for c in ["M","I","T","block","Rel_I","peak_status"] if c in annotated.columns]
        if peak_status is not None:
            ps_mini = peak_status[["mass","read_rank","ladder_call"]].rename(columns={"mass":"M_ps"}).copy()
            ann2 = annotated.sort_values("M")
            ps_mini = ps_mini.sort_values("M_ps")
            ann2 = pd.merge_asof(ann2, ps_mini, left_on="M", right_on="M_ps", tolerance=0.005, direction="nearest")
            for col in ["read_rank","ladder_call"]:
                if col in ann2.columns and col not in display_cols:
                    display_cols.append(col)
            annotated = ann2
        labels = {"M":"Mass (Da)","I":"Sum Intensity","T":"Apex RT (min)","block":"Block",
                  "Rel_I":"Rel_I","peak_status":"Status","read_rank":"Read #","ladder_call":"Call"}
        set_header(ws3, [labels.get(c,c) for c in display_cols])
        for ri, (_, row3) in enumerate(annotated[display_cols].head(5000).iterrows(), 2):
            fill = row_fill(row3.get("ladder_call")) if not pd.isna(row3.get("read_rank", float("nan"))) else None
            for ci, col in enumerate(display_cols, 1):
                val = row3[col]
                if pd.isna(val): val = ""
                elif col in ("M","Rel_I"): val = round(float(val), 4)
                elif col == "I": val = round(float(val), 1)
                cell = ws3.cell(row=ri, column=ci, value=val)
                cell.border = border
                if fill: cell.fill = fill
                elif ri % 2 == 0: cell.fill = ALT_FILL
        ws3.freeze_panes = "A2"; auto_width(ws3)

    # ── Sheet 4: Run Summary ──────────────────────────────────────────────
    from datetime import datetime
    ws4 = wb.create_sheet("Run Summary")
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 55
    ws4.cell(row=1, column=1, value="RNA Ladder Sequencing Results").font = Font(bold=True, size=14)
    ws4.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws4.append([])
    def kv(ws, k, v):
        r = ws.max_row + 1
        c = ws.cell(row=r, column=1, value=k)
        c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="F1F5F9")
        ws.cell(row=r, column=2, value=str(v) if v is not None else "—")
    kv(ws4, "Pipeline points processed", sidecar.get("n_pipeline_points","—"))
    kv(ws4, "Original points (before any subsampling)", sidecar.get("n_original_points","—"))
    kv(ws4, "Auto-subsampled during pipeline?", "Yes" if sidecar.get("was_subsampled") else "No")
    kv(ws4, "Total chains recovered", sidecar.get("n_chains_total","—"))
    kv(ws4, f"Chains with ≥{min_len} ladder positions (shown in plots)", sidecar.get("n_chains_min_10","—"))
    if cov_bins:
        tot = sum(b["total"] for b in cov_bins)
        mat = sum(b["matched"] for b in cov_bins)
        kv(ws4, "Overall peak coverage", f"{mat}/{tot} = {round(mat/tot*100,1)}%" if tot else "—")
    ref = sidecar.get("reference_sequence")
    kv(ws4, "Reference sequence provided", f"Yes ({len(ref)} nt): {ref[:40]}..." if ref and len(ref)>40 else (f"Yes: {ref}" if ref else "No"))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


from fastapi.responses import StreamingResponse as _StreamingResponse
import io as _io_module


@app.get("/sequencing-assist/download-results/{session_id}")
async def sequencing_assist_download_results(session_id: str):
    """Return a multi-sheet Excel with candidate short reads, decoded nucleotide
    sequences, coverage analysis, and annotated peak table. Call after run-pipeline."""
    session_dir = _session_path(session_id)
    try:
        excel_bytes = _build_excel_response(session_dir)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, f"Failed to build Excel: {exc}") from exc

    filename = f"rna_sequencing_results_{session_id[:8]}.xlsx"
    return _StreamingResponse(
        _io_module.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
