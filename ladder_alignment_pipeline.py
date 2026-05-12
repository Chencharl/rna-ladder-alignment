"""RNA ladder-to-sequence alignment pipeline.

Layout: rows = reference positions, columns = ladders (5 sub-cols each).
Sub-columns per ladder: Base | Mass | Intensity | RT | delta-mass

Key logic notes
---------------
Reject threshold is direction-dependent:
    5' end: delta < -18 Da is rejected  (large adduct losses expected)
    3' end: delta <  -1 Da is rejected  (tighter because 3' fragments are shorter)

Start-position placement:
    The first RNA position is assigned from the first non-High fragment mass using
    standard positive half-up rounding:
        first_rna_pos = round_half_up(first_mass / 320)
    The ladder offset is then chosen so that this first RNA row lands exactly
    on first_rna_pos, even when High appears before it in the ordered rows.
    After initial placement, ladders with strongly negative average delta mass
    are shifted upward and all delta masses are recalculated:
        avg_delta < -300 Da -> move up 2 positions
        avg_delta <  -80 Da -> move up 1 position

High calibrant handling:
    High rows are kept at their original ladder positions. Their delta masses
    are displayed, but they are excluded from RNA run classification.
"""

import argparse
import sys
from dataclasses import dataclass

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@dataclass
class AlignConfig:
    reject_below_5p: float       = -19.0   # 5' reject threshold (Da); -18.01 = dehydration artifact, margin set to -19
    reject_below_3p: float       =  -1.0   # 3' reject threshold (Da)
    strict_abs_da: float         =   0.09  # |delta| < this -> perfect match
    strict_min_run: int          =   3     # min consecutive positions for perfect
    stable_offset_diff_da: float =   0.6   # max delta-to-delta change for shifted
    stable_offset_min_run: int   =   4     # min consecutive positions for shifted
    noisy_jump_da: float         =  50.0   # single jump above this -> noisy
    order_strategy: str          = 'row_order'
    # order_strategy: 'row_order' | 'mass_asc' | 'column:<name>'

    def reject_threshold(self, direction: str) -> float:
        return self.reject_below_3p if direction == '3' else self.reject_below_5p


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

BASE_HEX = {
    'A': ('C6EFCE', '275B1C'),
    'U': ('FFEB9C', '7D5A00'),
    'G': ('BDD7EE', '1B3E5D'),
    'C': ('FFC7CE', '9C0006'),
}

DELTA_COLOURS = {
    'perfect':       ('00A65A', 'FFFFFF'),
    'shifted':       ('F97316', 'FFFFFF'),
    'noisy_shifted': ('FCD34D', '92400E'),
    'rejected':      ('A0A0A0', 'FFFFFF'),
    'normal':        ('F2F2F2', '555555'),
}

LADDER_COLOURS = {
    'perfect':       ('00A65A', 'FFFFFF', 'PERFECT'),
    'mixed':         ('059669', 'FFFFFF', 'PERFECT + SHIFTED'),
    'shifted':       ('F97316', 'FFFFFF', 'SHIFTED'),
    'noisy_shifted': ('FCD34D', '92400E', 'NOISY SHIFTED'),
    'noisy_mixed':   ('D97706', 'FFFFFF', 'PERFECT + NOISY SHIFTED'),
    'rejected':      ('A0A0A0', 'FFFFFF', 'REJECTED'),
    'normal':        ('E8EDF5', '1F3864', ''),
}

SUBCOLS = [
    ('base',      'Base'),
    ('mass',      'Mass'),
    ('intensity', 'Intens.'),
    ('rt',        'RT'),
    ('delta',     'dmass'),
]
N_SUB = len(SUBCOLS)


def _fill(h):   return PatternFill('solid', fgColor=h)
def _font(color, bold=False, size=8):
    return Font(name='Arial', size=size, bold=bold, color=color)

_THIN   = Side(style='thin',   color='BFC8D9')
_MED    = Side(style='medium', color='2F5496')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CTR    = Alignment(horizontal='center', vertical='center')
_LEFT   = Alignment(horizontal='left',   vertical='center')

_HDR_FILL  = _fill('2F5496')
_HDR_FONT  = _font('FFFFFF', bold=True)
_THEO_FILL = _fill('1F4E79')
_THEO_FONT = _font('FFFFFF', bold=True)


# ---------------------------------------------------------------------------
# Delta classification
# ---------------------------------------------------------------------------

def _run_indices(flags, min_len):
    result, start = set(), None
    for i, v in enumerate(flags):
        if v:
            if start is None:
                start = i
        elif start is not None:
            if i - start >= min_len:
                result.update(range(start, i))
            start = None
    if start is not None and len(flags) - start >= min_len:
        result.update(range(start, len(flags)))
    return result


def _shifted_indices(deltas, cfg):
    n = len(deltas)
    if n < 2:
        return set()
    flags, run_start = [False] * n, None
    for i in range(1, n):
        if abs(deltas[i] - deltas[i - 1]) < cfg.stable_offset_diff_da:
            if run_start is None:
                run_start = i - 1
        else:
            if run_start is not None:
                run_len = i - run_start
                if run_len >= cfg.stable_offset_min_run:
                    for j in range(run_start, run_start + run_len):
                        flags[j] = True
            run_start = None
    if run_start is not None:
        run_len = n - run_start
        if run_len >= cfg.stable_offset_min_run:
            for j in range(run_start, run_start + run_len):
                flags[j] = True
    return {i for i, f in enumerate(flags) if f}


def classify_positions(deltas, cfg, reject_threshold):
    """Priority: perfect > shifted > rejected > normal."""
    perfect = _run_indices([abs(d) < cfg.strict_abs_da for d in deltas],
                           cfg.strict_min_run)
    shifted = _shifted_indices(deltas, cfg)
    labels = []
    for i, d in enumerate(deltas):
        if i in perfect:
            labels.append('perfect')
        elif i in shifted:
            labels.append('shifted')
        elif d < reject_threshold:
            labels.append('rejected')
        else:
            labels.append('normal')
    return labels


def _ladder_class(pos_labels, noisy):
    has_p = 'perfect' in pos_labels
    has_s = 'shifted' in pos_labels
    if has_p and has_s:  return 'noisy_mixed' if noisy else 'mixed'
    if has_p:            return 'perfect'
    if has_s:            return 'noisy_shifted' if noisy else 'shifted'
    if 'rejected' in pos_labels: return 'rejected'
    return 'normal'


# ---------------------------------------------------------------------------
# Alignment
# ---------------------------------------------------------------------------

def _round_half_up_positive(x):
    """Standard 四舍五入 for positive masses."""
    return int(np.floor(float(x) + 0.5))


def _ordered_group(group, cfg):
    s = cfg.order_strategy
    if s.startswith('column:'):
        col = s.split(':', 1)[1]
        if col in group.columns:
            return group.sort_values(col).reset_index(drop=True), f'column:{col}'
    if s == 'mass_asc':
        return group.sort_values('monoisotopic_mass').reset_index(drop=True), 'mass_asc'
    return group.reset_index(drop=True), 'row_order'


def _build_valid(ordered, high_indices, start_pos, max_pos):
    """Keep every row, including High. Do not shift positions."""
    valid = []
    for i in ordered.index:
        p = start_pos + i
        if 1 <= p <= max_pos:
            valid.append((p, ordered.loc[i]))
    return valid


def _position_shift_for_average_delta(avg_delta):
    """Return upward position correction from the initial average delta mass."""
    if avg_delta < -300:
        return -2
    if avg_delta < -80:
        return -1
    return 0


def align_ladders(df, theo, cfg=None, direction='5'):
    """Align all ladder groups to the theoretical sequence.

    Parameters
    ----------
    df        : ladder DataFrame (Sheet1 from the input xlsx)
    theo      : theoretical sequence DataFrame
    cfg       : AlignConfig (defaults used if None)
    direction : '5' or '3', controls the reject threshold
    """
    if cfg is None:
        cfg = AlignConfig()

    max_pos        = int(theo['position'].max())
    theo_lookup    = dict(zip(theo['position'].astype(int), theo['theo_mass']))
    reject_thresh  = cfg.reject_threshold(direction)
    ladder_maps    = {}
    ladder_meta    = {}

    for name, group in df.groupby('ladder_number', sort=False):
        ordered, strategy = _ordered_group(group, cfg)
        n_iter   = ordered['n_iteration'].iloc[0]
        non_high = ordered[ordered['base_name'] != 'High']
        if non_high.empty:
            raise ValueError(f"ladder {name!r} has no RNA fragment rows")
        first_rna_index = int(non_high.index[0])
        first_mass = float(non_high.iloc[0]['monoisotopic_mass'])
        first_mass_div_320 = first_mass / 320.0

        first_rna_pos = _round_half_up_positive(first_mass_div_320)
        start_pos = first_rna_pos - first_rna_index
        anchor_pos = first_rna_pos
        align_diff = 0.0

        high_indices = {
            i for i in ordered.index
            if ordered.loc[i, 'base_name'] == 'High'
        }

        initial_valid = _build_valid(ordered, high_indices, start_pos, max_pos)
        initial_delta_vals = [
            row['monoisotopic_mass'] - theo_lookup[p]
            for p, row in initial_valid
            if p in theo_lookup
        ]
        initial_delta_mean = float(np.mean(initial_delta_vals)) if initial_delta_vals else 0.0
        position_shift_correction = _position_shift_for_average_delta(initial_delta_mean)
        start_pos += position_shift_correction

        valid = _build_valid(ordered, high_indices, start_pos, max_pos)
        delta_by_pos = {
            p: row['monoisotopic_mass'] - theo_lookup[p]
            for p, row in valid
            if p in theo_lookup
        }
        rna_valid = [(p, row) for p, row in valid if row['base_name'] != 'High']
        rna_deltas = [
            row['monoisotopic_mass'] - theo_lookup[p]
            for p, row in rna_valid if p in theo_lookup
        ]
        pos_lbls = classify_positions(rna_deltas, cfg, reject_thresh)

        noisy = (len(rna_deltas) >= 2 and
                 any(abs(rna_deltas[i] - rna_deltas[i - 1]) > cfg.noisy_jump_da
                     for i in range(1, len(rna_deltas))))
        overall = _ladder_class(pos_lbls, noisy)

        # Collision check
        seen, collisions = set(), 0
        for p, _ in valid:
            collisions += (p in seen)
            seen.add(p)

        pos_map = {}
        label_by_pos = {}
        rna_valid_with_delta = [(p, row) for p, row in rna_valid if p in theo_lookup]
        for (p, row), d, lbl in zip(rna_valid_with_delta, rna_deltas, pos_lbls):
            cell_lbl = 'noisy_shifted' if (lbl == 'shifted' and noisy) else lbl
            label_by_pos[p] = (d, cell_lbl)
        for p, row in valid:
            if row['base_name'] == 'High':
                pos_map[p] = {
                    'base':        'High',
                    'mass':        row['monoisotopic_mass'],
                    'intensity':   row['sum_intensity'],
                    'rt':          row['apex_rt'],
                    'delta':       delta_by_pos.get(p),
                    'delta_label': 'normal',
                }
                continue
            d, cell_lbl = label_by_pos.get(p, (delta_by_pos.get(p), 'normal'))
            pos_map[p] = {
                'base':        row['base_name'],
                'mass':        row['monoisotopic_mass'],
                'intensity':   row['sum_intensity'],
                'rt':          row['apex_rt'],
                'delta':       d,
                'delta_label': cell_lbl,
            }

        delta_vals       = [v['delta'] for v in pos_map.values() if v['delta'] is not None]
        placed_positions = list(pos_map.keys())
        jumps = ([abs(rna_deltas[i] - rna_deltas[i - 1]) for i in range(1, len(rna_deltas))]
                 if len(rna_deltas) >= 2 else [])

        ladder_maps[name] = pos_map
        ladder_meta[name] = {
            'n_iteration':              n_iter,
            'order_strategy':           strategy,
            'first_rna_index':          first_rna_index,
            'first_mass':               first_mass,
            'first_mass_div_320':       first_mass_div_320,
            'rounded_first_mass_div_320': first_rna_pos,
            'first_rna_pos':            first_rna_pos,
            'anchor_pos':               anchor_pos,
            'start_pos':                start_pos,
            'position_shift_correction': position_shift_correction,
            'initial_delta_mean':       initial_delta_mean,
            'pos_adjusted':             (position_shift_correction != 0),
            'end_pos':                  start_pos + len(ordered) - 1,
            'n_elements':               len(ordered),
            'n_placed':                 len(valid),
            'placed_position_span':     (max(placed_positions) - min(placed_positions)
                                         if placed_positions else 0),
            'placement_collision_count': collisions,
            'align_diff':               align_diff,
            'delta_mean':               float(np.mean(delta_vals)) if delta_vals else 0.0,
            'delta_std':                float(np.std(delta_vals)) if len(delta_vals) > 1 else 0.0,
            'max_abs_delta_jump':       float(max(jumps)) if jumps else 0.0,
            'n_rejected_positions':     pos_lbls.count('rejected'),
            'reject_threshold_da':      reject_thresh,
            'noisy':                    noisy,
            'overall':                  overall,
        }

    order = sorted(ladder_meta, key=lambda x: ladder_meta[x]['n_iteration'])
    return order, ladder_meta, ladder_maps


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def _header_cell(ws, row, col, text):
    c = ws.cell(row, col, text)
    c.font = _HDR_FONT; c.fill = _HDR_FILL; c.alignment = _CTR; c.border = _BORDER


def _write_summary_sheet(wb, sheet_name, ladder_order, ladder_meta, sort_key):
    ws   = wb.create_sheet(sheet_name)
    cols = [
        'ladder', 'n_iter', 'order_strategy',
        'first_rna_i', 'first_mass', 'first_mass/320', 'first_rna_pos',
        'initial_delta_mean', 'pos_shift', 'anchor', 'start', 'adjusted?', 'end',
        'n_elem', 'n_placed', 'span', 'collisions',
        'align_diff', 'delta_mean', 'delta_std',
        'max_jump', 'n_rejected', 'reject_threshold',
        'noisy', 'classification',
    ]
    for ci, h in enumerate(cols, 1):
        _header_cell(ws, 1, ci, h)

    for ladder in sorted(ladder_order, key=sort_key):
        m  = ladder_meta[ladder]
        bg, fg, tag = LADDER_COLOURS[m['overall']]
        ws.append([
            ladder,
            m['n_iteration'],
            m['order_strategy'],
            m['first_rna_index'],
            round(m['first_mass'], 4),
            round(m['first_mass_div_320'], 4),
            m['rounded_first_mass_div_320'],
            round(m['initial_delta_mean'], 4),
            m['position_shift_correction'],
            m['anchor_pos'],
            m['start_pos'],
            'yes' if m['pos_adjusted'] else '',
            m['end_pos'],
            m['n_elements'],
            m['n_placed'],
            m['placed_position_span'],
            m['placement_collision_count'],
            round(m['align_diff'],          4),
            round(m['delta_mean'],          4),
            round(m['delta_std'],           4),
            round(m['max_abs_delta_jump'],  4),
            m['n_rejected_positions'],
            m['reject_threshold_da'],
            'yes' if m['noisy'] else '',
            tag or 'normal',
        ])
        r = ws.max_row
        for ci in range(1, len(cols) + 1):
            c = ws.cell(r, ci)
            c.border = _BORDER; c.alignment = _CTR
            c.font = Font(name='Arial', size=8)
        ws.cell(r, len(cols)).fill = _fill(bg)
        ws.cell(r, len(cols)).font = Font(name='Arial', size=8, bold=True, color=fg)
        if m['placement_collision_count'] > 0:
            ws.cell(r, 17).fill = _fill('FFC7CE')
        if m['pos_adjusted']:
            ws.cell(r, 12).fill = _fill('FFEB9C')

    for ci in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = 'A2'


def build_excel(ladder_order, ladder_meta, ladder_maps,
                theo, direction, sample_name, out_path):
    max_pos     = int(theo['position'].max())
    base_col    = theo.columns[0]
    theo_by_pos = {int(r['position']): (r[base_col], r['theo_mass'])
                   for _, r in theo.iterrows()}

    wb  = Workbook()
    ws  = wb.active
    ws.title = f"{direction}' Ladder Alignment"

    REF_COLS  = 3
    POS_COL   = 1
    BASE_COL  = 2
    TMASS_COL = 3

    def lcol(li, si):
        return REF_COLS + li * N_SUB + si + 1

    total_cols = REF_COLS + len(ladder_order) * N_SUB

    if sample_name:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        tc = ws.cell(1, 1, f"{sample_name}  -  {direction}' Ladder Alignment")
        tc.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        tc.fill = _fill('1A2C5B'); tc.alignment = _CTR
        ws.row_dimensions[1].height = 20
        hdr1 = 2
    else:
        hdr1 = 1

    hdr2 = hdr1 + 1
    row0 = hdr2 + 1

    for r in [hdr1, hdr2]:
        ws.row_dimensions[r].height = 20

    for col, lbl in [(POS_COL, 'Pos'),
                     (BASE_COL, f"Theo ({direction}')"),
                     (TMASS_COL, 'Theo Mass')]:
        ws.merge_cells(start_row=hdr1, start_column=col,
                       end_row=hdr2,   end_column=col)
        c = ws.cell(hdr1, col, lbl)
        c.font = _THEO_FONT; c.fill = _THEO_FILL; c.alignment = _CTR; c.border = _BORDER

    for li, ladder in enumerate(ladder_order):
        meta = ladder_meta[ladder]
        bg, fg, tag = LADDER_COLOURS[meta['overall']]
        col_s = lcol(li, 0)
        col_e = lcol(li, N_SUB - 1)

        ws.merge_cells(start_row=hdr1, start_column=col_s,
                       end_row=hdr1,   end_column=col_e)
        label = (
            f"{ladder} (iter {meta['n_iteration']})  "
            f"m1/320={meta['first_mass_div_320']:.2f} "
            f"-> RNA pos {meta['first_rna_pos']} (start {meta['start_pos']})"
        )
        if meta['position_shift_correction']:
            label += f"  [shift {meta['position_shift_correction']}]"
        if tag:
            label += f"  [{tag}]"
        if meta['pos_adjusted']:
            label += f"  [pos adj]"
        c = ws.cell(hdr1, col_s, label)
        c.fill = _fill(bg); c.border = _BORDER; c.alignment = _CTR
        c.font = Font(name='Arial', size=7, bold=True, color=fg)

        for si, (_, slbl) in enumerate(SUBCOLS):
            c = ws.cell(hdr2, lcol(li, si), slbl)
            c.fill = _fill(bg); c.border = _BORDER; c.alignment = _CTR
            c.font = Font(name='Arial', size=6, bold=True, color=fg)

    stripe = [_fill('F7F9FC'), _fill('EAEEF5')]

    for p in range(1, max_pos + 1):
        er = row0 + p - 1
        ws.row_dimensions[er].height = 14
        base_b, theo_mass = theo_by_pos[p]
        bf, bfg = BASE_HEX.get(base_b, ('CCCCCC', '000000'))

        c = ws.cell(er, POS_COL, p)
        c.fill = _THEO_FILL; c.border = _BORDER; c.alignment = _CTR
        c.font = Font(name='Arial', size=8, bold=True, color='FFFFFF')

        c = ws.cell(er, BASE_COL, base_b)
        c.fill = _fill(bf); c.border = _BORDER; c.alignment = _CTR
        c.font = Font(name='Arial', size=8, bold=True, color=bfg)

        c = ws.cell(er, TMASS_COL, round(float(theo_mass), 3))
        c.number_format = '0.000'
        c.fill = _THEO_FILL; c.border = _BORDER; c.alignment = _CTR
        c.font = Font(name='Arial', size=7, color='FFFFFF')

        for li, ladder in enumerate(ladder_order):
            data      = ladder_maps[ladder].get(p)
            row_stripe = stripe[p % 2]

            for si, (sub_key, _) in enumerate(SUBCOLS):
                c = ws.cell(er, lcol(li, si))
                c.border = _BORDER; c.alignment = _CTR

                if data is None:
                    c.fill = row_stripe if sub_key == 'base' else _fill('FAFAFA')
                    c.font = Font(name='Arial', size=6, color='CCCCCC')
                    continue

                base = data['base']
                dlbl = data['delta_label']
                bf_b, bfg_b = BASE_HEX.get(base, ('DDDDDD', '333333'))

                if sub_key == 'base':
                    c.value = base
                    c.fill  = _fill(bf_b)
                    c.font  = Font(name='Arial', size=8, bold=True, color=bfg_b)
                elif sub_key == 'mass':
                    c.value = round(data['mass'], 3)
                    c.number_format = '0.000'
                    c.fill = _fill('F5F5F5')
                    c.font = Font(name='Arial', size=6, color='333333')
                elif sub_key == 'intensity':
                    c.value = int(data['intensity'])
                    c.number_format = '#,##0'
                    c.fill = _fill('EBF3FF')
                    c.font = Font(name='Arial', size=6, color='003366')
                elif sub_key == 'rt':
                    c.value = round(data['rt'], 3)
                    c.number_format = '0.000'
                    c.fill = _fill('FFF9EC')
                    c.font = Font(name='Arial', size=6, color='664400')
                elif sub_key == 'delta':
                    dv = data['delta']
                    if dv is None:
                        c.value = 'HIGH'
                        c.fill = _fill('E5E7EB')
                        c.font = Font(name='Arial', size=6, bold=True, color='4B5563')
                    else:
                        c.value = round(dv, 4)
                        c.number_format = '+0.0000;-0.0000;0.0000'
                        dbg, dfg = DELTA_COLOURS[dlbl]
                        c.fill = _fill(dbg)
                        c.font = Font(name='Arial', size=6,
                                      bold=(dlbl in ('perfect', 'shifted', 'noisy_shifted')),
                                      color=dfg)

    ws.column_dimensions[get_column_letter(POS_COL)].width   = 5
    ws.column_dimensions[get_column_letter(BASE_COL)].width  = 7
    ws.column_dimensions[get_column_letter(TMASS_COL)].width = 11
    sub_widths = {'base': 4.5, 'mass': 8.5, 'intensity': 9.0, 'rt': 6.5, 'delta': 8.5}
    for li in range(len(ladder_order)):
        for si, (key, _) in enumerate(SUBCOLS):
            ws.column_dimensions[get_column_letter(lcol(li, si))].width = sub_widths[key]

    ws.freeze_panes = ws.cell(row0, lcol(0, 0))

    leg = row0 + max_pos + 1
    ws.cell(leg, POS_COL, 'Base:').font = Font(name='Arial', bold=True, size=8)
    for ci, (b, (bg, fg)) in enumerate(BASE_HEX.items(), start=REF_COLS + 1):
        c = ws.cell(leg, ci, b)
        c.fill = _fill(bg); c.border = _BORDER; c.alignment = _CTR
        c.font = Font(name='Arial', size=7, bold=True, color=fg)

    leg2 = leg + 1
    ws.cell(leg2, POS_COL, 'dmass:').font = Font(name='Arial', bold=True, size=8)
    for ci, (lbl, (bg, fg)) in enumerate(DELTA_COLOURS.items(), start=REF_COLS + 1):
        c = ws.cell(leg2, ci, lbl)
        c.fill = _fill(bg); c.border = _BORDER; c.alignment = _CTR
        c.font = Font(name='Arial', size=7, bold=True, color=fg)

    _write_summary_sheet(
        wb, 'Credibility Summary', ladder_order, ladder_meta,
        sort_key=lambda x: ladder_meta[x]['n_iteration'])
    _write_summary_sheet(
        wb, 'Summary by Class', ladder_order, ladder_meta,
        sort_key=lambda x: (['perfect', 'mixed', 'shifted', 'noisy_shifted',
                              'noisy_mixed', 'normal', 'rejected']
                             .index(ladder_meta[x]['overall'])))

    ws_t = wb.create_sheet('Theo Reference')
    for ci, h in enumerate(['Position', 'Base', 'Theo Mass (Da)'], 1):
        _header_cell(ws_t, 1, ci, h)
    for _, tr in theo.iterrows():
        p = int(tr['position']); b = tr[base_col]
        ws_t.append([p, b, round(float(tr['theo_mass']), 4)])
        r = ws_t.max_row
        for ci in range(1, 4):
            c = ws_t.cell(r, ci); c.border = _BORDER; c.alignment = _CTR
            c.font = Font(name='Arial', size=8)
        bf, bfg = BASE_HEX.get(b, ('CCCCCC', '000000'))
        ws_t.cell(r, 2).fill = _fill(bf)
        ws_t.cell(r, 2).font = Font(name='Arial', size=8, bold=True, color=bfg)
    for ci in range(1, 4):
        ws_t.column_dimensions[get_column_letter(ci)].width = 16
    ws_t.freeze_panes = 'A2'

    wb.save(out_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='RNA ladder alignment pipeline')
    parser.add_argument('--ladder',    required=True)
    parser.add_argument('--theo',      required=True)
    parser.add_argument('--output',    required=True)
    parser.add_argument('--direction', default='5', choices=['5', '3'])
    parser.add_argument('--name',      default='')
    parser.add_argument('--order',     default='row_order')
    args = parser.parse_args()

    df = pd.read_csv(args.ladder)
    required = {'base_name', 'monoisotopic_mass', 'sum_intensity',
                'apex_rt', 'n_iteration', 'ladder_number'}
    if missing := required - set(df.columns):
        sys.exit(f'missing columns: {missing}')

    theo = pd.read_csv(args.theo)
    if {'theo_mass', 'position'} - set(theo.columns):
        sys.exit("theo CSV must have 'theo_mass' and 'position'")

    cfg   = AlignConfig(order_strategy=args.order)
    order, meta, maps = align_ladders(df, theo, cfg, direction=args.direction)

    from collections import Counter
    cc = Counter(v['overall'] for v in meta.values())
    adj = sum(1 for v in meta.values() if v['pos_adjusted'])
    print(f"direction={args.direction}  ladders={len(order)}  "
          f"perfect={cc['perfect']+cc['mixed']}  shifted={cc['shifted']}  "
          f"noisy={cc['noisy_shifted']+cc['noisy_mixed']}  "
          f"rejected={cc['rejected']}  normal={cc['normal']}  "
          f"pos_adjusted={adj}")

    build_excel(order, meta, maps, theo, args.direction, args.name, args.output)
    print(f"saved: {args.output}")


if __name__ == '__main__':
    main()
